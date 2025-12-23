"""
文案助手界面组件
提供AI对白生成的用户界面
"""

from typing import Dict, List, Optional
import sys

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QComboBox, QPushButton, QLineEdit, QTextEdit,
    QListWidget, QListWidgetItem, QSplitter, QMessageBox, QFileDialog, QDialog,
    QFrame, QScrollArea, QTabWidget, QGroupBox, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal, QTimer, QDateTime, QSize
from PyQt6.QtGui import QFont, QTextCursor, QKeySequence, QShortcut, QPixmap
from PyQt6.QtWidgets import QApplication

from duoki_editor.ui.toast import Toast

import re
import pandas as pd

from duoki_editor.utils.config_manager import ConfigManager
from duoki_editor.utils.constants_loader import ConstantsLoader

from duoki_editor.core.auth_manager import AuthManager
from .coze_client import CozeAPIClient, CozeWorkerThread
from .feishu_sync import sync_ai_content_to_feishu
from duoki_editor.modules.ai_image_generator.ai_image_generator import ImageGenerationWorker, CARTOON_STYLE_SUFFIX, BATCH_CREATE_STORY_PROMPT
import os
import io
from PIL import Image
from duoki_editor.modules.ai_image_generator.ai_image_generator import add_watermark


class PlainTextEdit(QTextEdit):
    """自定义QTextEdit，只接受纯文本粘贴，避免Excel格式问题"""
    
    def insertFromMimeData(self, source):
        """重写粘贴方法，只插入纯文本"""
        if source.hasText():
            # 获取纯文本内容
            plain_text = source.text()
            # 清理可能的格式字符
            plain_text = plain_text.replace('\r\n', '\n').replace('\r', '\n')
            # 插入纯文本
            self.insertPlainText(plain_text)
        else:
            # 如果没有文本内容，调用父类方法
            super().insertFromMimeData(source)


class ContentGeneratorWidget(QWidget):
    """文案助手主界面"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.config_manager = ConfigManager()
        
        # 初始化常量加载器
        self.constants_loader = ConstantsLoader()
        
        # 初始化认证管理器
        self.auth_manager = AuthManager()
        
        # 初始化组件
        self.coze_client = None
        self.current_worker = None
        self.username_worker = None
        
        # 当前状态
        self.current_platform = "扣子"
        self.current_agent = "多奇对白生成器"
        self.current_conversation_id = ""
        self.is_chatting = False
        self._batch_generating = False
        self._batch_frames = []
        self._batch_index = 0
        self._batch_wen_running = False
        self._batch_ci_running = False
        self._batch_wen_current_frame = None
        self._batch_ci_current_frame = None
        self._batch_tu_workers: Dict = {}
        self._batch_ci_wait_timer = QTimer(self)
        self._batch_ci_wait_timer.setSingleShot(True)
        self._batch_ci_wait_timer.timeout.connect(self._on_batch_ci_wait_timeout)
        self._batch_tu_wait_timer = QTimer(self)
        self._batch_tu_wait_timer.setSingleShot(True)
        self._batch_tu_wait_timer.timeout.connect(self._on_batch_tu_wait_timeout)
        
        # 飞书同步相关
        self.current_record_id = None  # 存储当前记录的record_id
        self.tab_upload_status = {}  # 跟踪每个tab页的上报状态 {tab_index: has_uploaded}
        
        # 初始化UI
        self._init_ui()
        self._load_conversation_id()
    
    def get_user_name(self) -> str:
        """获取用户名，优先从输入框获取，如果输入框为空则从主窗口获取"""
        # 优先从用户名输入框获取
        if hasattr(self, 'username_edit'):
            username_from_input = self.username_edit.text().strip()
            if username_from_input:
                return username_from_input
        
        # 如果输入框为空，从主窗口获取
        parent = self.parent()
        while parent:
            if hasattr(parent, 'user_name'):
                return parent.user_name
            parent = parent.parent()
        return ""
    
    def update_username_display(self, username: str):
        """更新用户名显示"""
        if hasattr(self, 'username_edit'):
            self.username_edit.setText(username)
            if username:
                self.username_edit.setPlaceholderText("")
                self.username_edit.setReadOnly(True)  # 有用户名时设为只读
            else:
                self.username_edit.setPlaceholderText("请输入用户名")
                self.username_edit.setReadOnly(False)  # 没有用户名时允许编辑
                # 使用toast提示获取用户名失败
                Toast.warning("获取用户名失败，请手动输入", self)
    
    def _init_username_display(self):
        """初始化用户名显示"""
        # 检查主窗口是否已经有用户名
        username = self.get_user_name()
        if username:
            self.update_username_display(username)
        

    
    def _init_ui(self):
        """初始化用户界面"""
        # 创建主布局
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        
        # 创建tab框架
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # 创建文案助手tab
        content_generator_tab = self._create_content_generator_tab()
        self.tab_widget.addTab(content_generator_tab, "文案助手")
        npc_pipeline_tab = self._create_npc_pipeline_tab()
        self.tab_widget.addTab(npc_pipeline_tab, "角色生产线")
        
        # 初始化用户名显示
        self._init_username_display()
    
    def _create_content_generator_tab(self) -> QWidget:
        """创建文案助手tab内容"""
        tab_widget = QWidget()
        layout = QHBoxLayout(tab_widget)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # 创建左右分割器
        splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(splitter)
        
        # 左侧控制面板
        left_panel = self._create_left_panel()
        splitter.addWidget(left_panel)
        
        # 右侧对话面板
        right_panel = self._create_right_panel()
        splitter.addWidget(right_panel)
        
        # 设置分割器比例 (左侧固定宽度，右侧自适应)
        splitter.setSizes([300, 800])
        splitter.setStretchFactor(0, 0)  # 左侧不拉伸
        splitter.setStretchFactor(1, 1)  # 右侧拉伸
        
        return tab_widget

    class SquareImageContainer(QLabel):
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.setStyleSheet("background-color: #000000; color: white;")
            self._original_pixmap = None
            self._loading = False
            self._spinner_timer = QTimer(self)
            self._spinner_frames = ['⠋','⠙','⠹','⠸','⠼','⠴','⠦','⠧','⠇','⠏']
            self._spinner_index = 0
            self._spinner_message = ""
            self._spinner_timer.timeout.connect(self._on_spinner_tick)
        def _on_spinner_tick(self):
            if not self._loading:
                return
            self._spinner_index = (self._spinner_index + 1) % len(self._spinner_frames)
            frame = self._spinner_frames[self._spinner_index]
            self.setText(f"{self._spinner_message} {frame}")
        def start_loading(self, message: str = "生成中"):
            self._loading = True
            self._spinner_message = message
            self._spinner_index = 0
            self._original_pixmap = None
            self.setText(f"{message} {self._spinner_frames[self._spinner_index]}")
            self._spinner_timer.start(100)
        def stop_loading(self):
            self._loading = False
            self._spinner_timer.stop()
            self.setText("")
        def set_original_pixmap(self, pix: QPixmap):
            self._original_pixmap = pix
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                s = self.size()
                sp = self._original_pixmap.scaled(s, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                QLabel.setPixmap(self, sp)
        def resizeEvent(self, event):
            super().resizeEvent(event)
            h = self.height()
            if h > 0:
                self.setFixedWidth(h)
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                s = self.size()
                sp = self._original_pixmap.scaled(s, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                QLabel.setPixmap(self, sp)
        def mousePressEvent(self, event):
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                p = self.parent()
                while p is not None:
                    if hasattr(p, '_show_image_preview'):
                        p._show_image_preview(self._original_pixmap, '图片预览')
                        break
                    p = p.parent()
            QLabel.mousePressEvent(self, event)
        def sizeHint(self):
            return QSize(300, 300)
        def clear(self):
            self._original_pixmap = None
            self._loading = False
            self._spinner_timer.stop()
            QLabel.clear(self)

    class ThreeViewContainer(QLabel):
        def __init__(self, parent=None):
            super().__init__(parent)
            self._original_pixmap = None
            self.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.setStyleSheet("background-color: #000000; color: white;")
        def resizeEvent(self, event):
            super().resizeEvent(event)
            h = self.height()
            if h > 0:
                self.setFixedWidth(int(h * 1.5))
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                s = self.size()
                sp = self._original_pixmap.scaled(s, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                QLabel.setPixmap(self, sp)
        def set_original_pixmap(self, pix):
            self._original_pixmap = pix
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                s = self.size()
                sp = self._original_pixmap.scaled(s, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                QLabel.setPixmap(self, sp)
        def mousePressEvent(self, event):
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                p = self.parent()
                while p is not None:
                    if hasattr(p, '_show_image_preview'):
                        p._show_image_preview(self._original_pixmap, '图片预览')
                        break
                    p = p.parent()
            QLabel.mousePressEvent(self, event)

    class TemplateFrame(QWidget):
        def __init__(self, name: str, on_click, parent=None):
            super().__init__(parent)
            self._name = name
            self._on_click = on_click
            self._buttons_widget = None
            self.btn_wen = None
            self.btn_ci = None
            self.btn_tu = None
            self._selected = False
            self.setObjectName("template_frame")
            self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            self.setStyleSheet("#template_frame{background-color: transparent; border: 1px solid #ffffff; border-radius: 4px;}")
            self.lab_wen = None
            self.lab_ci = None
            self.lab_tu = None
            self._left_name_label = None
            self._done_wen = False
            self._done_ci = False
            self._done_tu = False
            self._complete = False
            self._in_progress_step = None
            self._failed_step = None
        def _apply_frame_style(self):
            bg = "#66ff66" if self._complete else "transparent"
            if self._selected:
                self.setStyleSheet(f"#template_frame{{background-color: {bg}; border: 3px solid #8888ff; border-radius: 4px;}}")
            else:
                self.setStyleSheet(f"#template_frame{{background-color: {bg}; border: 1px solid #ffffff; border-radius: 4px;}}")
        def _update_completion_style(self):
            self._complete = bool(self._done_wen and self._done_ci and self._done_tu)
            if self._complete:
                if self.lab_wen is not None:
                    self.lab_wen.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
                if self.lab_ci is not None:
                    self.lab_ci.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
                if self.lab_tu is not None:
                    self.lab_tu.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
                if self._left_name_label is not None:
                    self._left_name_label.setStyleSheet("border: none; color: black;")
            self._apply_frame_style()
        def set_buttons_widget(self, w: QWidget):
            self._buttons_widget = w
            if self._buttons_widget is not None:
                self._buttons_widget.setVisible(False)
        def set_selected(self, selected: bool):
            self._selected = bool(selected)
            self._apply_frame_style()
            if self._buttons_widget is not None:
                self._buttons_widget.setVisible(True if self._selected else False)
        def mousePressEvent(self, event):
            if callable(self._on_click):
                self._on_click(self)
            QWidget.mousePressEvent(self, event)
        def set_wen_done(self):
            if self.lab_wen is not None:
                self.lab_wen.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
            self._done_wen = True
            self._update_completion_style()
        def set_ci_done(self):
            if self.lab_ci is not None:
                self.lab_ci.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
            self._done_ci = True
            self._update_completion_style()
        def set_tu_done(self):
            if self.lab_tu is not None:
                self.lab_tu.setStyleSheet("background-color: #66ff66; color: black; border-radius: 2px;")
            self._done_tu = True
            self._update_completion_style()
        def mark_step_in_progress(self, step: str):
            self._in_progress_step = step
            if step == 'wen' and self.lab_wen is not None:
                self.lab_wen.setStyleSheet("background-color: #ffa500; color: black; border-radius: 2px;")
            elif step == 'ci' and self.lab_ci is not None:
                self.lab_ci.setStyleSheet("background-color: #ffa500; color: black; border-radius: 2px;")
            elif step == 'tu' and self.lab_tu is not None:
                self.lab_tu.setStyleSheet("background-color: #ffa500; color: black; border-radius: 2px;")
            if self._left_name_label is not None:
                self._left_name_label.setStyleSheet("border: none; color: #ffa500;")
        def mark_step_failed(self, step: str):
            self._failed_step = step
            if step == 'wen' and self.lab_wen is not None:
                self.lab_wen.setStyleSheet("background-color: #ff8888; color: black; border-radius: 2px;")
            elif step == 'ci' and self.lab_ci is not None:
                self.lab_ci.setStyleSheet("background-color: #ff8888; color: black; border-radius: 2px;")
            elif step == 'tu' and self.lab_tu is not None:
                self.lab_tu.setStyleSheet("background-color: #ff8888; color: black; border-radius: 2px;")
            if self._left_name_label is not None:
                self._left_name_label.setStyleSheet("border: none; color: #ff0000;")

    def _create_npc_pipeline_tab(self) -> QWidget:
        tab = QWidget()
        layout = QHBoxLayout(tab)
        layout.setContentsMargins(5, 5, 5, 5)
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(8, 8, 8, 8)
        left_layout.setSpacing(8)
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("角色名字："))
        self.npc_role_combo = QComboBox()
        refs = self.constants_loader.get_image_references()
        self.npc_role_combo.addItem("请选择一个角色")
        self.npc_role_combo.addItems(list(refs.keys()))
        self.npc_role_combo.setCurrentIndex(0)
        self.npc_role_combo.currentTextChanged.connect(self._on_npc_role_changed)
        try:
            self.npc_role_combo.setFixedWidth(150)
        except Exception:
            pass
        row1.addWidget(self.npc_role_combo)
        row1.addStretch(1)
        self.npc_template_count_label = QLabel("共0个模板")
        row1.addWidget(self.npc_template_count_label)
        self.npc_batch_generate_btn = QPushButton("☢️批量生成")
        self.npc_batch_generate_btn.setFixedWidth(100)
        self.npc_batch_generate_btn.clicked.connect(self.on_npc_batch_generate)
        row1.addWidget(self.npc_batch_generate_btn)
        left_layout.addLayout(row1)
        self.npc_blank_scroll = QScrollArea()
        self.npc_blank_scroll.setWidgetResizable(True)
        self.npc_blank_grid_container = QWidget()
        self.npc_blank_grid = QGridLayout(self.npc_blank_grid_container)
        self.npc_blank_grid.setContentsMargins(8, 8, 8, 8)
        self.npc_blank_grid.setSpacing(8)
        self.npc_blank_grid.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self.npc_blank_grid.setColumnStretch(0, 1)
        self.npc_blank_scroll.setWidget(self.npc_blank_grid_container)
        left_layout.addWidget(self.npc_blank_scroll)
        self.npc_keyword_text = QTextEdit()
        self.npc_keyword_text.setPlaceholderText("请填写角色的全局剧情或全局行为动机…")
        self.npc_keyword_text.setFixedHeight(80)
        left_layout.addWidget(self.npc_keyword_text)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(8, 8, 8, 8)
        right_layout.setSpacing(8)
        row_top = QHBoxLayout()
        self.npc_select_output_dir_btn = QPushButton("选择输出目录")
        self.npc_select_output_dir_btn.clicked.connect(self.on_npc_select_output_dir)
        self.npc_output_dir_display = QLineEdit()
        self.npc_output_dir_display.setReadOnly(True)
        self.npc_open_output_dir_btn = QPushButton("打开目录")
        self.npc_open_output_dir_btn.clicked.connect(self.on_npc_open_output_dir)
        self.npc_start_generate_btn = QPushButton("开始生成")
        self.npc_start_generate_btn.clicked.connect(self.on_npc_start_generate)
        row_top.addWidget(self.npc_select_output_dir_btn)
        row_top.addWidget(self.npc_output_dir_display)
        row_top.addWidget(self.npc_open_output_dir_btn)
        row_top.addStretch(1)
        row_top.addWidget(self.npc_start_generate_btn)
        right_layout.addLayout(row_top)

        bottom_row = QHBoxLayout()
        self.npc_three_view_container = ContentGeneratorWidget.ThreeViewContainer()
        bottom_row.addWidget(self.npc_three_view_container, 1)
        self.npc_prompt_text = QTextEdit()
        self.npc_prompt_text.setReadOnly(True)
        bottom_row.addWidget(self.npc_prompt_text, 1)
        self.npc_image_container = ContentGeneratorWidget.SquareImageContainer()
        bottom_row.addWidget(self.npc_image_container, 1)
        right_layout.addLayout(bottom_row, 1)

        self.dialog_table = QTableWidget(0, 2)
        self.dialog_table.setHorizontalHeaderLabels(["npc", "dialog"])
        self.dialog_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.dialog_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        right_layout.addWidget(self.dialog_table, 2)

        layout.addWidget(left, 1)
        layout.addWidget(right, 3)
        layout.setStretch(0, 1)
        layout.setStretch(1, 3)

        default_out = self.config_manager.ensure_npc_output_directory()
        self.npc_output_dir_display.setText(default_out)
        self._load_npc_blank_frames()
        self._update_npc_template_count_label()
        return tab

    class ImagePreviewDialog(QDialog):
        def __init__(self, pix: QPixmap, parent=None, title: str = ''):
            super().__init__(parent)
            if title:
                self.setWindowTitle(title)
            self._original_pixmap = pix
            v = QVBoxLayout(self)
            self._label = QLabel(self)
            self._label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            v.addWidget(self._label)
            self._update_pixmap()
        def _update_pixmap(self):
            if self._original_pixmap is not None and not self._original_pixmap.isNull():
                s = self._label.size()
                if s.width() > 0 and s.height() > 0:
                    sp = self._original_pixmap.scaled(s, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                    self._label.setPixmap(sp)
        def resizeEvent(self, event):
            super().resizeEvent(event)
            self._update_pixmap()

    def _show_image_preview(self, pix: QPixmap, title: str = '图片预览'):
        dlg = ContentGeneratorWidget.ImagePreviewDialog(pix, self, title)
        dlg.resize(800, 600)
        dlg.exec()

    def _on_npc_role_changed(self, role_name: str):
        if hasattr(self, 'dialog_table'):
            self.dialog_table.setRowCount(0)
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            self.npc_prompt_text.clear()
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.clear()
        for f in getattr(self, '_npc_template_frames', []) or []:
            f.set_selected(False)
        self._selected_template_frame = None
        val = role_name.strip() if role_name else ""
        if not val or val == "请选择一个角色":
            if hasattr(self, 'npc_three_view_container'):
                self.npc_three_view_container.clear()
            print("角色未选择")
            # 更新提示词中的npc1占位
            self._npc1_current = ''
            self._refresh_npc_prompt_with_role()
            # 清空所有模板状态为默认灰色
            for f in getattr(self, '_npc_template_frames', []) or []:
                self._reset_template_frame_status(f, reset_all=True)
            self._update_npc_template_count_label()
            return
        refs = self.constants_loader.get_image_references()
        p = refs.get(val, '')
        if not p:
            if hasattr(self, 'npc_three_view_container'):
                self.npc_three_view_container.clear()
            print(f"未找到角色参考: {val}")
            # 即使没有参考图，也更新提示词中的npc1占位
            self._npc1_current = val
            self._refresh_npc_prompt_with_role()
            # 刷新所有模板状态为所选角色对应的进度
            self._refresh_all_template_frames_status()
            return
        import os
        import duoki_editor as de
        pkg_root = os.path.dirname(os.path.abspath(de.__file__))
        if not os.path.isabs(p):
            abs_path = os.path.join(pkg_root, p)
        else:
            abs_path = p
        if not os.path.exists(abs_path):
            self.npc_three_view_container.clear()
            print(f"参考图片不存在: {abs_path}")
            # 文件不存在时仍更新提示词
            self._npc1_current = val
            self._refresh_npc_prompt_with_role()
            self._refresh_all_template_frames_status()
            return
        pix = QPixmap(abs_path)
        if pix.isNull():
            self.npc_three_view_container.clear()
            print(f"加载参考图片失败: {abs_path}")
            # 加载失败时仍更新提示词
            self._npc1_current = val
            self._refresh_npc_prompt_with_role()
            self._refresh_all_template_frames_status()
            return
        self.npc_three_view_container.set_original_pixmap(pix)
        # 成功加载参考图后，同步更新提示词中的npc1占位
        self._npc1_current = val
        self._refresh_npc_prompt_with_role()
        self._refresh_all_template_frames_status()
        if getattr(self, '_selected_template_frame', None):
            self._update_selected_template_details(self._npc1_current)
        
    def _refresh_npc_prompt_with_role(self):
        npcname = ''
        if hasattr(self, 'npc_role_combo'):
            t = self.npc_role_combo.currentText().strip()
            npcname = '' if (not t or t == '请选择一个角色') else t
        self._npc1_current = npcname
        npc2 = getattr(self, '_npc2_current', '')
        original_speakers = getattr(self, '_original_speakers', []) or []
        if original_speakers:
            expected = []
            for s in original_speakers:
                if s == 'npc1':
                    expected.append(npcname)
                elif s == 'npc2':
                    expected.append(npc2)
                else:
                    expected.append(s)
            self._expected_speakers = expected
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            old = self.npc_prompt_text.toPlainText()
            if old:
                lines = old.splitlines()
                for i, line in enumerate(lines):
                    if line.startswith('npc1:'):
                        lines[i] = f"npc1:{npcname}"
                    elif line.startswith('关键字：多应用'):
                        lines[i] = f"关键字：多应用{npcname}的人设特征"
                new_text = "\n".join(lines)
                self._npc_final_message = new_text
                self.npc_prompt_text.setPlainText(new_text)
                print("NPC提示词已根据角色更新（fallback）")

    def _load_npc_blank_frames(self):
        import os
        import duoki_editor as de
        pkg_root = os.path.dirname(os.path.abspath(de.__file__))
        blank_dir = os.path.join(pkg_root, 'resources', 'data', 'blank')
        self._npc_blank_dir = blank_dir
        if not os.path.isdir(blank_dir):
            return
        for i in range(self.npc_blank_grid.count()-1, -1, -1):
            item = self.npc_blank_grid.itemAt(i)
            if item is not None:
                w = item.widget()
                if w is not None:
                    w.setParent(None)
                self.npc_blank_grid.removeItem(item)
        files = []
        for name in os.listdir(blank_dir):
            p = os.path.join(blank_dir, name)
            if os.path.isfile(p) and not name.startswith("~$"):
                files.append(name)
        files.sort()
        col_count = 1
        idx = 0
        self._npc_template_frames = []
        self._selected_template_frame = None
        for name in files:
            frame = ContentGeneratorWidget.TemplateFrame(name, self._on_template_frame_clicked)
            frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            f_layout = QHBoxLayout(frame)
            f_layout.setContentsMargins(8, 8, 8, 8)
            left_name = QLabel(name)
            left_name.setStyleSheet("border: none;")
            f_layout.addWidget(left_name, 1)
            right_col = QWidget()
            right_v = QVBoxLayout(right_col)
            right_v.setContentsMargins(0, 0, 0, 0)
            right_v.setSpacing(4)
            row_labels = QHBoxLayout()
            row_labels.setContentsMargins(0, 0, 0, 0)
            row_labels.setSpacing(6)
            row_labels.setAlignment(Qt.AlignmentFlag.AlignRight)
            lab_wen = QLabel("文")
            lab_wen.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lab_wen.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
            lab_ci = QLabel("词")
            lab_ci.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lab_ci.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
            lab_tu = QLabel("图")
            lab_tu.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lab_tu.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
            lab_wen.setFixedWidth(24)
            lab_ci.setFixedWidth(24)
            lab_tu.setFixedWidth(24)
            lab_wen.setStyleSheet("color: #888888;")
            lab_ci.setStyleSheet("color: #888888;")
            lab_tu.setStyleSheet("color: #888888;")
            row_labels.addWidget(lab_wen)
            row_labels.addWidget(lab_ci)
            row_labels.addWidget(lab_tu)
            frame.lab_wen = lab_wen
            frame.lab_ci = lab_ci
            frame.lab_tu = lab_tu
            frame._left_name_label = left_name
            row_buttons_w = QWidget()
            row_buttons = QHBoxLayout(row_buttons_w)
            row_buttons.setContentsMargins(0, 0, 0, 0)
            row_buttons.setSpacing(6)
            row_buttons.setAlignment(Qt.AlignmentFlag.AlignRight)
            btn_wen = QPushButton("🔄️")
            btn_ci = QPushButton("🔄️")
            btn_tu = QPushButton("🔄️")
            btn_wen.setFixedWidth(24)
            btn_ci.setFixedWidth(24)
            btn_tu.setFixedWidth(24)
            frame.btn_wen = btn_wen
            frame.btn_ci = btn_ci
            frame.btn_tu = btn_tu
            btn_wen.clicked.connect(lambda checked=False, fr=frame: self._on_template_btn_wen_clicked(fr))
            btn_ci.clicked.connect(lambda checked=False, fr=frame: self._on_template_btn_ci_clicked(fr))
            btn_tu.clicked.connect(lambda checked=False, fr=frame: self._on_template_btn_tu_clicked(fr))
            row_buttons.addWidget(btn_wen)
            row_buttons.addWidget(btn_ci)
            row_buttons.addWidget(btn_tu)
            right_v.addLayout(row_labels)
            right_v.addWidget(row_buttons_w)
            frame.set_buttons_widget(row_buttons_w)
            f_layout.addWidget(right_col)
            row = idx // col_count
            col = idx % col_count
            self.npc_blank_grid.addWidget(frame, row, col)
            self._npc_template_frames.append(frame)
            idx += 1

        self._update_npc_template_count_label()

    def _update_npc_template_count_label(self):
        frames = list(getattr(self, '_npc_template_frames', []) or [])
        n = len(frames)
        done = 0
        for f in frames:
            if getattr(f, '_complete', False):
                done += 1
        if hasattr(self, 'npc_template_count_label') and self.npc_template_count_label:
            self.npc_template_count_label.setText(f"已完成{done}/{n}个模板")

    def _on_template_frame_clicked(self, frame):
        if hasattr(self, 'dialog_table'):
            self.dialog_table.setRowCount(0)
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            self.npc_prompt_text.clear()
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.clear()
        if hasattr(self, '_selected_template_frame') and self._selected_template_frame is not None and self._selected_template_frame is not frame:
            self._selected_template_frame.set_selected(False)
        frame.set_selected(True)
        self._selected_template_frame = frame
        import os
        p = os.path.join(getattr(self, '_npc_blank_dir', ''), getattr(frame, '_name', ''))
        if not p or not os.path.exists(p):
            print("未找到xlsx文件")
            return
        if not (p.endswith('.xlsx') or p.endswith('.xls')):
            print("文件不是xlsx")
            return
        import pandas as pd
        df = pd.read_excel(p, engine='openpyxl')
        if 'speaker' not in df.columns or 'param1' not in df.columns:
            print("缺少必要列: speaker 或 param1")
            return
        lines = []
        original_speakers = []
        for i in range(len(df)):
            s = str(df.iloc[i].get('speaker') or '').strip()
            a = str(df.iloc[i].get('param1') or '')
            original_speakers.append(s)
            lines.append(f"{s}:{a}")
        text = "\n".join(lines)
        npcname = self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else ''
        if npcname == "请选择一个角色":
            npcname = ''
        file_key = str(getattr(frame, '_name', '')).split('-', 1)[0]
        mm = self.constants_loader.get_map_manager() if hasattr(self, 'constants_loader') else {}
        npc2 = mm.get(file_key, '')
        self._npc1_current = npcname
        self._npc2_current = npc2
        expected = []
        for s in original_speakers:
            if s == 'npc1':
                expected.append(npcname)
            elif s == 'npc2':
                expected.append(npc2)
            else:
                expected.append(s)
        self._template_text_block = text
        self._original_speakers = original_speakers
        self._expected_speakers = expected
        if hasattr(self, 'npc_prompt_text'):
            self.npc_prompt_text.clear()
            self.npc_prompt_text.setPlainText(text)
        if getattr(self, '_npc1_current', None) is not None:
            self._update_selected_template_details(self._npc1_current)

    def _compose_npc_prompt_text(self, tpl_text: str, npc1: str, npc2: str) -> str:
        tpl = str(tpl_text or '').strip()
        parts = []
        parts.append(tpl)
        parts.append("")
        parts.append(f"npc1:{npc1}")
        parts.append(f"npc2:{npc2}")
        parts.append("")
        kw_line = f"关键字：多应用{npc1}的人设特征"
        if hasattr(self, 'npc_keyword_text') and self.npc_keyword_text:
            kw = self.npc_keyword_text.toPlainText().strip()
            if kw:
                kw_line = f"{kw_line}，{kw}"
        parts.append(kw_line)
        parts.append("")
        parts.append("工作模式：自动筛选")
        return "\n".join(parts)

    def on_npc_select_output_dir(self):
        current = self.config_manager.get_npc_output_directory()
        directory = QFileDialog.getExistingDirectory(self, "选择输出目录", current or './')
        if directory:
            absdir = os.path.abspath(directory)
            self.config_manager.set_npc_output_directory(absdir)
            self.npc_output_dir_display.setText(absdir)
            os.makedirs(absdir, exist_ok=True)
            print(f"NPC生产线输出目录: {absdir}")

    def on_npc_open_output_dir(self):
        d = self.npc_output_dir_display.text().strip() or self.config_manager.get_npc_output_directory()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择输出目录")
            return
        d = os.path.abspath(d)
        os.makedirs(d, exist_ok=True)
        if os.name == 'nt':
            os.startfile(d)
        else:
            import subprocess
            if sys.platform == 'darwin':
                subprocess.run(["open", d])
            else:
                subprocess.run(["xdg-open", d])
        print(f"已打开输出目录: {d}")

    def on_npc_start_generate(self):
        # 若正在聊天，则执行取消逻辑（需已收到chat_id）
        if self.is_chatting:
            if not getattr(self, '_chat_id_received', False):
                print("=== 尚未收到chat_id，无法中止对话 ===")
                return
            is_stream = self.stream_checkbox.isChecked()
            if is_stream:
                self._abort_stream_chat_npc()
            else:
                self._abort_non_stream_chat_npc()
            return

        # 第0步：检查是否选择了NPC
        if hasattr(self, 'npc_role_combo'):
            npc_val = self.npc_role_combo.currentText().strip()
            if (not npc_val) or npc_val == "请选择一个角色":
                Toast.warning("请选择一个NPC", self)
                return

        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            if hasattr(self.npc_image_container, 'clear'):
                self.npc_image_container.clear()
            print("已清空NPC图片容器")
        f = getattr(self, '_selected_template_frame', None)
        if f is not None:
            f.mark_step_in_progress('wen')

        tpl = getattr(self, '_template_text_block', '')
        npc1 = getattr(self, '_npc1_current', '')
        if hasattr(self, 'npc_role_combo') and (not npc1):
            t = self.npc_role_combo.currentText().strip()
            npc1 = '' if (not t or t == "请选择一个角色") else t
            self._npc1_current = npc1
        npc2 = getattr(self, '_npc2_current', '')
        base_text = str(tpl or '').strip()
        if not base_text and hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            base_text = self.npc_prompt_text.toPlainText().strip()
        if not base_text:
            Toast.warning("请先选择一个模板并生成提示词", self)
            return
        final_text = self._compose_npc_prompt_text(base_text, npc1, npc2)
        self._npc_final_message = final_text
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            self.npc_prompt_text.setPlainText(final_text)
        print("=== NPC对白生文提示词 ===")
        print(final_text)
        print("=== NPC对白生文提示词结束 ===")

        # 会话检查（内容对话ID）
        conv_id = self.config_manager.get_coze_content_conversation_dialog_id()
        if not conv_id:
            api_base = self.config_manager.get_coze_api()
            token = self.config_manager.get_coze_token()
            if not api_base or not token:
                print("配置错误：缺少Coze API或Token")
                return
            self.coze_client = CozeAPIClient(api_base, token)
            self.current_worker = CozeWorkerThread(self.coze_client, 'create_conversation')
            self.current_worker.conversation_created.connect(self._on_npc_conversation_created)
            self.current_worker.conversation_failed.connect(self._on_npc_conversation_failed)
            self._npc_pending_message = getattr(self, '_npc_final_message', '').strip()
            self.npc_start_generate_btn.setText("创建会话中...")
            self.npc_start_generate_btn.setEnabled(False)
            print("正在创建内容对话ID...")
            self.current_worker.start()
            return
        self.current_conversation_id = conv_id
        if hasattr(self, 'session_id_edit'):
            self.session_id_edit.setText(conv_id)

        # 检查准备好的提示词并发送消息到Coze（使用NPC上下文）
        message = getattr(self, '_npc_final_message', '').strip()
        if not message:
            Toast.warning("请先选择一个模板并生成提示词", self)
            return
        self._send_message(message=message, target='npc')

    def _on_npc_conversation_created(self, conversation_id: str):
        print(f"内容对话ID创建成功: {conversation_id}")
        self.config_manager.set_coze_content_conversation_dialog_id(conversation_id)
        self.current_conversation_id = conversation_id
        if hasattr(self, 'session_id_edit'):
            self.session_id_edit.setText(conversation_id)
        msg = getattr(self, '_npc_pending_message', '').strip()
        if not msg:
            msg = getattr(self, '_npc_final_message', '').strip()
        self._npc_pending_message = ''
        self._send_message(message=msg, target='npc')

    def _on_npc_conversation_failed(self, error_msg: str):
        print(f"内容对话ID创建失败: {error_msg}")
        self.npc_start_generate_btn.setText("开始生成")
        self.npc_start_generate_btn.setEnabled(True)
        if getattr(self, '_batch_generating', False):
            self._batch_skip_current_and_advance('内容会话创建失败')

    def _lock_npc_controls(self):
        if hasattr(self, 'npc_role_combo'):
            self.npc_role_combo.setEnabled(False)
        if hasattr(self, 'npc_select_output_dir_btn'):
            self.npc_select_output_dir_btn.setEnabled(False)
        if hasattr(self, 'npc_open_output_dir_btn'):
            self.npc_open_output_dir_btn.setEnabled(False)
        if hasattr(self, 'npc_blank_scroll') and not getattr(self, '_batch_generating', False):
            self.npc_blank_scroll.setEnabled(False)
        print("已锁定NPC生产线控件")

    def _unlock_npc_controls(self):
        if hasattr(self, 'npc_role_combo'):
            self.npc_role_combo.setEnabled(True)
        if hasattr(self, 'npc_select_output_dir_btn'):
            self.npc_select_output_dir_btn.setEnabled(True)
        if hasattr(self, 'npc_open_output_dir_btn'):
            self.npc_open_output_dir_btn.setEnabled(True)
        if hasattr(self, 'npc_blank_scroll') and not getattr(self, '_batch_generating', False):
            self.npc_blank_scroll.setEnabled(True)
        print("已解锁NPC生产线控件")
    
    def _get_batch_frames(self) -> List:
        return list(getattr(self, '_npc_template_frames', []) or [])
    
    def _reset_batch_runtime_state(self):
        self._batch_wen_running = False
        self._batch_ci_running = False
        self._batch_wen_current_frame = None
        self._batch_ci_current_frame = None
        self._batch_tu_workers = {}
        if hasattr(self, '_batch_ci_wait_timer') and self._batch_ci_wait_timer:
            self._batch_ci_wait_timer.stop()
        if hasattr(self, '_batch_tu_wait_timer') and self._batch_tu_wait_timer:
            self._batch_tu_wait_timer.stop()
    
    def on_npc_batch_generate(self):
        npc_val = self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else ''
        if (not npc_val) or npc_val == '请选择一个角色':
            Toast.warning("请选择一个NPC", self)
            print("[批量生成] 中止：未选择NPC")
            return
        if not self._check_user_name():
            print("[批量生成] 中止：未获取到用户名")
            return
        frames = list(getattr(self, '_npc_template_frames', []) or [])
        if not frames:
            print("[批量生成] 中止：模板列表为空")
            return
        if getattr(self, '_batch_generating', False):
            print("[批量生成] 已在进行中，忽略重复点击")
            return
        self._npc1_current = npc_val
        self._refresh_all_template_frames_status()
        self._reset_batch_runtime_state()
        self._batch_generating = True
        self._batch_index = 0
        self._batch_frames = frames
        self.npc_batch_generate_btn.setText("批量中...")
        self.npc_batch_generate_btn.setEnabled(False)
        print(f"[批量生成] 开始，共{len(frames)}个模板")
        for f in frames:
            f.set_selected(False)
            if getattr(f, 'btn_wen', None):
                f.btn_wen.setEnabled(False)
            if getattr(f, 'btn_ci', None):
                f.btn_ci.setEnabled(False)
            if getattr(f, 'btn_tu', None):
                f.btn_tu.setEnabled(False)
        self._selected_template_frame = None
        self._batch_schedule_wen_next()
        self._batch_scan_for_ci_candidates()
        self._batch_scan_for_tu_candidates()
    
    def _batch_schedule_wen_next(self):
        if not getattr(self, '_batch_generating', False):
            return
        if self._batch_wen_running:
            return
        if self.is_chatting:
            QTimer.singleShot(500, self._batch_schedule_wen_next)
            return
        frames = self._get_batch_frames()
        if not frames:
            self._maybe_finish_batch()
            return
        for frame in frames:
            if getattr(frame, '_failed_step', None):
                continue
            if getattr(frame, '_done_wen', False):
                continue
            self._batch_wen_running = True
            self._batch_wen_current_frame = frame
            message = self._build_npc_prompt_message_for_frame(frame)
            if not message:
                frame.mark_step_failed('wen')
                print(f"[批量生成] 模板生文提示词为空，标记失败: {getattr(frame, '_name', '')}")
                self._batch_wen_running = False
                self._batch_wen_current_frame = None
                continue
            self._npc_final_message = message
            print(f"[批量生成] 生文进程启动: {getattr(frame, '_name', '')}")
            self.on_npc_start_generate()
            return
        print("[批量生成] 生文进程已无可处理模板")
        self._maybe_finish_batch()
    
    def _batch_scan_for_ci_candidates(self):
        if not getattr(self, '_batch_generating', False):
            return
        if self._batch_ci_running:
            return
        if self.is_chatting:
            QTimer.singleShot(500, self._batch_scan_for_ci_candidates)
            return
        frames = self._get_batch_frames()
        if not frames:
            self._maybe_finish_batch()
            return
        candidates = []
        pending_wen = False
        for frame in frames:
            if getattr(frame, '_failed_step', None):
                continue
            if getattr(frame, '_done_wen', False) and not getattr(frame, '_done_ci', False):
                candidates.append(frame)
            elif not getattr(frame, '_done_wen', False) and not getattr(frame, '_done_ci', False):
                pending_wen = True
        if candidates:
            frame = candidates[0]
            self._batch_ci_running = True
            self._batch_ci_current_frame = frame
            npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
            if npc1 == '请选择一个角色':
                npc1 = ''
            npc2 = self._get_npc2_by_frame(frame)
            self._npc2_current = npc2
            tpl_text = self._build_ci_tpl_text_from_excel(frame, npc1, npc2)
            if not tpl_text:
                frame.mark_step_failed('ci')
                print(f"[批量生成] 模板生词源文案为空，标记失败: {getattr(frame, '_name', '')}")
                self._batch_ci_running = False
                self._batch_ci_current_frame = None
                self._batch_scan_for_ci_candidates()
                return
            message = self._compose_npc_prompt_text(tpl_text, npc1, npc2)
            print(f"[批量生成] 生词进程启动: {getattr(frame, '_name', '')}")
            self._ensure_image_conversation_and_send_image(message)
            return
        if pending_wen:
            if not self._batch_ci_wait_timer.isActive():
                print("[批量生成] 生词进程等待生文完成，5秒后重试")
                self._batch_ci_wait_timer.start(5000)
        else:
            print("[批量生成] 生词进程已无可处理模板")
            self._maybe_finish_batch()
    
    def _batch_scan_for_tu_candidates(self):
        if not getattr(self, '_batch_generating', False):
            return
        frames = self._get_batch_frames()
        if not frames:
            self._maybe_finish_batch()
            return
        active = self._batch_tu_workers or {}
        active_frames = list(active.values())
        candidates = []
        pending_ci = False
        for frame in frames:
            if getattr(frame, '_failed_step', None):
                continue
            done_ci = getattr(frame, '_done_ci', False)
            done_tu = getattr(frame, '_done_tu', False)
            if done_ci and not done_tu and frame not in active_frames:
                candidates.append(frame)
            elif not done_ci and not done_tu and getattr(frame, '_done_wen', False) and not getattr(frame, '_failed_step', None):
                pending_ci = True
        started = 0
        for frame in candidates:
            if len(self._batch_tu_workers) >= 5:
                break
            ok = self._batch_start_tu_for_frame(frame)
            if ok:
                started += 1
        if not candidates and pending_ci:
            if not self._batch_tu_wait_timer.isActive():
                print("[批量生成] 生图进程等待生词完成，5秒后重试")
                self._batch_tu_wait_timer.start(5000)
        elif not candidates and not self._batch_tu_workers and not pending_ci:
            print("[批量生成] 生图进程已无可处理模板")
            self._maybe_finish_batch()

    def _on_batch_ci_wait_timeout(self):
        if not getattr(self, '_batch_generating', False):
            return
        print("[批量生成] 生词进程超时检查，重新扫描可处理模板")
        self._batch_scan_for_ci_candidates()

    def _on_batch_tu_wait_timeout(self):
        if not getattr(self, '_batch_generating', False):
            return
        print("[批量生成] 生图进程超时检查，重新扫描可处理模板")
        self._batch_scan_for_tu_candidates()

    def _maybe_finish_batch(self):
        if not getattr(self, '_batch_generating', False):
            return
        if self._batch_wen_running or self._batch_ci_running:
            return
        if getattr(self, '_batch_tu_workers', None):
            if len(self._batch_tu_workers) > 0:
                return
        frames = self._get_batch_frames()
        if not frames:
            self._finish_batch_generate()
            return
        for f in frames:
            if getattr(f, '_failed_step', None) is None and not getattr(f, '_complete', False):
                return
        self._finish_batch_generate()
    
    def _batch_skip_current_and_advance(self, reason: str = ''):
        name = ''
        f = getattr(self, '_selected_template_frame', None)
        if f:
            name = getattr(f, '_name', '')
        msg = f"[批量生成] 跳过模板: {name}"
        if reason:
            msg += f"，原因：{reason}"
        print(msg)
        if not getattr(self, '_batch_generating', False):
            return
        if f is not None and getattr(f, '_failed_step', None) is None:
            step = None
            if getattr(self, '_batch_wen_current_frame', None) is f:
                step = 'wen'
            elif getattr(self, '_batch_ci_current_frame', None) is f:
                step = 'ci'
            if step:
                f.mark_step_failed(step)
        self._batch_wen_running = False
        self._batch_ci_running = False
        self._batch_wen_current_frame = None
        self._batch_ci_current_frame = None
        self._batch_schedule_wen_next()
        self._batch_scan_for_ci_candidates()
        self._batch_scan_for_tu_candidates()

    def _advance_batch_to_next_template(self):
        if not getattr(self, '_batch_generating', False):
            return
        self._batch_schedule_wen_next()
        self._batch_scan_for_ci_candidates()
        self._batch_scan_for_tu_candidates()
    
    def _finish_batch_generate(self):
        self._batch_generating = False
        self._batch_index = 0
        self._batch_frames = []
        self.npc_batch_generate_btn.setText("☢️批量生成")
        self.npc_batch_generate_btn.setEnabled(True)
        frames = self._get_batch_frames()
        for f in frames:
            if getattr(f, 'btn_wen', None):
                f.btn_wen.setEnabled(True)
            if getattr(f, 'btn_ci', None):
                f.btn_ci.setEnabled(True)
            if getattr(f, 'btn_tu', None):
                f.btn_tu.setEnabled(True)
        self._reset_batch_runtime_state()
        print("[批量生成] 流程结束")
    
    def _build_npc_prompt_message_for_frame(self, frame) -> str:
        base_dir = getattr(self, '_npc_blank_dir', '')
        name = getattr(frame, '_name', '')
        if not base_dir or not name:
            return ''
        p = os.path.join(base_dir, name)
        if not p or not os.path.exists(p):
            print("未找到xlsx文件")
            return ''
        if not (p.endswith('.xlsx') or p.endswith('.xls')):
            print("文件不是xlsx")
            return ''
        df = pd.read_excel(p, engine='openpyxl')
        if 'speaker' not in df.columns or 'param1' not in df.columns:
            print("缺少必要列: speaker 或 param1")
            return ''
        lines = []
        original_speakers = []
        for i in range(len(df)):
            s = str(df.iloc[i].get('speaker') or '').strip()
            a = str(df.iloc[i].get('param1') or '')
            original_speakers.append(s)
            lines.append(f"{s}:{a}")
        text = "\n".join(lines)
        npcname = self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else ''
        if npcname == "请选择一个角色":
            npcname = ''
        file_key = str(getattr(frame, '_name', '')).split('-', 1)[0]
        mm = self.constants_loader.get_map_manager() if hasattr(self, 'constants_loader') else {}
        npc2 = mm.get(file_key, '')
        self._npc1_current = npcname
        self._npc2_current = npc2
        expected = []
        for s in original_speakers:
            if s == 'npc1':
                expected.append(npcname)
            elif s == 'npc2':
                expected.append(npc2)
            else:
                expected.append(s)
        self._template_text_block = text
        self._original_speakers = original_speakers
        self._expected_speakers = expected
        self._npc_final_message = ''
        return text
    
    def _build_ci_tpl_text_from_excel(self, frame, npc1: str, npc2: str) -> str:
        excel_path = self._get_content_excel_path(npc1, frame)
        if not excel_path or not os.path.exists(excel_path):
            print("[批量生成] 未找到生文结果xlsx")
            return ''
        df = pd.read_excel(excel_path, engine='openpyxl')
        col_name = f"{npc1}-{npc2}-a7e0"
        if col_name not in list(df.columns):
            print("[批量生成] 生文结果中缺少目标列")
            return ''
        rows = []
        for i in range(len(df)):
            sp = str(df.iloc[i].get('speaker') or '').strip()
            dlg = str(df.iloc[i].get(col_name) or '').strip()
            if not dlg:
                continue
            if sp == 'npc1':
                disp = npc1
            elif sp == 'npc2':
                disp = npc2
            elif sp == 'duoki':
                disp = 'duoki'
            else:
                disp = sp
            rows.append(f"{disp}:{dlg}")
        return "\n".join(rows)
    
    def _get_image_prompt_from_excel(self, frame, npc1: str) -> str:
        excel_path = self._get_content_excel_path(npc1, frame)
        if not excel_path or not os.path.exists(excel_path):
            return ''
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        sheet = wb['image_prompt'] if 'image_prompt' in wb.sheetnames else None
        if sheet is None:
            return ''
        v = str(sheet['A1'].value or '').strip()
        return v
    
    def _batch_start_tu_for_frame(self, frame) -> bool:
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        prompt = self._get_image_prompt_from_excel(frame, npc1)
        if not prompt:
            print(f"[批量生成] 模板无生图提示词，跳过: {getattr(frame, '_name', '')}")
            return False
        api_token = self.config_manager.get_wuyun_token()
        if not api_token:
            print("错误: 未找到 wuyun API Token，请检查配置")
            frame.mark_step_failed('tu')
            return False
        p = str(prompt or '').strip()
        frame.mark_step_in_progress('tu')
        ref_paths = self._build_ref_image_paths(p)
        print(f"[批量生成] 生图进程参考图匹配数量: {len(ref_paths)}")
        if ref_paths:
            print(f"[批量生成] 生图进程参考图列表: {ref_paths}")
        full_prompt = f"{p}, {BATCH_CREATE_STORY_PROMPT}" if BATCH_CREATE_STORY_PROMPT else p
        data_dict = {'prompt': full_prompt, 'style': CARTOON_STYLE_SUFFIX, 'filename': '', 'ref_image_paths': ref_paths, 'width': None, 'height': None}
        worker = ImageGenerationWorker(
            api_token=api_token,
            prompt_data_list=[data_dict],
            model="gpt-4o-image-vip",
            size="1024x1024",
            num_images=1
        )
        if self._batch_tu_workers is None:
            self._batch_tu_workers = {}
        self._batch_tu_workers[worker] = frame
        worker.image_generated.connect(lambda data, img, w=worker: self._on_batch_npc_image_generated(w, data, img))
        worker.error_occurred.connect(lambda msg, data, w=worker: self._on_batch_npc_image_error(w, msg, data))
        worker.finished.connect(lambda w=worker: self._on_batch_npc_image_finished(w))
        print(f"[批量生成] 生图进程启动线程: {getattr(frame, '_name', '')}")
        worker.start()
        return True

    def _on_batch_npc_image_generated(self, worker, data_dict: Dict, image_bytes: bytes):
        frame = None
        if getattr(self, '_batch_tu_workers', None) is not None:
            frame = self._batch_tu_workers.get(worker)
        if frame is None:
            print("[批量生成] 生图进程收到结果，但未找到对应模板")
            return
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        image_path = self._get_image_path(npc1, frame)
        out_dir = os.path.dirname(image_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        working_bytes = image_bytes
        with Image.open(io.BytesIO(working_bytes)) as im:
            sw, sh = im.size
            width, height = 720, 720
            if sw > 0 and sh > 0 and width > 0 and height > 0:
                s = max(width / sw, height / sh)
                nw = max(1, int(round(sw * s)))
                nh = max(1, int(round(sh * s)))
                tmp = im.resize((nw, nh), Image.Resampling.LANCZOS)
                l = max(0, (nw - width) // 2)
                t = max(0, (nh - height) // 2)
                r = l + width
                b = t + height
                out = tmp.crop((l, t, r, b))
            else:
                out = im
            if out.mode != 'RGB':
                out = out.convert('RGB')
            buf = io.BytesIO()
            out.save(buf, format='JPEG', quality=90)
            working_bytes = buf.getvalue()
        working_bytes = add_watermark(working_bytes)
        with open(image_path, 'wb') as f:
            f.write(working_bytes)
        print(f"[批量生成] 生图进程已保存图片: {image_path}")
        frame.set_tu_done()
        self._update_npc_template_count_label()
        if getattr(self, '_batch_tu_workers', None) is not None:
            self._batch_tu_workers.pop(worker, None)
        self._batch_scan_for_tu_candidates()
        self._maybe_finish_batch()

    def _on_batch_npc_image_error(self, worker, error_message: str, data_dict: Dict):
        frame = None
        if getattr(self, '_batch_tu_workers', None) is not None:
            frame = self._batch_tu_workers.get(worker)
        name = getattr(frame, '_name', '') if frame is not None else ''
        print(f"[批量生成] 生图进程失败: 模板={name}, 错误={error_message}")
        detail = None
        full_body = None
        if isinstance(data_dict, dict):
            detail = data_dict.get('error_detail')
            full_body = data_dict.get('error_full_body')
        if detail:
            print(f"[批量生成] 生图详细错误: {detail}")
        if full_body:
            print("[批量生成] 生图返回体:\n" + str(full_body))
        popup_msg = error_message
        extra_parts = []
        if detail:
            extra_parts.append(f"详细错误:\n{detail}")
        if full_body:
            extra_parts.append(f"返回体:\n{full_body}")
        if extra_parts:
            popup_msg = popup_msg + "\n\n" + "\n\n".join(extra_parts)
        if frame is not None:
            frame.mark_step_failed('tu')
        if getattr(self, '_batch_tu_workers', None) is not None:
            self._batch_tu_workers.pop(worker, None)
        QMessageBox.critical(self, "生成失败", popup_msg)
        self._batch_scan_for_tu_candidates()
        self._maybe_finish_batch()

    def _on_batch_npc_image_finished(self, worker):
        if getattr(self, '_batch_tu_workers', None) is not None and worker in self._batch_tu_workers:
            self._batch_tu_workers.pop(worker, None)
            print("[批量生成] 生图线程结束，移除记录")
        self._batch_scan_for_tu_candidates()
        self._maybe_finish_batch()
    
    def _create_left_panel(self) -> QWidget:
        """创建左侧控制面板"""
        panel = QWidget()
        panel.setFixedWidth(280)
        
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)
        
        # 第一组：平台、智能体、创建会话
        session_group = QGroupBox()
        session_group_layout = QVBoxLayout(session_group)
        session_group_layout.setContentsMargins(3, 3, 3, 3)
        session_group_layout.setSpacing(5)
        
        # 平台选择
        platform_label = QLabel("平台:")
        platform_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        session_group_layout.addWidget(platform_label)
        
        self.platform_combo = QComboBox()
        self.platform_combo.addItems(list(self.constants_loader.get_ai_platforms().keys()))
        self.platform_combo.setCurrentText(self.current_platform)
        self.platform_combo.currentTextChanged.connect(self._on_platform_changed)
        session_group_layout.addWidget(self.platform_combo)
        
        # 智能体选择
        agent_label = QLabel("智能体:")
        agent_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        session_group_layout.addWidget(agent_label)
        
        self.agent_combo = QComboBox()
        self._update_agent_combo()
        self.agent_combo.currentTextChanged.connect(self._on_agent_changed)
        session_group_layout.addWidget(self.agent_combo)
        
        # 会话按钮
        self.session_btn = QPushButton("创建会话")
        self.session_btn.clicked.connect(self._create_conversation)
        session_group_layout.addWidget(self.session_btn)
        
        layout.addWidget(session_group)
        
        # 第二组：用户名、会话ID、流式选项
        user_group = QGroupBox()
        user_group_layout = QVBoxLayout(user_group)
        user_group_layout.setContentsMargins(3, 3, 3, 3)
        user_group_layout.setSpacing(5)
        
        # 用户名输入框
        username_label = QLabel("用户名:")
        username_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        user_group_layout.addWidget(username_label)
        
        self.username_edit = QLineEdit()
        self.username_edit.setReadOnly(True)
        self.username_edit.setPlaceholderText("正在获取用户名...")
        user_group_layout.addWidget(self.username_edit)
        
        # 会话ID输入框
        session_id_label = QLabel("会话ID:")
        session_id_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        user_group_layout.addWidget(session_id_label)
        
        self.session_id_edit = QLineEdit()
        self.session_id_edit.setReadOnly(True)
        self.session_id_edit.setPlaceholderText("点击创建会话按钮获取会话ID")
        user_group_layout.addWidget(self.session_id_edit)
        
        # 流式复选框
        self.stream_checkbox = QCheckBox("流式")
        self.stream_checkbox.setChecked(True)  # 默认选中
        user_group_layout.addWidget(self.stream_checkbox)
        
        layout.addWidget(user_group)
        
        # 添加弹性空间
        layout.addStretch()
        
        return panel
    
    def _create_right_panel(self) -> QWidget:
        """创建右侧对话面板"""
        panel = QWidget()
        
        # 创建水平分割器，分为三列：输入面板、输出面板、历史记录列表
        splitter = QSplitter(Qt.Orientation.Horizontal)
        layout = QHBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(splitter)
        
        # 第一列：输入面板
        input_panel = self._create_input_panel()
        splitter.addWidget(input_panel)
        
        # 第二列：输出面板
        output_panel = self._create_output_panel()
        splitter.addWidget(output_panel)
        
        # 第三列：历史记录列表
        history_panel = self._create_history_panel()
        splitter.addWidget(history_panel)
        
        # 设置三列的宽度比例为 2:3:1（输入:输出:历史）
        splitter.setSizes([200, 300, 100])
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        splitter.setStretchFactor(2, 1)
        
        return panel
    

    
    def _create_input_panel(self) -> QWidget:
        """创建输入面板"""
        panel = QWidget()
        
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # 标题
        title_label = QLabel("提示词输入")
        title_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 输入框 - 撑满高度，使用自定义PlainTextEdit避免Excel格式问题
        self.input_text = PlainTextEdit()
        self.input_text.setPlaceholderText("请输入您的提示词...")
        # 移除最大高度限制，让输入框撑满高度
        layout.addWidget(self.input_text, 1)  # stretch factor = 1
        
        # 发送/中止按钮
        self.send_btn = QPushButton("发送")
        self.send_btn.clicked.connect(self._send_message)
        layout.addWidget(self.send_btn)
        
        return panel
    
    def _create_output_panel(self) -> QWidget:
        """创建输出面板"""
        panel = QWidget()
        
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # 标题
        title_label = QLabel("智能体输出")
        title_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 输出框
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setPlaceholderText("AI回复将在这里显示...")
        layout.addWidget(self.output_text)
        
        # 格式化按钮（隐藏，改为自动格式化）
        self.format_btn = QPushButton("格式化")
        self.format_btn.setEnabled(False)  # 初始状态不可用
        self.format_btn.setVisible(False)  # 隐藏按钮
        self.format_btn.clicked.connect(self._format_content)
        layout.addWidget(self.format_btn)
        
        # Tab容器用于显示解析后的数据
        self.data_tab_widget = QTabWidget()
        self.data_tab_widget.setVisible(False)  # 初始隐藏
        self.data_tab_widget.currentChanged.connect(self._on_tab_changed)  # 监听tab切换事件
        layout.addWidget(self.data_tab_widget)
        
        return panel
    

    
    def _update_agent_combo(self):
        """更新智能体下拉列表"""
        self.agent_combo.clear()
        
        # 获取当前平台对应的智能体
        ai_platforms = self.constants_loader.get_ai_platforms()
        platform_key = ai_platforms.get(self.current_platform)
        if platform_key:
            ai_agents = self.constants_loader.get_ai_agents()
            agents = ai_agents.get(self.current_platform, [])
            self.agent_combo.addItems(agents)
            
            # 设置默认选择
            if agents and self.current_agent in agents:
                self.agent_combo.setCurrentText(self.current_agent)
            elif agents:
                self.current_agent = agents[0]
                self.agent_combo.setCurrentText(self.current_agent)
    
    def _load_conversation_id(self):
        """加载当前智能体的会话ID"""
        if self.current_agent == "多奇对白生成器":
            conversation_id = self.config_manager.get_coze_conversation_dialog_id()
            self.current_conversation_id = conversation_id or ""
            self.session_id_edit.setText(self.current_conversation_id)
    

    
    def _on_platform_changed(self, platform: str):
        """平台选择改变"""
        self.current_platform = platform
        self._update_agent_combo()
    
    def _on_agent_changed(self, agent: str):
        """智能体选择改变"""
        self.current_agent = agent
        self._load_conversation_id()
    

    
    def _create_conversation(self):
        """创建新会话"""
        if not self._check_user_name():
            if getattr(self, '_batch_generating', False):
                print("[批量生成] 中止：未获取到用户名")
                self._finish_batch_generate()
            return
        
        # 获取API配置
        api_base = self.config_manager.get_coze_api()
        token = self.config_manager.get_coze_token()
        
        if not api_base or not token:
            QMessageBox.warning(self, "配置错误", "请先配置Coze API地址和令牌")
            return
        
        # 创建API客户端
        self.coze_client = CozeAPIClient(api_base, token)
        
        # 创建工作线程
        self.current_worker = CozeWorkerThread(self.coze_client, 'create_conversation')
        self.current_worker.conversation_created.connect(self._on_conversation_created)
        self.current_worker.conversation_failed.connect(self._on_conversation_failed)
        
        # 禁用按钮
        self.session_btn.setEnabled(False)
        self.session_btn.setText("创建中...")
        
        # 启动线程
        self.current_worker.start()
    
    def _send_message(self, message=None, target='dialog', bot_id_override=None):
        """发送消息，可指定目标上下文"""
        if self.is_chatting:
            if not getattr(self, '_chat_id_received', False):
                print("=== 尚未收到chat_id，无法中止对话 ===")
                return
            is_stream = self.stream_checkbox.isChecked()
            if target == 'npc':
                if is_stream:
                    self._abort_stream_chat_npc()
                else:
                    self._abort_non_stream_chat_npc()
            else:
                if is_stream:
                    self._abort_stream_chat()
                else:
                    self._abort_non_stream_chat()
            return

        if not self._check_user_name():
            return

        if message is None:
            message = self.input_text.toPlainText().strip()
        if not message:
            QMessageBox.warning(self, "输入错误", "请输入提示词")
            return

        if not self.current_conversation_id:
            QMessageBox.warning(self, "会话错误", "请先创建会话")
            return

        api_base = self.config_manager.get_coze_api()
        token = self.config_manager.get_coze_token()
        bot_id = bot_id_override or self.config_manager.get_coze_duoki_content_bot_id()
        user_name = self.get_user_name()
        if not all([api_base, token, bot_id]):
            QMessageBox.warning(self, "配置错误", "请检查Coze相关配置")
            return

        self.coze_client = CozeAPIClient(api_base, token)
        is_stream = self.stream_checkbox.isChecked()

        self.current_worker = CozeWorkerThread(
            self.coze_client,
            'send_message',
            bot_id=bot_id,
            conversation_id=self.current_conversation_id,
            user_name=user_name,
            message=message,
            stream=is_stream
        )

        if target == 'npc':
            self.current_worker.message_received.connect(self._on_message_received_npc)
            self.current_worker.chat_finished.connect(self._on_chat_finished_npc)
            self.current_worker.chat_failed.connect(self._on_chat_failed_npc)
            self.coze_client.chat_id_received.connect(self._on_chat_id_received_npc)
        elif target == 'npc_image':
            self.current_worker.message_received.connect(self._on_message_received_npc_image)
            self.current_worker.chat_finished.connect(self._on_chat_finished_npc_image)
            self.current_worker.chat_failed.connect(self._on_chat_failed_npc)
            self.coze_client.chat_id_received.connect(self._on_chat_id_received_npc)
        else:
            self.current_worker.message_received.connect(self._on_message_received)
            self.current_worker.chat_finished.connect(self._on_chat_finished)
            self.current_worker.chat_failed.connect(self._on_chat_failed)
            self.coze_client.chat_id_received.connect(self._on_chat_id_received)

        if is_stream:
            if target == 'npc':
                self.current_worker.chat_cancelled.connect(self._on_stream_chat_cancelled_npc)
                self.current_worker.stream_aborted.connect(self._on_stream_aborted_npc)
                self.current_worker.cancel_status_verified.connect(self._on_cancel_status_verified_npc)
            elif target == 'npc_image':
                self.current_worker.chat_cancelled.connect(self._on_stream_chat_cancelled_npc)
                self.current_worker.stream_aborted.connect(self._on_stream_aborted_npc)
                self.current_worker.cancel_status_verified.connect(self._on_cancel_status_verified_npc)
            else:
                self.current_worker.chat_cancelled.connect(self._on_stream_chat_cancelled)
                self.current_worker.stream_aborted.connect(self._on_stream_aborted)
                self.current_worker.cancel_status_verified.connect(self._on_cancel_status_verified)
        else:
            if target == 'npc':
                self.current_worker.chat_cancelled.connect(self._on_non_stream_chat_cancelled_npc)
            elif target == 'npc_image':
                self.current_worker.chat_cancelled.connect(self._on_non_stream_chat_cancelled_npc)
            else:
                self.current_worker.chat_cancelled.connect(self._on_non_stream_chat_cancelled)

        if target == 'npc':
            handler = getattr(self, '_on_non_stream_aborted_npc', None)
            if handler is None:
                handler = self._on_non_stream_aborted
            self.current_worker.non_stream_aborted.connect(handler)
        elif target == 'npc_image':
            handler = getattr(self, '_on_non_stream_aborted_npc', None)
            if handler is None:
                handler = self._on_non_stream_aborted
            self.current_worker.non_stream_aborted.connect(handler)
        else:
            self.current_worker.non_stream_aborted.connect(self._on_non_stream_aborted)

        self.is_chatting = True
        self._chat_id_received = False
        self._chat_target = target
        if target == 'npc':
            self.npc_start_generate_btn.setText("连接中...")
            self.npc_start_generate_btn.setEnabled(False)
            self.npc_prompt_text.clear()
            self.npc_prompt_text.append("正在连接...")
            self._lock_npc_controls()
        elif target == 'npc_image':
            self.npc_start_generate_btn.setText("连接中...")
            self.npc_start_generate_btn.setEnabled(False)
            self.npc_prompt_text.clear()
            self._lock_npc_controls()
        else:
            self.send_btn.setText("连接中...")
            self.send_btn.setEnabled(False)
            self.output_text.clear()
            self.output_text.append("正在连接...")
            self._lock_input_controls()

        self._cancel_processed = False
        self.data_tab_widget.setVisible(False)
        self.format_btn.setVisible(False)
        self.format_btn.setEnabled(False)
        self.current_worker.start()

    def _on_message_received_npc(self, content: str):
        if self.is_chatting:
            current_text = self.npc_prompt_text.toPlainText()
            if current_text == "正在生成回复...\n":
                self.npc_prompt_text.clear()
            cursor = self.npc_prompt_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.npc_prompt_text.setTextCursor(cursor)
            self.npc_prompt_text.insertPlainText(content)
            cursor = self.npc_prompt_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.npc_prompt_text.setTextCursor(cursor)

    def _on_chat_finished_npc(self, chat_data: Dict):
        print("=== NPC相关文案生成完成 ===")
        self._reset_chat_state_npc()
        Toast.info("NPC相关文案生成完成", self)
        output = self.npc_prompt_text.toPlainText().strip()
        if not output:
            output = str(chat_data.get('full_content', '')).strip()
        if output:
            self._format_npc_output(output)
            ok = self._validate_and_maybe_regenerate(output)
            if getattr(self, '_batch_generating', False) and ok:
                self._batch_wen_running = False
                self._batch_wen_current_frame = None
                self._batch_schedule_wen_next()
                self._batch_scan_for_ci_candidates()
                self._batch_scan_for_tu_candidates()

    def _on_chat_failed_npc(self, error_msg: str):
        print(f"=== NPC相关文案生成失败: {error_msg} ===")
        step = 'ci' if getattr(self, '_chat_target', '') == 'npc_image' else 'wen'
        f = getattr(self, '_selected_template_frame', None)
        if f is not None:
            f.mark_step_failed(step)
        self._reset_chat_state_npc()
        QMessageBox.critical(self, "对话失败", error_msg)
        if getattr(self, '_batch_generating', False):
            self._batch_skip_current_and_advance('生文失败')

    def _on_chat_id_received_npc(self, chat_id: str):
        print(f"=== NPC收到chat_id: {chat_id}，切换按钮为中止状态 ===")
        if self.is_chatting and not self._chat_id_received:
            self._chat_id_received = True
            self.npc_start_generate_btn.setText("中止")
            self.npc_start_generate_btn.setEnabled(True)
            self.npc_prompt_text.clear()
            self.npc_prompt_text.append("正在生成回复...\n")
            print("NPC按钮已切换为中止状态")

    def _on_message_received_npc_image(self, content: str):
        if self.is_chatting:
            cursor = self.npc_prompt_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.npc_prompt_text.setTextCursor(cursor)
            self.npc_prompt_text.insertPlainText(content)

    def _on_chat_finished_npc_image(self, chat_data: Dict):
        print("=== 生图提示词生成完成 ===")
        full = str(chat_data.get('full_content', '')).strip()
        if getattr(self, '_batch_generating', False):
            self._reset_chat_state_npc()
            if full:
                self._append_image_prompt_to_last_excel(full)
            else:
                print("[批量生成] 生图提示词为空，未写入Excel")
            self._batch_ci_running = False
            self._batch_ci_current_frame = None
            self._batch_scan_for_ci_candidates()
            self._batch_scan_for_tu_candidates()
            return
        self.npc_prompt_text.clear()
        if full:
            self.npc_prompt_text.setPlainText(full)
        Toast.info("生图提示词生成完成", self)
        self._append_image_prompt_to_last_excel(full)
        self._start_npc_image_generation_from_prompt(full)

    def _reset_chat_state_npc(self):
        self.is_chatting = False
        self.npc_start_generate_btn.setText("开始生成")
        self.npc_start_generate_btn.setEnabled(True)
        self.current_worker = None
        self._chat_id_received = False
        self._unlock_npc_controls()

    def _abort_stream_chat_npc(self):
        self.npc_start_generate_btn.setText("取消中...")
        self.npc_start_generate_btn.setEnabled(False)
        self.npc_prompt_text.append("\n正在取消对话...")
        worker = self.current_worker
        if worker and worker.isRunning():
            worker.cancel_stream_chat()
            worker.cancel()
        else:
            self._reset_chat_state_npc()

    def _abort_non_stream_chat_npc(self):
        worker = self.current_worker
        if worker and worker.isRunning():
            worker.cancel_non_stream_chat()
            worker.cancel()
        self._reset_chat_state_npc()

    def _on_stream_chat_cancelled_npc(self, message: str):
        step = '生图提示词生成' if getattr(self, '_chat_target', '') == 'npc_image' else 'NPC相关文案生成'
        print(f"{step}（流式）被取消: {message}")
        current_output = self.npc_prompt_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            self.npc_prompt_text.append(f"\n\n--- {message} ---")
        else:
            self.npc_prompt_text.setPlainText(f"--- {message} ---")
        if getattr(self, '_cancel_processed', False):
            return
        self._cancel_processed = True
        self._reset_chat_state_npc()
        Toast.info(f"{step}已取消", self)

    def _on_non_stream_chat_cancelled_npc(self, message: str):
        step = '生图提示词生成' if getattr(self, '_chat_target', '') == 'npc_image' else 'NPC相关文案生成'
        print(f"{step}（非流式）被取消: {message}")
        current_output = self.npc_prompt_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            self.npc_prompt_text.append(f"\n\n--- {message} ---")
        else:
            self.npc_prompt_text.setPlainText(f"--- {message} ---")
        if getattr(self, '_cancel_processed', False):
            return
        self._cancel_processed = True
        self._reset_chat_state_npc()
        Toast.info(f"{step}已取消", self)

    def _on_cancel_status_verified_npc(self, cancel_result: dict):
        step = '生图提示词生成' if getattr(self, '_chat_target', '') == 'npc_image' else 'NPC相关文案生成'
        print(f"=== 取消状态验证结果（{step}）: {cancel_result} ===")
        if cancel_result['success'] and cancel_result['status'] == 'canceled':
            self.npc_prompt_text.append("\n\n--- 流式对话已成功取消 ---")
            Toast.info(f"{step}已取消", self)
            self._reset_chat_state_npc()
        else:
            error_msg = cancel_result.get('message', '取消失败')
            self.npc_prompt_text.append(f"\n取消失败: {error_msg}")
            self.npc_start_generate_btn.setText("中止")
            self.npc_start_generate_btn.setEnabled(True)

    def _on_stream_aborted_npc(self, content: str):
        if getattr(self, '_cancel_processed', False):
            return
        self._cancel_processed = True
        self._reset_chat_state_npc()
        step = '生图提示词生成' if getattr(self, '_chat_target', '') == 'npc_image' else 'NPC相关文案生成'
        Toast.info(f"[调试] {step}中止", self)

    def _validate_and_maybe_regenerate(self, original_output: str):
        expected = getattr(self, '_expected_speakers', []) or []
        formatted = getattr(self, '_formatted_npcs', []) or []
        if expected and formatted and len(expected) == len(formatted):
            ok = True
            for i in range(len(expected)):
                if str(expected[i]).strip() != str(formatted[i]).strip():
                    ok = False
                    break
            print(f"Step3匹配结果: {'通过' if ok else '不通过'}")
            if ok:
                self._save_formatted_dialogs_to_excel()
                if not getattr(self, '_batch_generating', False):
                    self._start_image_generation_from_first_line(original_output)
                return True
        print("Step3匹配结果: 不通过")
        Toast.warning("校验未通过，立即重新生成", self)
        msg = getattr(self, '_npc_final_message', '').strip()
        if msg:
            self._send_message(message=msg, target='npc')
        return False

    def _format_npc_output(self, content: str):
        rows = []
        formatted_npcs = []
        formatted_dialogs = []
        for line in content.splitlines():
            t = line.strip()
            if not t:
                continue
            m = re.search(r'[：:]', t)
            idx = m.start() if m else -1
            if idx == -1:
                continue
            npc = t[:idx].strip()
            dialog = t[idx+1:].strip()
            if npc and dialog and npc != '评价':
                rows.append((npc, dialog))
                formatted_npcs.append(npc)
                formatted_dialogs.append(dialog)
        if hasattr(self, 'dialog_table') and not getattr(self, '_batch_generating', False):
            self.dialog_table.setRowCount(0)
            self.dialog_table.setRowCount(len(rows))
            from PyQt6.QtWidgets import QTableWidgetItem
            for i, (n, d) in enumerate(rows):
                self.dialog_table.setItem(i, 0, QTableWidgetItem(n))
                self.dialog_table.setItem(i, 1, QTableWidgetItem(d))
            print(f"NPC格式化完成，共添加 {len(rows)} 行")
        self._formatted_npcs = formatted_npcs
        self._formatted_dialogs = formatted_dialogs
        print(f"Step2格式化后npc数组: {formatted_npcs}")

    def _start_image_generation_from_first_line(self, original_output: str):
        line = ''
        for l in original_output.splitlines():
            if l.strip():
                line = l.strip()
                break
        if not line:
            return
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        npc2 = getattr(self, '_npc2_current', '')
        msg = line.replace('{user_name}', '').replace('{npc1_name}', npc1).replace('{npc2_name}', npc2)
        self._ensure_image_conversation_and_send_image(msg)

    def _start_npc_image_generation_from_prompt(self, prompt_text: str):
        p = str(prompt_text or '').strip()
        if not p:
            print("[NPC图片生成] 提示词为空，跳过")
            return
        f = getattr(self, '_selected_template_frame', None)
        if f is not None:
            f.mark_step_in_progress('tu')
        api_token = self.config_manager.get_wuyun_token()
        if not api_token:
            QMessageBox.critical(self, "错误", "未找到 wuyun API Token，请检查配置")
            return
        # 在生成前根据提示词匹配参考图附件
        ref_paths = self._build_ref_image_paths(p)
        print(f"[NPC图片生成] 参考图匹配数量: {len(ref_paths)}")
        if ref_paths:
            print(f"[NPC图片生成] 参考图列表: {ref_paths}")
        full_prompt = f"{p}, {BATCH_CREATE_STORY_PROMPT}" if BATCH_CREATE_STORY_PROMPT else p
        data_dict = {'prompt': full_prompt, 'style': CARTOON_STYLE_SUFFIX, 'filename': '', 'ref_image_paths': ref_paths, 'width': None, 'height': None}
        print("[NPC图片生成] 开始，锁定控件并显示加载动画")
        self.npc_start_generate_btn.setText("生图中...")
        self.npc_start_generate_btn.setEnabled(False)
        self._lock_npc_controls()
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.start_loading("生图中")
        self._npc_image_worker = ImageGenerationWorker(
            api_token=api_token,
            prompt_data_list=[data_dict],
            model="gpt-4o-image-vip",
            size="1024x1024",
            num_images=1
        )
        self._npc_image_worker.progress_updated.connect(self._on_npc_image_progress)
        self._npc_image_worker.image_generated.connect(self._on_npc_image_generated)
        self._npc_image_worker.error_occurred.connect(self._on_npc_image_error)
        self._npc_image_worker.finished.connect(self._on_npc_image_generation_finished)
        self._npc_image_worker.start()

    def _build_ref_image_paths(self, prompt: str) -> list:
        """根据提示词匹配 constants 中的 image_references，返回存在的绝对路径列表"""
        refs = self.constants_loader.get_image_references() or {}
        if not isinstance(refs, dict):
            return []
        matched_rel_paths = []
        for key, rel in refs.items():
            if key and isinstance(key, str) and key in prompt:
                matched_rel_paths.append(rel)
        # 去重保持顺序
        seen = set()
        matched_rel_paths = [r for r in matched_rel_paths if not (r in seen or seen.add(r))]
        if not matched_rel_paths:
            return []
        import duoki_editor as de
        base_dir = os.path.dirname(os.path.abspath(de.__file__))
        abs_paths = []
        for rel in matched_rel_paths:
            abs_path = rel if os.path.isabs(rel) else os.path.join(base_dir, rel)
            if os.path.exists(abs_path):
                abs_paths.append(abs_path)
            else:
                print(f"[NPC图片生成] 参考图片不存在: {abs_path}")
        return abs_paths

    def _get_orig_base_from_name(self, name: str) -> str:
        stem = os.path.splitext(str(name or ''))[0]
        stem = re.sub(r'-blank', '', stem)
        stem = re.sub(r'[\s\-_]+$', '', stem)
        return stem

    def _get_npc2_by_frame(self, frame) -> str:
        mm = self.constants_loader.get_map_manager() if hasattr(self, 'constants_loader') else {}
        file_key = str(getattr(frame, '_name', '')).split('-', 1)[0]
        return mm.get(file_key, '')

    def _get_content_excel_path(self, npc1: str, frame) -> str:
        dir_root = self.config_manager.get_npc_output_directory()
        content_dir = os.path.join(dir_root, (npc1 or 'unknown'), 'content')
        name = getattr(frame, '_name', '')
        orig = self._get_orig_base_from_name(name)
        return os.path.join(content_dir, f"{orig}-{npc1}.xlsx")

    def _get_image_path(self, npc1: str, frame) -> str:
        dir_root = self.config_manager.get_npc_output_directory()
        image_dir = os.path.join(dir_root, (npc1 or 'unknown'), 'image')
        npc2 = self._get_npc2_by_frame(frame)
        orig = self._get_orig_base_from_name(getattr(frame, '_name', ''))
        return os.path.join(image_dir, f"{orig}_{npc1}-{npc2}.jpg")

    def _refresh_template_frame_status(self, frame, npc1: str):
        if not frame or not npc1:
            return
        excel_path = self._get_content_excel_path(npc1, frame)
        npc2 = self._get_npc2_by_frame(frame)
        col_name = f"{npc1}-{npc2}-a7e0"
        if os.path.exists(excel_path):
            import pandas as pd
            df = pd.read_excel(excel_path, engine='openpyxl')
            if col_name in list(df.columns):
                frame.set_wen_done()
                print(f"已检测到Excel列: {col_name}")
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            if 'image_prompt' in wb.sheetnames:
                ws = wb['image_prompt']
                v = str(ws['A1'].value or '').strip()
                if v:
                    frame.set_ci_done()
                    print("已检测到image_prompt并存在内容")
        image_path = self._get_image_path(npc1, frame)
        if os.path.exists(image_path):
            frame.set_tu_done()
            print(f"已检测到图片: {image_path}")

    def _refresh_all_template_frames_status(self):
        npc1 = getattr(self, '_npc1_current', '') or ''
        frames = list(getattr(self, '_npc_template_frames', []) or [])
        if not frames:
            return
        if not npc1:
            for f in frames:
                self._reset_template_frame_status(f, reset_all=True)
            self._update_npc_template_count_label()
            return
        for f in frames:
            self._reset_template_frame_status(f, reset_all=True)
            self._refresh_template_frame_status(f, npc1)
        self._update_npc_template_count_label()

    def _update_selected_template_details(self, npc1: str):
        frame = getattr(self, '_selected_template_frame', None)
        if not frame or not npc1:
            return
        excel_path = self._get_content_excel_path(npc1, frame)
        npc2 = self._get_npc2_by_frame(frame)
        col_name = f"{npc1}-{npc2}-a7e0"
        if os.path.exists(excel_path):
            import pandas as pd
            df = pd.read_excel(excel_path, engine='openpyxl')
            if col_name in list(df.columns):
                rows = []
                for i in range(len(df)):
                    sp = str(df.iloc[i].get('speaker') or '').strip()
                    dlg = str(df.iloc[i].get(col_name) or '').strip()
                    if not dlg:
                        continue
                    if sp == 'npc1':
                        disp = npc1
                    elif sp == 'npc2':
                        disp = npc2
                    elif sp == 'duoki':
                        disp = 'duoki'
                    else:
                        disp = sp
                    rows.append((disp, dlg))
                if hasattr(self, 'dialog_table'):
                    self.dialog_table.setRowCount(0)
                    self.dialog_table.setRowCount(len(rows))
                    from PyQt6.QtWidgets import QTableWidgetItem
                    for i, (n, d) in enumerate(rows):
                        self.dialog_table.setItem(i, 0, QTableWidgetItem(n))
                        self.dialog_table.setItem(i, 1, QTableWidgetItem(d))
                    print(f"已加载Excel到对话表，共 {len(rows)} 行")
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheet = wb['image_prompt'] if 'image_prompt' in wb.sheetnames else None
            if sheet is not None:
                v = str(sheet['A1'].value or '').strip()
                if v:
                    if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
                        self.npc_prompt_text.setPlainText(v)
                    print("已将image_prompt写入提示词框")
        image_path = self._get_image_path(npc1, frame)
        if os.path.exists(image_path):
            pix = QPixmap(image_path)
            if hasattr(self, 'npc_image_container') and self.npc_image_container:
                self.npc_image_container.set_original_pixmap(pix)
            print(f"已显示图片: {image_path}")

    def _save_formatted_dialogs_to_excel(self):
        import os
        import pandas as pd
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        npc2 = getattr(self, '_npc2_current', '')
        dialogs = getattr(self, '_formatted_dialogs', []) or []
        dir_root = self.config_manager.get_npc_output_directory()
        content_dir = os.path.join(dir_root, (npc1 or 'unknown'), 'content')
        os.makedirs(content_dir, exist_ok=True)
        base_dir = getattr(self, '_npc_blank_dir', '')
        frame = getattr(self, '_selected_template_frame', None)
        if getattr(self, '_batch_generating', False):
            frame = getattr(self, '_batch_wen_current_frame', frame)
        name = getattr(frame, '_name', '') if frame is not None else ''
        src_path = os.path.join(base_dir, name) if base_dir and name else ''
        if not src_path:
            print("[NPC内容存储] 原模板路径为空")
            return
        df = pd.read_excel(src_path, engine='openpyxl')
        df_out = df.copy()
        col_name = f"{npc1}-{npc2}-a7e0"
        insert_idx = df_out.columns.get_loc('param1') + 1 if 'param1' in df_out.columns else len(df_out.columns)
        m = len(df_out)
        vals = dialogs[:m]
        if len(vals) < m:
            vals = vals + [''] * (m - len(vals))
        df_out.insert(insert_idx, col_name, vals)
        orig_base = os.path.splitext(os.path.basename(name))[0] if name else 'output'
        orig_base = re.sub(r'-blank', '', orig_base)
        orig_base = re.sub(r'[\s\-_]+$', '', orig_base)
        out_name = f"{orig_base}-{npc1}.xlsx"
        out_path = os.path.join(content_dir, out_name)
        df_out.to_excel(out_path, index=False)
        print(f"[NPC内容存储] 已生成文件: {out_path}")
        self._last_saved_content_path = out_path
        self._mark_selected_template('wen')

    def _append_image_prompt_to_last_excel(self, prompt_text: str):
        import os
        from openpyxl import load_workbook
        from PyQt6.QtCore import QDateTime
        p = getattr(self, '_last_saved_content_path', '')
        if not p or not os.path.exists(p):
            npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
            if npc1 == '请选择一个角色':
                npc1 = ''
            frame = getattr(self, '_selected_template_frame', None)
            if getattr(self, '_batch_generating', False):
                frame = getattr(self, '_batch_ci_current_frame', frame)
            if npc1 and frame:
                candidate = self._get_content_excel_path(npc1, frame)
                if os.path.exists(candidate):
                    p = candidate
                else:
                    dir_root = self.config_manager.get_npc_output_directory()
                    content_dir = os.path.join(dir_root, (npc1 or 'unknown'), 'content')
                    name = getattr(frame, '_name', '')
                    orig = self._get_orig_base_from_name(name)
                    prefix = f"{orig}-{npc1}"
                    if os.path.isdir(content_dir):
                        for nm in os.listdir(content_dir):
                            low = nm.lower()
                            if low.startswith(prefix.lower()) and low.endswith('.xlsx'):
                                p = os.path.join(content_dir, nm)
                                break
            if not p or not os.path.exists(p):
                QMessageBox.critical(self, "错误", "未找到上一步生成的xlsx文件")
                print("[NPC内容存储] 未找到上一步生成的xlsx文件")
                if getattr(self, '_batch_generating', False):
                    self._batch_skip_current_and_advance('找不到xlsx文件')
                return
        wb = load_workbook(p)
        ts = QDateTime.currentDateTime().toString("yyyyMMddHHmmss")
        base = "image_prompt"
        name = base if base not in wb.sheetnames else f"{base}_{ts}"
        ws = wb.create_sheet(title=name)
        ws["A1"] = prompt_text
        wb.save(p)
        print(f"[NPC内容存储] 已写入生图提示词到文件: {p}，sheet: {name}")
        self._last_saved_content_path = p
        self._mark_selected_template('ci')

    def _on_npc_image_progress(self, message: str):
        self.npc_start_generate_btn.setText(f"生图中 {message}")

    def _on_npc_image_generated(self, data_dict: Dict, image_bytes: bytes):
        print("[NPC图片生成] 收到图片数据，开始保存与展示")
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        npc2 = getattr(self, '_npc2_current', '')
        dir_root = self.config_manager.get_npc_output_directory()
        out_dir = os.path.join(dir_root, (npc1 or 'unknown'), 'image')
        os.makedirs(out_dir, exist_ok=True)
        frame = getattr(self, '_selected_template_frame', None)
        name = getattr(frame, '_name', '') if frame is not None else ''
        stem = os.path.splitext(name)[0] if name else 'image'
        idx = stem.find('blank')
        prefix = stem[:idx] if idx >= 0 else stem
        prefix = re.sub(r'[\s\-_]+$', '', prefix)
        filename = f"{prefix}_{npc1}-{npc2}.jpg"
        save_path = os.path.join(out_dir, filename)
        working_bytes = image_bytes
        with Image.open(io.BytesIO(working_bytes)) as im:
            sw, sh = im.size
            width, height = 720, 720
            if sw > 0 and sh > 0 and width > 0 and height > 0:
                s = max(width / sw, height / sh)
                nw = max(1, int(round(sw * s)))
                nh = max(1, int(round(sh * s)))
                tmp = im.resize((nw, nh), Image.Resampling.LANCZOS)
                l = max(0, (nw - width) // 2)
                t = max(0, (nh - height) // 2)
                r = l + width
                b = t + height
                out = tmp.crop((l, t, r, b))
            else:
                out = im
            if out.mode != 'RGB':
                out = out.convert('RGB')
            buf = io.BytesIO()
            out.save(buf, format='JPEG', quality=90)
            working_bytes = buf.getvalue()
        working_bytes = add_watermark(working_bytes)
        with open(save_path, 'wb') as f:
            f.write(working_bytes)
        print(f"[NPC图片生成] 已保存图片: {save_path}")
        self._mark_selected_template('tu')
        pix = QPixmap()
        pix.loadFromData(working_bytes)
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.stop_loading()
            self.npc_image_container.set_original_pixmap(pix)

    def _on_npc_image_error(self, error_message: str, data_dict: Dict):
        f = getattr(self, '_selected_template_frame', None)
        if f is not None:
            f.mark_step_failed('tu')
        print(f"[NPC图片生成] 失败: {error_message}")
        detail = None
        full_body = None
        if isinstance(data_dict, dict):
            detail = data_dict.get('error_detail')
            full_body = data_dict.get('error_full_body')
        if detail:
            print(f"[NPC图片生成] 详细错误: {detail}")
        if full_body:
            print("[NPC图片生成] 返回体:\n" + str(full_body))
        popup_msg = error_message
        extra_parts = []
        if detail:
            extra_parts.append(f"详细错误:\n{detail}")
        if full_body:
            extra_parts.append(f"返回体:\n" + str(full_body))
        if extra_parts:
            popup_msg = popup_msg + "\n\n" + "\n\n".join(extra_parts)
        QMessageBox.critical(self, "生成失败", popup_msg)
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.stop_loading()
            self.npc_image_container.setText("生成失败")
        self._reset_chat_state_npc()
        if getattr(self, '_batch_generating', False):
            self._batch_skip_current_and_advance('生图失败')

    def _on_npc_image_generation_finished(self):
        print("[NPC图片生成] 完成，解锁控件")
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.stop_loading()
        self._reset_chat_state_npc()
        if getattr(self, '_batch_generating', False):
            self._batch_index = int(getattr(self, '_batch_index', 0)) + 1
            self._advance_batch_to_next_template()

    def _ensure_image_conversation_and_send_image(self, message: str):
        f = getattr(self, '_selected_template_frame', None)
        if f is not None:
            f.mark_step_in_progress('ci')
        conv_id = self.config_manager.get_coze_image_conversation_dialog_id()
        if not conv_id:
            api_base = self.config_manager.get_coze_api()
            token = self.config_manager.get_coze_token()
            if not api_base or not token:
                print('配置错误：缺少Coze API或Token')
                return
            self.coze_client = CozeAPIClient(api_base, token)
            self.current_worker = CozeWorkerThread(self.coze_client, 'create_conversation')
            self.current_worker.conversation_created.connect(lambda cid: self._on_image_conversation_created(cid, message))
            self.current_worker.conversation_failed.connect(self._on_image_conversation_failed)
            self.npc_start_generate_btn.setText('创建会话中...')
            self.npc_start_generate_btn.setEnabled(False)
            print('正在创建图片对话ID...')
            self.current_worker.start()
            return
        self.current_conversation_id = conv_id
        bot_id = self.config_manager.get_coze_duoki_image_bot_id()
        self._send_message(message=message, target='npc_image', bot_id_override=bot_id)

    def _on_image_conversation_created(self, conversation_id: str, message: str):
        print(f'图片对话ID创建成功: {conversation_id}')
        self.config_manager.set_coze_image_conversation_dialog_id(conversation_id)
        self.current_conversation_id = conversation_id
        bot_id = self.config_manager.get_coze_duoki_image_bot_id()
        self._send_message(message=message, target='npc_image', bot_id_override=bot_id)

    def _on_image_conversation_failed(self, error_msg: str):
        print(f'图片对话ID创建失败: {error_msg}')
        self.npc_start_generate_btn.setText('开始生成')
        self.npc_start_generate_btn.setEnabled(True)
        if getattr(self, '_batch_generating', False):
            self._batch_skip_current_and_advance('图片会话创建失败')
    
    def _abort_stream_chat(self):
        """
        严格的流式对话中止流程：
        1. 发送取消请求
        2. 等待取消状态验证
        3. 只有确认取消成功后才重置UI状态
        """
        # 更新按钮状态，显示正在取消
        self.send_btn.setText("取消中...")
        self.send_btn.setEnabled(False)
        self.output_text.append("\n正在取消对话...")
        
        # 保存当前worker的引用，避免在操作过程中被其他线程修改
        worker = self.current_worker
        if worker and worker.isRunning():
            try:
                # 调用严格的取消流程，不立即重置UI状态
                worker.cancel_stream_chat()
                # 取消线程
                worker.cancel()
            except Exception as e:
                print(f"中止流式对话时发生错误: {e}")
                # 如果发生异常，强制重置状态
                self._reset_chat_state()
        else:
            # 如果没有运行的worker，直接重置状态
            self._reset_chat_state()
    
    def _abort_non_stream_chat(self):
        """中止非流式对话"""
        # 保存当前worker的引用，避免在操作过程中被其他线程修改
        worker = self.current_worker
        if worker and worker.isRunning():
            try:
                # 先尝试调用API取消非流式对话
                worker.cancel_non_stream_chat()
                # 然后取消线程
                worker.cancel()
            except Exception as e:
                print(f"中止非流式对话时发生错误: {e}")
        
        self._reset_chat_state()
    
    def _on_chat_id_received(self, chat_id: str):
        """收到chat_id后，将按钮切换为中止状态"""
        print(f"=== UI收到chat_id: {chat_id}，切换按钮为中止状态 ===")
        
        if self.is_chatting and not self._chat_id_received:
            self._chat_id_received = True
            self.send_btn.setText("中止")
            self.send_btn.setEnabled(True)
            self.output_text.clear()
            self.output_text.append("正在生成回复...\n")
            print("按钮已切换为中止状态，现在可以中止对话")
    
    def _on_cancel_status_verified(self, cancel_result: dict):
        """
        处理严格的取消状态验证结果
        只有确认取消成功（status=canceled）后才重置UI状态
        """
        print(f"=== 收到取消状态验证结果: {cancel_result} ===")
        
        if cancel_result['success'] and cancel_result['status'] == 'canceled':
            # 取消成功，重置UI状态
            print("取消成功，重置UI状态")
            # 使用统一的输出格式
            self.output_text.append(f"\n\n--- 流式对话已成功取消1 ---")
            # 显示toast提示
            Toast.info("流式对话已取消", self)
            self._reset_chat_state()
        else:
            # 取消失败，显示错误信息但不重置状态，保持"取消中..."状态
            error_msg = cancel_result.get('message', '取消失败')
            print(f"取消失败: {error_msg}")
            self.output_text.append(f"\n取消失败: {error_msg}")
            self.output_text.append("\n请等待对话自然结束或重试取消")
            
            # 重新启用中止按钮，允许用户重试
            self.send_btn.setText("中止")
            self.send_btn.setEnabled(True)
    
    def _reset_chat_state(self):
        """重置聊天状态"""
        self.is_chatting = False
        self.send_btn.setText("发送")
        self.send_btn.setEnabled(True)
        self.current_worker = None
        self._chat_id_received = False  # 重置chat_id接收标志
        
        # 解锁输入控件，恢复正常编辑状态
        self._unlock_input_controls()
    
    def _lock_input_controls(self):
        """锁定输入控件，防止在AI生成过程中被修改"""
        # 锁定主要输入框
        self.input_text.setReadOnly(True)
        
        # 锁定平台和智能体选择
        self.platform_combo.setEnabled(False)
        self.agent_combo.setEnabled(False)
        
        # 锁定会话相关控件
        self.session_btn.setEnabled(False)
        self.stream_checkbox.setEnabled(False)
        
        print("已锁定输入控件，防止在生成过程中被修改")
    
    def _unlock_input_controls(self):
        """解锁输入控件，恢复正常编辑状态"""
        # 解锁主要输入框
        self.input_text.setReadOnly(False)
        
        # 解锁平台和智能体选择
        self.platform_combo.setEnabled(True)
        self.agent_combo.setEnabled(True)
        
        # 解锁会话相关控件
        self.session_btn.setEnabled(True)
        self.stream_checkbox.setEnabled(True)
        
        print("已解锁输入控件，恢复正常编辑状态")
    
    def _check_user_name(self) -> bool:
        """检查用户名"""
        user_name = self.get_user_name()
        if not user_name:
            # 检查输入框是否可编辑（即是否处于手动输入状态）
            if hasattr(self, 'username_edit') and not self.username_edit.isReadOnly():
                Toast.warning("请先在用户名输入框中输入用户名", self)
                self.username_edit.setFocus()  # 将焦点设置到用户名输入框
            else:
                Toast.warning("没有获取到用户名，请稍等", self)
            return False
        return True
    
    def _on_conversation_created(self, conversation_id: str):
        """会话创建成功"""
        self.current_conversation_id = conversation_id
        self.session_id_edit.setText(conversation_id)
        
        # 保存到配置
        self.config_manager.set_coze_conversation_dialog_id(conversation_id)
        
        # 恢复按钮状态
        self.session_btn.setEnabled(True)
        self.session_btn.setText("创建会话")
        
        QMessageBox.information(self, "成功", f"会话创建成功\nID: {conversation_id}")
    
    def _on_conversation_failed(self, error_msg: str):
        """会话创建失败"""
        self.session_btn.setEnabled(True)
        self.session_btn.setText("创建会话")
        
        QMessageBox.critical(self, "创建失败", f"会话创建失败:\n{error_msg}")
    
    def _on_message_received(self, content: str):
        """接收到消息内容"""
        if self.is_chatting:
            # 如果是第一次接收消息，清空"正在生成回复..."
            current_text = self.output_text.toPlainText()
            
            if current_text == "正在生成回复...\n":
                self.output_text.clear()
            
            # 移动光标到文末，然后追加新内容
            cursor = self.output_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.output_text.setTextCursor(cursor)
            self.output_text.insertPlainText(content)
            
            # 确保滚动到底部
            cursor = self.output_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            self.output_text.setTextCursor(cursor)
    
    def _on_chat_finished(self, chat_data: Dict):
        """对话完成"""
        print(f"=== 对话完成 ===")
        
        self._reset_chat_state()
        
        # 获取输出内容
        output_content = self.output_text.toPlainText().strip()
        
        # 如果有输出内容，先解析再决定是否添加到历史记录
        if output_content:
            # 解析AI内容（只解析一次）
            parsed_data = self._parse_ai_content(output_content)
            
            # 如果解析成功，进行同步、格式化和添加到历史记录
            if parsed_data:
                # 先同步数据到飞书多维表格获取record_id（传入已解析的数据）
                self._sync_to_feishu_with_parsed_data(parsed_data)
                
                # 然后添加到历史记录（此时current_record_id已经设置）
                self._add_to_history(output_content)
                
                # 最后自动格式化内容（传入已解析的数据）
                self._format_content_with_parsed_data(parsed_data)
        
        # 对话完成后保留输入框内容，不清空
        full_content = chat_data.get('full_content', '')
        if full_content and full_content.strip():
            print("对话成功完成，保留输入框内容")
        else:
            print("对话完成但没有收到有效内容，保留输入框内容")
    
    def _on_chat_failed(self, error_msg: str):
        """对话失败"""
        print(f"=== 对话失败: {error_msg} ===")
        
        self._reset_chat_state()
        self.output_text.setPlainText(f"对话失败: {error_msg}")
        QMessageBox.critical(self, "对话失败", error_msg)
    
    def _on_stream_chat_cancelled(self, message: str):
        """处理流式对话被取消的信号"""
        print(f"流式对话被取消: {message}")
        
        # 保存取消消息，用于后续的UI更新
        self._cancel_message = message
        
        # 重置取消处理标志
        self._cancel_processed = False
        
        # 设置一个超时定时器，确保即使没有收到stream_aborted信号，UI也能重置
        # 这是为了处理流式连接已经结束但没有触发stream_aborted的情况
        if not hasattr(self, '_cancel_timeout_timer'):
            self._cancel_timeout_timer = QTimer()
            self._cancel_timeout_timer.setSingleShot(True)
            self._cancel_timeout_timer.timeout.connect(self._on_stream_cancel_timeout)
        
        # 500ms后如果还没有收到stream_aborted信号，就强制重置UI
        self._cancel_timeout_timer.start(500)
    
    def _on_non_stream_chat_cancelled(self, message: str):
        """处理非流式对话被取消的信号"""
        print(f"非流式对话被取消: {message}")
        
        # 首先更新输出文本框
        current_output = self.output_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            # 如果已经有输出内容，保留并添加取消提示
            self.output_text.append(f"\n\n--- {message} ---")
        else:
            # 如果没有输出内容，显示取消信息
            self.output_text.setPlainText(f"--- {message} ---")
        
        # 检查是否已经处理过取消操作（UI重置等）
        if getattr(self, '_cancel_processed', False):
            print(f"=== 非流式对话取消UI重置已处理，跳过重复处理 ===")
            return
            
        # 标记已处理
        self._cancel_processed = True
        
        # 直接重置UI状态（非流式不需要等待stream_aborted信号）
        self._reset_chat_state()
        
        # 显示toast提示
        Toast.info("非流式对话已取消", self)
    
    def _on_stream_cancel_timeout(self):
        """流式取消操作超时处理，强制重置UI状态"""
        print(f"=== 流式取消操作超时，强制重置UI状态 ===")
        
        # 获取取消消息
        cancel_message = getattr(self, '_cancel_message', '流式取消操作超时')
        
        # 首先更新输出文本框（这个必须执行）
        current_output = self.output_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            # 如果已经有输出内容，保留并添加取消提示
            self.output_text.append(f"\n\n--- {cancel_message} ---")
        else:
            # 如果没有输出内容，显示取消信息
            self.output_text.setPlainText(f"--- {cancel_message} ---")
        
        # 检查是否已经处理过取消操作
        if getattr(self, '_cancel_processed', False):
            print(f"=== 流式取消操作UI重置已处理，跳过超时处理 ===")
            return
            
        # 标记已处理，防止重复
        self._cancel_processed = True
        
        # 停止定时器
        if hasattr(self, '_cancel_timeout_timer'):
            self._cancel_timeout_timer.stop()
        
        # 直接重置UI状态
        self._reset_chat_state()
        
        # 显示toast提示
        Toast.info("流式对话已被手动中止", self)
        
        # 清理取消消息
        if hasattr(self, '_cancel_message'):
            delattr(self, '_cancel_message')
    
    def _on_stream_aborted(self, content: str):
        """
        流式连接被中止（旧逻辑，在严格取消流程中不应被触发）
        注意：此方法现在主要用于兼容性，新的严格取消流程使用_on_cancel_status_verified
        """
        print(f"=== [旧逻辑] 流式连接被中止，已接收内容长度: {len(content)} ===")
        print("[警告] 此方法在严格取消流程中不应被触发，请检查取消逻辑")
        
        # 首先更新输出文本框（移除重复的取消消息）
        current_output = self.output_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            # 如果已经有输出内容，添加调试信息
            self.output_text.append(f"\n\n[调试] 旧的流式中止逻辑被触发")
        else:
            # 如果没有输出内容，显示调试信息
            self.output_text.setPlainText(f"[调试] 旧的流式中止逻辑被触发（无内容）")
        
        # 检查是否已经处理过取消操作（UI重置等）
        if getattr(self, '_cancel_processed', False):
            print(f"=== 流式连接中止UI重置已处理，跳过重复处理 ===")
            return
            
        # 标记已处理
        self._cancel_processed = True
        
        # 取消超时定时器（如果存在）
        if hasattr(self, '_cancel_timeout_timer'):
            self._cancel_timeout_timer.stop()
        
        # 现在可以安全地重置UI状态
        self._reset_chat_state()
        
        # 显示调试toast提示
        Toast.info("[调试] 旧的流式中止逻辑", self)
        
        # 清理取消消息
        if hasattr(self, '_cancel_message'):
            delattr(self, '_cancel_message')
    
    def _on_non_stream_aborted(self, message: str):
        """
        非流式对话被中止（此方法在流式对话中不应被触发）
        注意：此方法仅用于非流式对话的取消处理
        """
        print(f"=== [非流式] 非流式对话被中止: {message} ===")
        
        # 首先更新输出文本框（移除重复的取消消息）
        current_output = self.output_text.toPlainText()
        if current_output and current_output != "正在生成回复...\n":
            # 如果已经有输出内容，添加调试信息
            self.output_text.append(f"\n\n[调试] 非流式对话中止逻辑被触发")
        else:
            # 如果没有输出内容，显示调试信息
            self.output_text.setPlainText(f"[调试] 非流式对话中止逻辑被触发")
        
        # 检查是否已经处理过取消操作（UI重置等）
        if getattr(self, '_cancel_processed', False):
            print(f"=== 非流式对话中止UI重置已处理，跳过重复处理 ===")
            return
            
        # 标记已处理
        self._cancel_processed = True
        
        # 取消超时定时器（如果存在）
        if hasattr(self, '_cancel_timeout_timer'):
            self._cancel_timeout_timer.stop()
        
        # 现在可以安全地重置UI状态
        self._reset_chat_state()
        
        # 显示调试toast提示
        Toast.info("[调试] 非流式对话中止逻辑", self)
    
    def _format_content(self):
        """格式化AI输出内容"""
        try:
            content = self.output_text.toPlainText().strip()
            if not content:
                QMessageBox.warning(self, "格式化错误", "没有内容可以格式化")
                return
            
            # 解析五个段落
            parsed_data = self._parse_ai_content(content)
            if not parsed_data:
                return
            
            # 使用解析后的数据进行格式化
            self._format_content_with_parsed_data(parsed_data)
            
        except Exception as e:
            error_msg = f"格式化过程中发生错误: {str(e)}"
            self.output_text.append(f"\n\n{error_msg}")
            QMessageBox.critical(self, "格式化错误", error_msg)
    
    def _format_content_with_parsed_data(self, parsed_data: dict):
        """使用已解析的数据进行格式化"""
        try:
            # 创建Tab页签和数据表格
            self._create_data_tabs(parsed_data)
            
            # 隐藏格式化按钮，显示Tab容器
            self.format_btn.setVisible(False)
            self.data_tab_widget.setVisible(True)
            
        except Exception as e:
            error_msg = f"格式化过程中发生错误: {str(e)}"
            self.output_text.append(f"\n\n{error_msg}")
            QMessageBox.critical(self, "格式化错误", error_msg)
    
    def _parse_ai_content(self, content: str) -> dict:
        """解析AI返回的内容，提取五个段落"""
        try:
            # 按行分割内容
            lines = content.split('\n')
            
            # 查找所有[xxx]开头的行
            section_starts = []
            for i, line in enumerate(lines):
                if re.match(r'^\[.*\]', line.strip()):
                    section_starts.append(i)
            
            # 查找所有"评价："开头的行
            evaluation_lines = []
            for i, line in enumerate(lines):
                if line.strip().startswith('评价：'):
                    evaluation_lines.append(i)
            
            # 提取段落
            sections = {}
            for i, start_idx in enumerate(section_starts):
                # 获取段落名称
                section_name = re.match(r'^\[(.*)\]', lines[start_idx].strip()).group(1)
                
                # 找到对应的评价行
                end_idx = len(lines)
                for eval_idx in evaluation_lines:
                    if eval_idx > start_idx:
                        end_idx = eval_idx
                        break
                
                # 提取段落内容（不包括[xxx]行、评价行和关键字开头的行）
                section_lines = []
                for j in range(start_idx + 1, end_idx):
                    line = lines[j].strip()
                    if line and not line.startswith('评价：') and not line.startswith('关键字'):
                        section_lines.append(line)
                
                # 解析每一行的角色和对白
                parsed_lines = []
                for line in section_lines:
                    char_dialog = self._parse_line_to_char_dialog(line)
                    if char_dialog:
                        parsed_lines.append(char_dialog)
                
                if parsed_lines:
                    sections[section_name] = parsed_lines
            
            # 检查是否有五个段落
            if len(sections) != 5:
                QMessageBox.warning(self, "格式问题", f"格式问题，内容可能未完整解析。找到{len(sections)}个段落，期望5个段落")
                return None
            
            return sections
            
        except Exception as e:
            QMessageBox.critical(self, "解析错误", f"内容解析失败: {str(e)}")
            return None
    
    def _parse_line_to_char_dialog(self, line: str) -> dict:
        """解析单行内容为角色和对白"""
        try:
            # 查找第一个冒号或空格
            colon_idx = -1
            space_idx = -1
            
            # 查找全角冒号
            if '：' in line:
                colon_idx = line.index('：')
            # 查找半角冒号
            elif ':' in line:
                colon_idx = line.index(':')
            # 查找半角空格
            elif ' ' in line:
                space_idx = line.index(' ')
            
            # 确定分割位置
            split_idx = -1
            if colon_idx != -1:
                split_idx = colon_idx + 1
            elif space_idx != -1:
                split_idx = space_idx + 1
            
            if split_idx == -1:
                return None
            
            # 分割角色名和对白
            char_name = line[:split_idx-1].strip()
            dialog = line[split_idx:].strip()
            
            if char_name and dialog:
                return {'char': char_name, 'dialog': dialog}
            else:
                return None
                
        except Exception as e:
            print(f"解析行失败: {line}, 错误: {str(e)}")
            return None
    
    def _create_data_tabs(self, parsed_data: dict):
        """创建Tab页签和数据表格"""
        # 清空现有的Tab
        self.data_tab_widget.clear()
        
        # 重置上报状态（重新生成内容）
        self.tab_upload_status.clear()
        print("重新生成内容，已重置所有tab的上报状态")
        
        # 定义Tab名称映射
        tab_names = {
            '方案一': '方案一',
            '方案二': '方案二', 
            '方案三': '方案三',
            '方案四': '方案四',
            '特别方案': '特别方案'
        }
        
        # 为每个段落创建Tab
        for section_key, section_data in parsed_data.items():
            # 确定Tab名称
            tab_name = tab_names.get(section_key, section_key)
            
            # 创建表格
            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels(['角色名', '对白内容'])
            
            # 设置表格数据
            table.setRowCount(len(section_data))
            for row, item in enumerate(section_data):
                table.setItem(row, 0, QTableWidgetItem(item['char']))
                table.setItem(row, 1, QTableWidgetItem(item['dialog']))
            
            # 设置表格样式
            table.horizontalHeader().setStretchLastSection(True)
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            table.setAlternatingRowColors(True)
            
            # 修改选择模式为支持多单元格选中
            table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
            table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
            
            # 添加复制快捷键
            copy_shortcut = QShortcut(QKeySequence.StandardKey.Copy, table)
            copy_shortcut.activated.connect(lambda t=table: self._copy_table_selection(t))
            
            # 添加到Tab容器
            self.data_tab_widget.addTab(table, tab_name)
    
    def _get_current_tab_index(self) -> Optional[int]:
        """获取当前选中的tab索引"""
        if hasattr(self, 'data_tab_widget') and self.data_tab_widget:
            return self.data_tab_widget.currentIndex()
        return None
    
    def _on_tab_changed(self, index: int):
        """tab切换事件处理，重置上报状态"""
        # 重置所有tab的上报状态
        self.tab_upload_status.clear()
    
    def _copy_table_selection(self, table: QTableWidget):
        """复制表格选中的内容到剪贴板"""
        try:
            # 获取选中的项目
            selected_items = table.selectedItems()
            if not selected_items:
                return
            
            # 获取选中区域的行列范围
            selected_ranges = table.selectedRanges()
            if not selected_ranges:
                return
            
            # 处理多个选中区域
            clipboard_data = []
            
            for selected_range in selected_ranges:
                # 获取选中区域的边界
                top_row = selected_range.topRow()
                bottom_row = selected_range.bottomRow()
                left_col = selected_range.leftColumn()
                right_col = selected_range.rightColumn()
                
                # 按行列顺序提取数据
                range_data = []
                for row in range(top_row, bottom_row + 1):
                    row_data = []
                    for col in range(left_col, right_col + 1):
                        item = table.item(row, col)
                        cell_text = item.text() if item else ""
                        row_data.append(cell_text)
                    range_data.append("\t".join(row_data))
                
                clipboard_data.extend(range_data)
            
            # 将数据复制到剪贴板
            clipboard_text = "\n".join(clipboard_data)
            clipboard = QApplication.clipboard()
            clipboard.setText(clipboard_text)
            
            print(f"已复制 {len(clipboard_data)} 个单元格到剪贴板")
            
            # 获取当前tab索引并检查是否已上报
            current_tab_index = self._get_current_tab_index()
            if current_tab_index is not None:
                # 检查是否已经上报过
                if self.tab_upload_status.get(current_tab_index, False):
                    print("该tab页已经上报过，跳过飞书更新")
                    return
                
                # 更新飞书记录的selected字段
                success = self._update_feishu_selected_field(table)
                if success:
                    # 标记该tab页已上报
                    self.tab_upload_status[current_tab_index] = True
                    print(f"Tab {current_tab_index} 已标记为已上报")
            
        except Exception as e:
            print(f"复制失败: {str(e)}")
    
    def _update_feishu_selected_field(self, table: QTableWidget) -> bool:
        """更新飞书记录的selected字段"""
        try:
            # 检查是否有record_id
            if not self.current_record_id:
                print("没有可用的record_id，跳过selected字段更新")
                
                # 提供重新同步的选项
                reply = QMessageBox.question(
                    self, 
                    "缺少记录ID", 
                    "当前历史记录没有关联的飞书记录ID。\n\n是否要重新同步当前内容到飞书？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    # 重新同步当前内容
                    current_content = self.output_text.toPlainText().strip()
                    if current_content:
                        parsed_data = self._parse_ai_content(current_content)
                        if parsed_data:
                            self._sync_to_feishu_with_parsed_data(parsed_data)
                            # 更新当前历史记录项的record_id
                            self._update_current_history_record_id()
                            # 如果同步成功，继续更新selected字段
                            if self.current_record_id:
                                return self._update_feishu_selected_field(table)
                        else:
                            QMessageBox.warning(self, "同步失败", "无法解析当前内容，同步失败")
                    else:
                        QMessageBox.warning(self, "同步失败", "当前没有可同步的内容")
                
                return False
            
            # 获取当前tab索引
            current_tab_index = -1
            for i in range(self.data_tab_widget.count()):
                if self.data_tab_widget.widget(i) == table:
                    current_tab_index = i + 1  # tab索引从1开始
                    break
            
            if current_tab_index == -1:
                print("无法确定当前tab索引")
                return False
            
            # 导入飞书同步模块并更新记录
            from .feishu_sync import FeishuBitableSync
            
            sync_client = FeishuBitableSync()
            success = sync_client.update_record(self.current_record_id, current_tab_index)
            
            if success:
                print(f"已更新飞书记录selected字段: record_id={self.current_record_id}, selected={current_tab_index}")
                return True
            else:
                print("飞书记录更新失败")
                return False
                
        except Exception as e:
            print(f"更新飞书selected字段时发生异常: {str(e)}")
            return False
    
    def _create_history_panel(self) -> QWidget:
        """创建历史记录面板"""
        panel = QWidget()
        
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # 标题
        title_label = QLabel("历史记录")
        title_label.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 历史记录列表
        self.history_list = QListWidget()
        self.history_list.setFont(QFont("Microsoft YaHei", 9))
        self.history_list.itemClicked.connect(self._on_history_item_clicked)
        layout.addWidget(self.history_list)
        
        # 初始化历史记录存储
        self.history_data = []
        
        return panel
    
    def _extract_characters_and_keyword_from_input(self) -> tuple:
        """从用户输入内容中提取角色和关键字信息"""
        try:
            # 获取用户输入内容
            input_content = self.input_text.toPlainText().strip()
            if not input_content:
                return "", "", ""
            
            lines = input_content.split('\n')
            npc1_name = ""
            npc2_name = ""
            keyword = ""
            
            # 从文末往前一行一行查找
            for line in reversed(lines):
                line = line.strip()
                if not line:
                    continue
                
                # 查找npc1（支持多种格式：npc1:、npc1：、npc1 多个空格、npc1 : 等）
                if line.startswith('npc1') and npc1_name == "":
                    # 使用正则表达式匹配：npc1 后面可能有多个空格，然后可能有冒号，再可能有多个空格，最后是名字
                    match = re.match(r'npc1\s*[：:]*\s*(.+)', line)
                    if match:
                        npc1_name = match.group(1).strip()
                
                # 查找npc2（支持多种格式）
                elif line.startswith('npc2') and npc2_name == "":
                    match = re.match(r'npc2\s*[：:]*\s*(.+)', line)
                    if match:
                        npc2_name = match.group(1).strip()
                
                # 查找关键字（支持多种格式）
                elif line.startswith('关键字') and keyword == "":
                    match = re.match(r'关键字\s*[：:]*\s*(.+)', line)
                    if match:
                        keyword = match.group(1).strip()
            
            return npc1_name, npc2_name, keyword
            
        except Exception as e:
            print(f"提取角色和关键字失败: {str(e)}")
            return "", "", ""
    
    def _add_to_history(self, content: str):
        """添加内容到历史记录"""
        try:
            # 从用户输入中提取角色和关键字
            npc1, npc2, keyword = self._extract_characters_and_keyword_from_input()
            
            # 创建历史记录项名称
            item_name = f"{npc1} {npc2} {keyword}"
            
            # 保存到历史数据，包含record_id（插入到开头，保持新到旧的顺序）
            history_item = {
                'name': item_name,
                'content': content,
                'timestamp': QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss"),
                'record_id': self.current_record_id  # 保存当前的record_id
            }
            self.history_data.insert(0, history_item)
            
            # 添加到列表显示（插入到顶部）
            list_item = QListWidgetItem(item_name)
            list_item.setToolTip(f"创建时间: {history_item['timestamp']}\nRecord ID: {self.current_record_id or '无'}")
            self.history_list.insertItem(0, list_item)
            
            # 选中新添加的项目（第一项）
            self.history_list.setCurrentRow(0)
            
            print(f"已添加历史记录: {item_name}, record_id: {self.current_record_id}")
            
        except Exception as e:
            print(f"添加历史记录失败: {str(e)}")
    
    def _on_history_item_clicked(self, item: QListWidgetItem):
        """处理历史记录项点击事件"""
        try:
            # 获取点击的项的索引
            row = self.history_list.row(item)
            if 0 <= row < len(self.history_data):
                # 获取对应的历史数据
                history_item = self.history_data[row]
                content = history_item['content']
                
                # 将内容加载到输出框
                self.output_text.setPlainText(content)
                
                # 更新当前的record_id
                self.current_record_id = history_item.get('record_id')
                
                # 自动格式化
                self._format_content()
                
                print(f"已加载历史记录: {history_item['name']}, record_id: {self.current_record_id}")
                
        except Exception as e:
            print(f"加载历史记录失败: {str(e)}")
    
    def _update_current_history_record_id(self):
        """更新当前选中的历史记录项的record_id"""
        try:
            current_item = self.history_list.currentItem()
            if current_item and self.current_record_id:
                # 获取当前选中项的索引
                row = self.history_list.row(current_item)
                if 0 <= row < len(self.history_data):
                    # 更新历史数据中的record_id
                    self.history_data[row]['record_id'] = self.current_record_id
                    
                    # 更新列表项的tooltip
                    history_item = self.history_data[row]
                    current_item.setToolTip(f"创建时间: {history_item['timestamp']}\nRecord ID: {self.current_record_id}")
                    
                    print(f"已更新历史记录项的record_id: {history_item['name']} -> {self.current_record_id}")
        except Exception as e:
            print(f"更新历史记录record_id失败: {str(e)}")
    
    def _sync_to_feishu(self, content: str):
        """同步AI生成内容到飞书多维表格"""
        try:
            # 解析AI内容为五个方案
            parsed_data = self._parse_ai_content(content)
            if not parsed_data:
                # 如果解析失败，将原始内容放在content_1中
                print("AI内容解析失败，将原始内容放在content_1中")
                parsed_data = {"原始内容": [{"char": "", "dialog": content}]}
            
            # 使用解析后的数据进行同步
            self._sync_to_feishu_with_parsed_data(parsed_data)
            
        except Exception as e:
            print(f"同步到飞书失败: {str(e)}")
            # 不显示错误弹窗，避免影响用户体验，只在控制台记录错误
    
    def _sync_to_feishu_with_parsed_data(self, parsed_data: dict):
        """使用已解析的数据同步到飞书多维表格"""
        try:
            # 提取npc1、npc2和关键字
            npc1, npc2, keywords = self._extract_characters_and_keyword_from_input()
            
            # 获取当前用户名
            user_name = self.get_user_name()
            
            # 初始化五个方案内容
            content_1 = content_2 = content_3 = content_4 = content_5 = ""
            
            if parsed_data and isinstance(parsed_data, dict):
                # 将解析后的段落转换为文本格式
                section_names = list(parsed_data.keys())
                
                for i, section_name in enumerate(section_names[:5]):  # 最多取前5个段落
                    section_data = parsed_data[section_name]
                    
                    # 将段落数据转换为文本（去掉方括号标记）
                    section_text = ""
                    for line_data in section_data:
                        if isinstance(line_data, dict) and 'char' in line_data and 'dialog' in line_data:
                            section_text += f"{line_data['char']}：{line_data['dialog']}\n"
                    
                    # 分配到对应的content字段
                    if i == 0:
                        content_1 = section_text.strip()
                    elif i == 1:
                        content_2 = section_text.strip()
                    elif i == 2:
                        content_3 = section_text.strip()
                    elif i == 3:
                        content_4 = section_text.strip()
                    elif i == 4:
                        content_5 = section_text.strip()
            
            # 调用飞书同步函数并保存record_id
            record_id = sync_ai_content_to_feishu(
                content_1=content_1,
                content_2=content_2,
                content_3=content_3,
                content_4=content_4,
                content_5=content_5,
                npc1=npc1,
                npc2=npc2,
                keywords=keywords,
                user=user_name
            )
            
            # 保存record_id
            if record_id:
                self.current_record_id = record_id
                print(f"已同步数据到飞书，record_id: {record_id}")
            else:
                print("飞书同步失败")
                self.current_record_id = None
            
            print(f"同步数据: npc1={npc1}, npc2={npc2}, keywords={keywords}, user={user_name}")
            print(f"方案数量: content_1={'有内容' if content_1 else '空'}, content_2={'有内容' if content_2 else '空'}, content_3={'有内容' if content_3 else '空'}, content_4={'有内容' if content_4 else '空'}, content_5={'有内容' if content_5 else '空'}")
            
        except Exception as e:
            print(f"同步到飞书失败: {str(e)}")
            # 不显示错误弹窗，避免影响用户体验，只在控制台记录错误
    
    def closeEvent(self, event):
        """组件关闭时的清理工作"""
        self._cleanup_threads()
        super().closeEvent(event)
    
    def _cleanup_threads(self):
        """清理所有活跃的线程"""
        try:
            # 清理当前工作线程
            if self.current_worker and self.current_worker.isRunning():
                print("正在清理CozeWorkerThread...")
                self.current_worker.cancel()
                # 等待线程结束，最多等待3秒
                if not self.current_worker.wait(3000):
                    print("警告: CozeWorkerThread未能在3秒内正常结束")
                    self.current_worker.terminate()
                    self.current_worker.wait(1000)
                self.current_worker = None
            
            # 清理用户名获取线程
            if self.username_worker and self.username_worker.isRunning():
                print("正在清理UsernameFetchWorker...")
                self.username_worker.quit()
                if not self.username_worker.wait(2000):
                    print("警告: UsernameFetchWorker未能在2秒内正常结束")
                    self.username_worker.terminate()
                    self.username_worker.wait(1000)
                self.username_worker = None
                
        except Exception as e:
            print(f"清理线程时发生错误: {str(e)}")
    
    def __del__(self):
        """析构函数，确保线程被正确清理"""
        self._cleanup_threads()
    def _mark_selected_template(self, part: str):
        f = getattr(self, '_selected_template_frame', None)
        if getattr(self, '_batch_generating', False):
            if part == 'wen':
                f = getattr(self, '_batch_wen_current_frame', f)
            elif part == 'ci':
                f = getattr(self, '_batch_ci_current_frame', f)
        if not f:
            return
        if part == 'wen':
            f.set_wen_done()
        elif part == 'ci':
            f.set_ci_done()
        elif part == 'tu':
            f.set_tu_done()
        self._update_npc_template_count_label()

    def _reset_template_frame_status(self, frame, reset_wen=False, reset_ci=False, reset_tu=False, reset_all=False):
        fr = frame or getattr(self, '_selected_template_frame', None)
        if not fr:
            return
        if reset_all:
            reset_wen = reset_ci = reset_tu = True
        if reset_wen:
            fr._done_wen = False
            if fr.lab_wen is not None:
                fr.lab_wen.setStyleSheet("background-color: transparent; color: #888888; border-radius: 2px;")
        if reset_ci:
            fr._done_ci = False
            if fr.lab_ci is not None:
                fr.lab_ci.setStyleSheet("background-color: transparent; color: #888888; border-radius: 2px;")
        if reset_tu:
            fr._done_tu = False
            if fr.lab_tu is not None:
                fr.lab_tu.setStyleSheet("background-color: transparent; color: #888888; border-radius: 2px;")
        fr._complete = False
        if fr._left_name_label is not None:
            fr._left_name_label.setStyleSheet("border: none;")
        fr._apply_frame_style()
        self._update_npc_template_count_label()

    def _compose_tpl_text_from_dialog_table(self) -> str:
        lines = []
        if hasattr(self, 'dialog_table') and self.dialog_table:
            rows = self.dialog_table.rowCount()
            for i in range(rows):
                nitem = self.dialog_table.item(i, 0)
                ditem = self.dialog_table.item(i, 1)
                npc = (nitem.text() if nitem else '').strip()
                dlg = (ditem.text() if ditem else '').strip()
                if npc and dlg:
                    lines.append(f"{npc}:{dlg}")
        return "\n".join(lines)

    def _on_template_btn_wen_clicked(self, frame):
        self._selected_template_frame = frame
        frame.set_selected(True)
        print("[角色生产线] btn_wen: 重置状态并从第一步开始生成文案")
        self._reset_template_frame_status(frame, reset_all=True)
        if hasattr(self, 'dialog_table') and self.dialog_table:
            self.dialog_table.setRowCount(0)
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.clear()
        self.on_npc_start_generate()

    def _on_template_btn_ci_clicked(self, frame):
        self._selected_template_frame = frame
        frame.set_selected(True)
        print("[角色生产线] btn_ci: 重置词/图状态，基于表格生成生图提示词")
        self._reset_template_frame_status(frame, reset_ci=True, reset_tu=True)
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            self.npc_prompt_text.clear()
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.clear()
        npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
        if npc1 == '请选择一个角色':
            npc1 = ''
        npc2 = getattr(self, '_npc2_current', '')
        tpl_text = self._compose_tpl_text_from_dialog_table()
        if not tpl_text:
            print("[角色生产线] 表格为空，无法生成生图提示词")
            return
        message = self._compose_npc_prompt_text(tpl_text, npc1, npc2)
        self._ensure_image_conversation_and_send_image(message)

    def _on_template_btn_tu_clicked(self, frame):
        self._selected_template_frame = frame
        frame.set_selected(True)
        print("[角色生产线] btn_tu: 重置图状态，使用当前提示词生图")
        self._reset_template_frame_status(frame, reset_tu=True)
        if hasattr(self, 'npc_image_container') and self.npc_image_container:
            self.npc_image_container.clear()
        prompt = ''
        if hasattr(self, 'npc_prompt_text') and self.npc_prompt_text:
            prompt = self.npc_prompt_text.toPlainText().strip()
        if not prompt:
            npc1 = getattr(self, '_npc1_current', '') or (self.npc_role_combo.currentText().strip() if hasattr(self, 'npc_role_combo') else '')
            if npc1 == '请选择一个角色':
                npc1 = ''
            frame_obj = getattr(self, '_selected_template_frame', None)
            if frame_obj is not None:
                excel_path = self._get_content_excel_path(npc1, frame_obj)
                if os.path.exists(excel_path):
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path, read_only=True)
                    if 'image_prompt' in wb.sheetnames:
                        ws = wb['image_prompt']
                        prompt = str(ws['A1'].value or '').strip()
        if not prompt:
            print("[角色生产线] 当前没有提示词，无法进行生图")
            return
        self._start_npc_image_generation_from_prompt(prompt)
