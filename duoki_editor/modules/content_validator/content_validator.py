"""
内容校验模块
用于校验Excel文件内容并提供修复建议
"""

import sys
import io
from contextlib import redirect_stdout, redirect_stderr
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                            QLabel, QFileDialog, QTableView, QHeaderView, QMessageBox,
                            QListWidget, QSplitter, QAbstractItemView, QListWidgetItem,
                            QCheckBox, QMenu, QApplication, QDialog, QDialogButtonBox, QTextEdit,
                            QLineEdit, QGridLayout, QGroupBox, QTabWidget, QRadioButton, QButtonGroup,
                            QTableWidget, QTableWidgetItem, QComboBox)
from PyQt6.QtCore import Qt, QAbstractTableModel, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QAction, QBrush, QPixmap
from PyQt6.QtNetwork import QNetworkAccessManager
import pandas as pd
import os
import re
import logging
import openpyxl
import requests
from typing import List, Dict, Any
from configparser import ConfigParser
from duoki_editor.utils.config_manager import ConfigManager
from duoki_editor.utils.constants_loader import get_app_name_map, get_npc_id_map
from duoki_editor.core.character_table_manager import CharacterTableManager
from duoki_editor.utils.worker_thread import ThreadPoolRunner
from duoki_editor.ui.toast import Toast


class TextRedirector:
    """将文本输出重定向到QTextEdit控件的类"""
    def __init__(self, text_widget):
        self.text_widget = text_widget
        
    def write(self, text):
        if text.strip():  # 只处理非空文本
            # 在主线程中更新UI
            self.text_widget.append(text.strip())
            # 滚动到底部
            self.text_widget.verticalScrollBar().setValue(
                self.text_widget.verticalScrollBar().maximum()
            )
            # 刷新界面
            QApplication.processEvents()
    
    def flush(self):
        pass


class LoggingHandler(logging.Handler):
    """自定义的logging处理器，将日志输出到QTextEdit控件"""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        
    def emit(self, record):
        msg = self.format(record)
        
        # 根据日志级别和内容添加颜色
        if record.levelno >= logging.ERROR:
            # 错误信息用红色
            colored_msg = f'<span style="color: red;">{msg}</span>'
        elif record.levelno >= logging.WARNING:
            # 警告信息用橙色
            colored_msg = f'<span style="color: orange;">{msg}</span>'
        elif "成功" in msg or "✅" in msg or "Successfully" in msg:
            # 成功信息用绿色
            colored_msg = f'<span style="color: green;">{msg}</span>'
        elif "开始" in msg or "Starting" in msg or "Processing" in msg or "Merging" in msg:
            # 重要进度信息用橙色
            colored_msg = f'<span style="color: orange;">{msg}</span>'
        else:
            # 普通信息保持默认颜色
            colored_msg = msg
        
        # 在主线程中更新UI
        self.text_widget.append(colored_msg)
        # 滚动到底部
        self.text_widget.verticalScrollBar().setValue(
            self.text_widget.verticalScrollBar().maximum()
        )
        # 刷新界面
        QApplication.processEvents()


class ValidationTableModel(QAbstractTableModel):
    """校验结果表格模型"""
    
    def __init__(self, data=None):
        super().__init__()
        self._data = data if data is not None else pd.DataFrame()
        self._headers = ["选择", "title", "row", "type", "原文", "修复建议"]
        self._checkboxes = []  # 存储checkbox状态
        self.select_all_state = True  # 全选状态
    
    def rowCount(self, parent=None):
        return len(self._data)
    
    def columnCount(self, parent=None):
        return len(self._headers)
    
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
            
        if role == Qt.ItemDataRole.DisplayRole:
            if index.column() == 0:  # 选择列
                return ""  # checkbox列不显示文本
            else:
                # 获取对应列的数据，列索引对应：1=title, 2=row, 3=type, 4=original_text, 5=suggestion
                column_names = ['title', 'row', 'type', 'original_text', 'suggestion']
                if index.column() - 1 < len(column_names):
                    column_name = column_names[index.column() - 1]
                    if hasattr(self._data, 'iloc') and not self._data.empty:
                        value = self._data.iloc[index.row()][column_name]
                        return str(value) if pd.notna(value) else ""
                return ""
        elif role == Qt.ItemDataRole.CheckStateRole:
            if index.column() == 0:  # 选择列
                if index.row() < len(self._checkboxes):
                    return Qt.CheckState.Checked if self._checkboxes[index.row()] else Qt.CheckState.Unchecked
                else:
                    return Qt.CheckState.Checked  # 默认选中
        elif role == Qt.ItemDataRole.ToolTipRole:
            # 为type列（第3列）添加tooltip，显示命中的错误条件文本
            if index.column() == 3:  # type列
                if hasattr(self._data, 'iloc') and not self._data.empty:
                    type_value = str(self._data.iloc[index.row()]['type'])
                    tooltip_text = self._generate_tooltip_for_type(type_value)
                    # 直接返回tooltip文本，让CSS样式控制宽度
                    return tooltip_text
            # 为原文列（第4列）添加tooltip，显示原文内容
            elif index.column() == 4:  # original_text列
                if hasattr(self._data, 'iloc') and not self._data.empty:
                    row_data = self._data.iloc[index.row()]
                    original_text = str(row_data['original_text'])
                    type_value = str(row_data['type'])
                    
                    # 如果是条件1，则显示特殊tooltip
                    if type_value == '1':
                        param1 = str(row_data['param1']) if 'param1' in row_data and pd.notna(row_data['param1']) else ""
                        tooltip = f"param1：{param1}\n当前文本：{original_text}"
                        return tooltip
                    # 否则返回原文内容作为tooltip
                    return original_text if pd.notna(original_text) else ""
            # 为修复建议列（第5列）添加tooltip，显示完整建议内容
            elif index.column() == 5:  # suggestion列
                if hasattr(self._data, 'iloc') and not self._data.empty:
                    suggestion = str(self._data.iloc[index.row()]['suggestion'])
                    # 只有当建议不为空时才返回tooltip
                    return suggestion if pd.notna(suggestion) and suggestion.strip() != "" else None
        elif role == Qt.ItemDataRole.ForegroundRole:
            # 为type列添加蓝色链接样式
            if index.column() == 3:  # type列
                return QColor(0, 100, 200)  # 蓝色
            # 为原文列添加橙色文本（如果命中条件1或已修复）
            elif index.column() == 4:  # original_text列
                if hasattr(self._data, 'iloc') and not self._data.empty:
                    row_data = self._data.iloc[index.row()]
                    type_value = str(row_data['type'])
                    # 检查是否有fixed标记
                    is_fixed = row_data.get('fixed', False) if hasattr(row_data, 'get') else False
                    if type_value == '1' or is_fixed:
                        return QColor('orange')  # 橙色
            # 为修复建议列添加橙色文本
            elif index.column() == 5:  # suggestion列
                return QColor('orange')  # 橙色
        elif role == Qt.ItemDataRole.FontRole:
            # 为type列添加下划线字体
            if index.column() == 3:  # type列
                font = QFont()
                font.setUnderline(True)
                return font
        
        return None
    
    def _generate_tooltip_for_type(self, type_value):
        """为type列的数字生成tooltip"""
        condition_descriptions = {
            '1': '条件1：英文相关变量内容与模板相比种类不匹配',
            '2': '条件2：不应出现中括号中只有数字的情况',
            '3': '条件3：内容中不应该出现冒号、波浪线及省略号',
            '4': '条件4：内容的首尾不应出现配对的单引号或反引号',
            '5': '条件5：知识点前后不应该出现双引号，知识点后不能有符号',
            '6': '条件6：内容中不应该出现连续标点符号',
            '7': '条件7：无方括号的{word_en}或{phrase_en}必须在纯英文语句中，{word_cn}或{phrase_cn}必须不在纯英文语句中',
            '8': '条件8：半角句号前后两句不应出现中英文混合',
            '9': '条件9：{words}两侧不能有中括号',
            '10': '条件10：中括号或大括号两侧的引号必须成对，不应出现反引号'
        }
        
        # 解析type值中的数字（可能是"1 2"这样的组合）
        numbers = type_value.split()
        tooltips = []
        
        for num in numbers:
            if num in condition_descriptions:
                tooltips.append(condition_descriptions[num])
        
        # 使用<br>标签进行换行，而不是\n
        return '<br>'.join(tooltips) if tooltips else type_value
    
    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid():
            return False
            
        if role == Qt.ItemDataRole.CheckStateRole and index.column() == 0:
            # 确保checkbox列表长度足够
            while len(self._checkboxes) <= index.row():
                self._checkboxes.append(True)
            
            self._checkboxes[index.row()] = (value == Qt.CheckState.Checked)
            self.dataChanged.emit(index, index, [role])
            return True
        elif role == Qt.ItemDataRole.EditRole and index.column() > 0:
            # 处理其他列的编辑（如修复建议列）
            column_names = ['title', 'row', 'type', 'original_text', 'suggestion']
            if index.column() - 1 < len(column_names):
                column_name = column_names[index.column() - 1]
                if hasattr(self._data, 'iloc') and not self._data.empty:
                    self._data.iloc[index.row(), self._data.columns.get_loc(column_name)] = value
                    self.dataChanged.emit(index, index, [role])
                    return True
            
        return False
    
    def flags(self, index):
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
            
        if index.column() == 0:  # 选择列
            return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable
        else:
            return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
    
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
            
        if orientation == Qt.Orientation.Horizontal:
            return self._headers[section]
            
        return str(section + 1)
    
    def load_data(self, data):
        """加载数据到模型"""
        self.beginResetModel()
        self._data = data if data is not None else pd.DataFrame()
        
        # 根据条件类型设置checkbox状态
        self._checkboxes = []
        if not self._data.empty:
            for index, row in self._data.iterrows():
                type_value = str(row.get('type', ''))
                # 如果命中条件1，取消勾选；其他条件默认勾选
                if re.search(r'\b1\b', type_value):
                    self._checkboxes.append(False)
                else:
                    self._checkboxes.append(True)
        
        self.endResetModel()
    
    def toggle_select_all(self):
        """切换全选状态"""
        self.select_all_state = not self.select_all_state
        self._checkboxes = [self.select_all_state] * len(self._data)
        # 通知视图更新
        if len(self._data) > 0:
            top_left = self.index(0, 0)
            bottom_right = self.index(len(self._data) - 1, 0)
            self.dataChanged.emit(top_left, bottom_right, [Qt.ItemDataRole.CheckStateRole])
    
    def get_selected_rows(self):
        """获取选中的行索引"""
        return [i for i, checked in enumerate(self._checkboxes) if checked]


class ContentValidator(QWidget):
    """内容校验主界面"""
    
    def __init__(self, data_manager=None, auth_manager=None):
        super().__init__()
        self.data_manager = data_manager
        self.auth_manager = auth_manager
        self.file_paths = []  # 存储完整文件路径
        self.validation_results = []  # 存储校验结果
        self.config_manager = ConfigManager()  # 初始化配置管理器
        self.current_file_index = None  # 当前选中的文件索引
        self.character_table_manager = CharacterTableManager()
        self.init_ui()
       
    def check_condition_1(self, text, param1_text):
        """条件1：判定内容对应的param1的文本中，包含"_en"的特殊格式{xxx}，例如{word_en}等，在当前内容中，种类是否齐备"""
        issues = []
        
        # 忽略特殊规则模式（不包含"_en"的特殊格式）
        ignore_patterns = ['{user_name}', '{npc1_name}', '{npc2_name}', '{npc1_name_origin}', '{npc2_name_origin}']
        
        # 如果文本和param1文本完全相同，则不需要检查
        if text == param1_text:
            return False, []
        
        # 查找所有包含"_en"的特殊格式{xxx}
        en_pattern = r'\{[a-zA-Z0-9_]*_en[a-zA-Z0-9_]*\}'
        
        # 在param1中查找所有包含"_en"的特殊格式
        param1_en_matches = re.findall(en_pattern, param1_text)
        # 过滤掉需要忽略的特殊模式
        param1_en_types = set()
        for match in param1_en_matches:
            if match not in ignore_patterns:
                param1_en_types.add(match)
        
        # 如果param1中没有包含"_en"的特殊格式，则不需要检查
        if not param1_en_types:
            return False, []
        
        # 在当前文本中查找所有包含"_en"的特殊格式
        text_en_matches = re.findall(en_pattern, text)
        # 过滤掉需要忽略的特殊模式
        text_en_types = set()
        for match in text_en_matches:
            if match not in ignore_patterns:
                text_en_types.add(match)
        
        # 检查当前文本中是否包含param1中所有类型的特殊格式
        missing_types = param1_en_types - text_en_types
        
        # 如果有缺失的类型，则命中条件1
        if missing_types:
            missing_types_str = ', '.join(missing_types)
            issues.append(f"内容中缺少param1中的特殊格式：{missing_types_str}")
            return True, issues
        
        # 如果所有类型都齐备，则不命中条件1
        return False, []

    def check_condition_2(self, text):
        """条件2：检查中括号加数字的内容"""
        issues = []
        
        # 查找[数字]模式
        pattern = r'\[\d+\]'
        matches = re.finditer(pattern, text)
        
        for match in matches:
            issues.append(f"发现中括号加数字：{match.group()}")
        
        return len(issues) > 0, issues

    def check_condition_3(self, text):
        """条件3：内容中不应该出现全角或半角冒号、波浪线、省略号、顿号"""
        issues = []
        
        # 检查是否包含全角冒号、半角冒号或波浪线
        colon_pattern = r'[:：~]'
        matches = re.findall(colon_pattern, text)
        
        if matches:
            issues.append(f"内容中包含不应出现的符号：{', '.join(matches)}，建议替换为全角逗号")
        
        # 检查是否包含省略号（……或…）
        ellipsis_pattern = r'[…]{1,2}'
        ellipsis_matches = re.findall(ellipsis_pattern, text)
        
        if ellipsis_matches:
            issues.append(f"内容中包含省略号：{', '.join(ellipsis_matches)}，建议替换为句号")
        
        # 检查是否包含顿号
        pause_pattern = r'、'
        pause_matches = re.findall(pause_pattern, text)
        
        if pause_matches:
            issues.append(f"内容中包含顿号：{', '.join(pause_matches)}，建议替换为逗号")
            
        return len(issues) > 0, issues

    def check_condition_4(self, text):
        """条件4：检查内容的首尾是否出现配对的单引号或反引号"""
        issues = []
        
        # 检查是否以单引号开头并以单引号结尾
        if text.startswith("'") and text.endswith("'"):
            # 提取不带引号的内容
            content_without_quotes = text[1:-1]
            issues.append(f"内容首尾出现配对的单引号，建议去掉：'{content_without_quotes}'")
        
        # 检查是否以两个单引号开头并以两个单引号结尾
        if text.startswith("''") and text.endswith("''"):
            # 提取不带引号的内容
            content_without_quotes = text[2:-2]
            issues.append(f"内容首尾出现配对的双单引号，建议去掉：''{content_without_quotes}''")
            
        # 检查是否有类似 '{word_en}' 的模式
        pattern = r"^'(\{[^}]+\})'$"
        matches = re.findall(pattern, text)
        if matches:
            for match in matches:
                issues.append(f"变量被单引号包围，建议去掉单引号：'{match}'，改为{match}")
        
        # 检查是否以反引号开头并以反引号结尾
        if text.startswith("`") and text.endswith("`"):
            # 提取不带反引号的内容
            content_without_backticks = text[1:-1]
            issues.append(f"内容首尾出现配对的反引号，建议去掉：`{content_without_backticks}`")
        
        # 检查是否有类似 `{word_en}` 的模式
        backtick_pattern = r"^`(\{[^}]+\})`$"
        backtick_matches = re.findall(backtick_pattern, text)
        if backtick_matches:
            for match in backtick_matches:
                issues.append(f"变量被反引号包围，建议去掉反引号：`{match}`，改为{match}")
        
        return len(issues) > 0, issues

    def check_condition_5(self, text):
        """条件5：知识点前后不应该出现双引号，方括号后不允许有标点符号"""
        issues = []
        
        # 检查知识点前后是否有双引号
        patterns = [
            r'"(\{word_en\})"',
            r'"(\[{word_en}\])"',
            r'"(\{phrase_en\})"',
            r'"(\[{phrase_en}\])"'
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text)
            if matches:
                for match in matches:
                    issues.append(f"知识点前后出现双引号：\"{match}\"，建议去掉双引号")
        
        # 检查方括号后面是否有标点符号
        # 匹配所有非字母、数字、汉字的符号，但排除[、{和各种空格
        bracket_after_pattern = r'\}\]([^\w\u4e00-\u9fff\[\{\s])'
        bracket_after_matches = re.findall(bracket_after_pattern, text)
        if bracket_after_matches:
            for match in bracket_after_matches:
                issues.append(f"方括号后出现标点符号：]{match}，建议去掉符号")
        
        # 检查方括号前面是否有特定的标点符号
        # 只检查：全角双引号"、全角单引号'、全角左括号（
        bracket_before_pattern = r'([\u201c\u300c\u2018\uff08])\[\{'
        bracket_before_matches = re.findall(bracket_before_pattern, text)
        if bracket_before_matches:
            for match in bracket_before_matches:
                issues.append(f"方括号前出现不允许的符号：{match}[，建议去掉符号")
        
        return len(issues) > 0, issues


    def check_condition_6(self, text):
        """条件6：检查连续标点符号和三个单引号（全角、半角或全半角混合）"""
        issues = []
        
        # 检查连续的标点符号（全角、半角或全半角混合）
        patterns = [
            (r'[！!]{2,}', '连续两个及以上感叹号'),
            (r'[？\?]{2,}', '连续两个及以上问号'),
            (r'[。\.]{2,}', '连续两个及以上点号句号'),
            (r"['']{2,}", '连续两个及以上单引号'),
            (r"[~～]{2,}", '连续两个及以上波浪号')
        ]
        
        for pattern, description in patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                issues.append(f"发现{description}：{match.group()}")
        
        return len(issues) > 0, issues
    
    def check_condition_7(self, text):
        """条件7：非知识点的{word_en}或{phrase_en}必须在纯英文语句中，{word_cn}或{phrase_cn}必须不在纯英文语句中"""
        issues = []
        
        # 查找所有的{word_en}和{phrase_en}
        en_patterns = [
            r'\{word_en\}',
            r'\{phrase_en\}'
        ]
        
        for pattern in en_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                # 检查是否被方括号包围，如果是则跳过
                start_pos = match.start()
                end_pos = match.end()
                
                # 检查前面是否有左方括号
                if start_pos > 0 and text[start_pos-1] == '[':
                    # 检查后面是否有右方括号
                    if end_pos < len(text) and text[end_pos] == ']':
                        continue  # 被方括号包围，跳过检查
                
                # 向前找第一个标点符号（增加]）
                sentence_start = 0
                for i in range(start_pos-1, -1, -1):
                    if i < 0 or text[i] in '！!？?。.""，,]':
                        sentence_start = i + 1
                        break
                
                # 向后找第一个标点符号（增加[）
                sentence_end = len(text)
                for i in range(end_pos, len(text)):
                    if i >= len(text) or text[i] in '！!？?。.""，,[':
                        sentence_end = i
                        break
                
                # 提取句子并去掉知识点
                sentence = text[sentence_start:sentence_end]
                sentence_without_var = sentence[:start_pos-sentence_start] + sentence[end_pos-sentence_start:]
                
                # 检查剩余部分是否只有英文、空格和数字
                # 空字符串被认为是中性的，既可以是英文也可以是中文，不报错
                if sentence_without_var.strip() and not self._is_english_or_number_sentence(sentence_without_var):
                    var_text = match.group()
                    issues.append(f"非知识点的{var_text}不在纯英文语句中，建议加上中括号：[{var_text}]")
        
        # 查找所有的{word_cn}和{phrase_cn}
        cn_patterns = [
            r'\{word_cn\}',
            r'\{phrase_cn\}'
        ]
        
        for pattern in cn_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                # 检查是否被方括号包围，如果是则跳过
                start_pos = match.start()
                end_pos = match.end()
                
                # 检查前面是否有左方括号
                if start_pos > 0 and text[start_pos-1] == '[':
                    # 检查后面是否有右方括号
                    if end_pos < len(text) and text[end_pos] == ']':
                        continue  # 被方括号包围，跳过检查
                
                # 向前找第一个标点符号（增加]）
                sentence_start = 0
                for i in range(start_pos-1, -1, -1):
                    if i < 0 or text[i] in '！!？?。.""，,]':
                        sentence_start = i + 1
                        break
                
                # 向后找第一个标点符号（增加[）
                sentence_end = len(text)
                for i in range(end_pos, len(text)):
                    if i >= len(text) or text[i] in '！!？?。.""，,[':
                        sentence_end = i
                        break
                
                # 提取句子并去掉知识点
                sentence = text[sentence_start:sentence_end]
                sentence_without_var = sentence[:start_pos-sentence_start] + sentence[end_pos-sentence_start:]
                
                # 检查剩余部分是否只有英文、空格和数字
                if sentence_without_var.strip() and self._is_english_or_number_sentence(sentence_without_var):
                    var_text = match.group()
                    # 建议将中文变量改为英文变量
                    if 'word_cn' in var_text:
                        suggestion = var_text.replace('word_cn', 'word_en')
                    else:  # phrase_cn
                        suggestion = var_text.replace('phrase_cn', 'phrase_en')
                    issues.append(f"非知识点的{var_text}在纯英文语句中，建议改为：{suggestion}")
        
        return len(issues) > 0, issues

    def check_condition_8(self, text):
        """条件8：检查半角句号前后两句是否都为英文或数字，因为半角句号可以是小数点"""
        issues = []
        
        # 查找所有半角句号的位置
        dot_positions = [m.start() for m in re.finditer(r'\.', text)]
        
        for dot_pos in dot_positions:
            # 向前找第一个标点符号
            before_start = 0
            for i in range(dot_pos-1, -1, -1):
                if i < 0 or text[i] in '！!？?。[]`，,':
                    before_start = i + 1
                    break
            
            # 向后找第一个标点符号
            after_end = len(text)
            for i in range(dot_pos+1, len(text)):
                if i >= len(text) or text[i] in '！!？?。[]`，,':
                    after_end = i
                    break
            
            # 获取句号前后的句子
            before_sentence = text[before_start:dot_pos]
            after_sentence = text[dot_pos+1:after_end]
            
            # 移除知识点后再检查前后句子是否都为英文或数字
            before_sentence_clean = self._remove_en_brackets(before_sentence)
            after_sentence_clean = self._remove_en_brackets(after_sentence)
            
            before_is_english_or_number = self._is_english_or_number_sentence(before_sentence_clean)
            after_is_english_or_number = self._is_english_or_number_sentence(after_sentence_clean)
            
            # 如果前后句子不都是英文或数字，则报告问题
            if not (before_is_english_or_number and after_is_english_or_number):
                context = text[max(0, dot_pos-10):min(len(text), dot_pos+11)]
                issues.append(f"半角句号前后句子语言不一致，可能导致外国口音：{context}")
        
        return len(issues) > 0, issues

    def check_condition_9(self, text):
        """条件9：{words}两侧不能有中括号"""
        issues = []
        
        # 检查{words}两侧是否有中括号
        pattern = r'\[\{words\}\]'
        matches = re.findall(pattern, text)
        if matches:
            issues.append(f"发现{len(matches)}处[{{words}}]，应该去掉中括号改为{{words}}")
        
        return len(issues) > 0, issues

    def check_condition_10(self, text):
        """条件10：{words}两侧不能有单引号或反引号"""
        issues = []
        pattern = r'\[\{[^}]+\}\]|\{[^}]+\}'
        for match in re.finditer(pattern, text):
            s = match.start()
            e = match.end()
            left = text[s-1] if s > 0 else ''
            right = text[e] if e < len(text) else ''
            left_q = left in ('\'', '"')
            right_q = right in ('\'', '"')
            if left_q ^ right_q:
                t = match.group()
                side = '左侧' if left_q else '右侧'
                issues.append(f"{t} {side}存在单独引号，建议去掉")
        if '`' in text:
            issues.append("发现反引号`，建议去掉")
        return len(issues) > 0, issues

    def auto_fix_condition_2(self, text):
        """条件2修复：去掉形如[80]这种中括号加数字的文本部分"""
        # 查找[数字]模式并删除
        pattern = r'\[\d+\]'
        fixed_text = re.sub(pattern, '', text)
        return fixed_text

    def auto_fix_condition_3(self, text):
        """条件3修复：将全角或半角冒号、波浪线替换为全角逗号，省略号替换为句号，顿号替换为逗号"""
        result = text
        
        # 将全角或半角冒号及波浪线替换为全角逗号
        result = re.sub(r'[:：~]', '，', result)
        
        # 将省略号（……或…）替换为句号
        result = re.sub(r'[…]{1,2}', '。', result)
        
        # 将顿号替换为逗号
        result = re.sub(r'、', '，', result)
        
        return result

    def auto_fix_condition_4(self, text):
        """条件4修复：去掉首尾配对的单引号或反引号"""
        # 处理单引号包围的情况
        if text.startswith("'") and text.endswith("'"):
            return text[1:-1]
        
        # 处理双单引号包围的情况
        if text.startswith("''") and text.endswith("''"):
            return text[2:-2]
            
        # 处理变量被单引号包围的情况
        pattern = r"^'(\{[^}]+\})'$"
        if re.match(pattern, text):
            return re.sub(pattern, r"\1", text)
        
        # 处理反引号包围的情况
        if text.startswith("`") and text.endswith("`"):
            return text[1:-1]
            
        # 处理变量被反引号包围的情况
        backtick_pattern = r"^`(\{[^}]+\})`$"
        if re.match(backtick_pattern, text):
            return re.sub(backtick_pattern, r"\1", text)
            
        return text

    def auto_fix_condition_5(self, text):
        """条件5修复：去掉知识点前后的双引号，去掉方括号后的所有标点符号"""
        patterns = [
            (r'"(\{word_en\})"', r'\1'),
            (r'"(\[{word_en}\])"', r'\1'),
            (r'"(\{phrase_en\})"', r'\1'),
            (r'"(\[{phrase_en}\])"', r'\1')
        ]
        
        result = text
        for pattern, replacement in patterns:
            result = re.sub(pattern, replacement, result)
        
        # 去掉方括号后的所有标点符号（除字母、数字、汉字、[、{和空格外）
        result = re.sub(r'\]([^\w\u4e00-\u9fff\[\{\s])', ']', result)
        
        # 去掉方括号前的特定标点符号（全角双引号”、半角双引号"、全角单引号'、半角单引号'、反引号`、全角右括号）、半角右括号)）
        result = re.sub(r'([\u201c\u300c\u2018\uff08])\[', '[', result)
            
        return result

    def auto_fix_condition_6(self, text):
        """条件6修复：将连续的标点符号去重，优先保留全角符号"""
        # 处理连续的感叹号（优先保留全角）
        text = re.sub(r'[！!]{2,}', lambda m: '！' if '！' in m.group(0) else '!', text)
        # 处理连续的问号（优先保留全角）
        text = re.sub(r'[？\?]{2,}', lambda m: '？' if '？' in m.group(0) else '?', text)
        # 处理连续的句号（优先保留全角）
        text = re.sub(r'[。\.]{2,}', lambda m: '。' if '。' in m.group(0) else '.', text)
        # 处理连续的单引号（优先保留全角）
        text = re.sub(r"['']{2,}", lambda m: "'" if "'" in m.group(0) else "'", text)
        # 处理连续的波浪号（优先保留全角）
        text = re.sub(r'[~～]{2,}', lambda m: '～' if '～' in m.group(0) else '~', text)
        
        return text

    def auto_fix_condition_7(self, text):
        """条件7修复：将不在纯英文语句中的{word_en}或{phrase_en}加上中括号，将在纯英文语句中的{word_cn}或{phrase_cn}改为对应的英文变量"""
        result = text
        
        # 第一步：处理英文变量 - 为不在纯英文语句中的{word_en}或{phrase_en}加上中括号
        en_patterns = [
            r'\{word_en\}',
            r'\{phrase_en\}'
        ]
        
        for pattern in en_patterns:
            matches = list(re.finditer(pattern, result))
            # 从后向前处理，避免位置变化
            for match in reversed(matches):
                start_pos = match.start()
                end_pos = match.end()
                var_text = match.group()
                
                # 检查是否已被方括号包围
                if (start_pos > 0 and result[start_pos-1] == '[' and 
                    end_pos < len(result) and result[end_pos] == ']'):
                    continue  # 已被方括号包围，跳过
                
                # 向前找第一个标点符号
                sentence_start = 0
                for i in range(start_pos-1, -1, -1):
                    if i < 0 or result[i] in '！!？?。.""，,':
                        sentence_start = i + 1
                        break
                
                # 向后找第一个标点符号
                sentence_end = len(result)
                for i in range(end_pos, len(result)):
                    if i >= len(result) or result[i] in '！!？?。.""，,':
                        sentence_end = i
                        break
                
                # 提取句子并去掉知识点
                sentence = result[sentence_start:sentence_end]
                sentence_without_var = sentence[:start_pos-sentence_start] + sentence[end_pos-sentence_start:]

                # 检查剩余部分是否只有英文、空格和数字
                if not sentence_without_var.strip() or not self._is_english_or_number_sentence(sentence_without_var):
                    # 在变量周围添加中括号
                    result = result[:start_pos] + '[' + var_text + ']' + result[end_pos:]
        
        # 第二步：处理中文变量 - 将在纯英文语句中的{word_cn}或{phrase_cn}改为对应的英文变量
        cn_patterns = [
            (r'\{word_cn\}', '{word_en}'),
            (r'\{phrase_cn\}', '{phrase_en}')
        ]
        
        for pattern, replacement in cn_patterns:
            matches = list(re.finditer(pattern, result))
            # 从后向前处理，避免位置变化
            for match in reversed(matches):
                start_pos = match.start()
                end_pos = match.end()
                var_text = match.group()
                
                # 检查是否被方括号包围，如果是则跳过
                if (start_pos > 0 and result[start_pos-1] == '[' and 
                    end_pos < len(result) and result[end_pos] == ']'):
                    continue  # 被方括号包围，跳过
                
                # 向前找第一个标点符号
                sentence_start = 0
                for i in range(start_pos-1, -1, -1):
                    if i < 0 or result[i] in '！!？?。.""，,':
                        sentence_start = i + 1
                        break
                
                # 向后找第一个标点符号
                sentence_end = len(result)
                for i in range(end_pos, len(result)):
                    if i >= len(result) or result[i] in '！!？?。.""，,':
                        sentence_end = i
                        break
                
                # 提取句子并去掉知识点
                sentence = result[sentence_start:sentence_end]
                sentence_without_var = sentence[:start_pos-sentence_start] + sentence[end_pos-sentence_start:]

                # 检查剩余部分是否只有英文、空格和数字
                if sentence_without_var.strip() and self._is_english_or_number_sentence(sentence_without_var):
                    # 将中文变量替换为英文变量
                    result = result[:start_pos] + replacement + result[end_pos:]
        
        return result
    
    def auto_fix_condition_8(self, text):
        """条件8修复：将半角句号直接变成全角句号"""
        # 将半角句号替换为全角句号
        fixed_text = text.replace('.', '。')
        return fixed_text

    def auto_fix_condition_9(self, text):
        """条件9修复：去掉{words}两侧的中括号"""
        # 将[{words}]替换为{words}
        fixed_text = re.sub(r'\[\{words\}\]', '{words}', text)
        return fixed_text

    def auto_fix_condition_10(self, text):
        """条件10修复：去掉不合规的引号和反引号"""
        pattern = r'\[\{[^}]+\}\]|\{[^}]+\}'
        result = text
        matches = list(re.finditer(pattern, result))
        for m in reversed(matches):
            s = m.start()
            e = m.end()
            left = result[s-1] if s > 0 else ''
            right = result[e] if e < len(result) else ''
            left_q = left in ('\'', '"')
            right_q = right in ('\'', '"')
            if left_q ^ right_q:
                if left_q:
                    j = s - 1
                    while j >= 0 and result[j] in ('\'', '"'):
                        j -= 1
                    result = result[:j+1] + result[s:]
                else:
                    j = e
                    while j < len(result) and result[j] in ('\'', '"'):
                        j += 1
                    result = result[:e] + result[j:]
        if '`' in result:
            result = result.replace('`', '')
        return result

    def _remove_en_brackets(self, text):
        """移除所有形如{xxx}或[{xxx}]且大括号中包含'en'的内容"""
        if not text:
            return text
        
        # 移除形如[{xxx_en}]的内容（中括号包围的大括号，且大括号中包含en）
        text = re.sub(r'\[\{[^}]*en[^}]*\}\]', '', text)
        
        # 移除形如{xxx_en}的内容（单独的大括号，且大括号中包含en）
        text = re.sub(r'\{[^}]*en[^}]*\}', '', text)
        
        return text
    
    def _get_sentence_before_position(self, text, position):
        """获取指定位置前的句子（被标点符号分割的文本段）"""
        # 定义句子分隔符（排除空格）
        separators = r'[！!？?。；;：:，,、<>#]'
        
        # 从位置向前查找最近的分隔符
        before_text = text[:position]
        matches = list(re.finditer(separators, before_text))
        
        if matches:
            # 找到最后一个分隔符，取其后的文本
            last_separator_pos = matches[-1].end()
            sentence = before_text[last_separator_pos:].strip()
        else:
            # 没有找到分隔符，取从开头到当前位置的文本
            sentence = before_text.strip()
        
        # 移除包含en的大括号内容
        sentence = self._remove_en_brackets(sentence)
        
        return sentence
    
    def _get_sentence_after_position(self, text, position):
        """获取指定位置后的句子（被标点符号分割的文本段）"""
        # 定义句子分隔符（排除空格）
        separators = r'[！!？?。；;：:，,、<>#]'
        
        # 从位置向后查找最近的分隔符
        after_text = text[position + 1:]  # 跳过当前的句号
        match = re.search(separators, after_text)
        
        if match:
            # 找到分隔符，取其前的文本
            sentence = after_text[:match.start()].strip()
        else:
            # 没有找到分隔符，取到文本末尾
            sentence = after_text.strip()
        
        # 移除包含en的大括号内容
        sentence = self._remove_en_brackets(sentence)
        
        return sentence
    
    def _is_english_or_number_sentence(self, text):
        """判断文本是否只包含英文、数字、空格和特定标点符号
        
        允许的字符：英文字母、数字、空格、"|"、"`"、"["、"]"和常见英文标点符号（包含排版引号 ’）
        """
        if not text:
            return True  # 空文本认为符合条件
        
        # 移除包含en的大括号内容
        text_clean = self._remove_en_brackets(text)
        
        # 严格检查是否只包含英文字母、数字、空格和允许的标点符号
        pattern = r"^[a-zA-Z0-9\s'.,!?;:()\-\"|\[\]’\"]*$"
        return re.match(pattern, text_clean.strip()) is not None
        
    def validate_cell_content(self, content, param1_content, column_name, row_index):
        """校验单元格内容"""
        if pd.isna(content) or str(content).strip() == '':
            return []
        
        content_str = str(content)
        param1_str = str(param1_content) if not pd.isna(param1_content) else ''
        
        validation_results = []
        hit_conditions = []
        all_issues = []
        
        # 检查条件1（不变）
        has_issue_1, issues_1 = self.check_condition_1(content_str, param1_str)
        if has_issue_1:
            hit_conditions.append('1')
            all_issues.extend(issues_1)
        
        # 检查条件2
        has_issue_2, issues_2 = self.check_condition_2(content_str)
        if has_issue_2:
            hit_conditions.append('2')
            all_issues.extend(issues_2)
        
        # 检查条件3
        has_issue_3, issues_3 = self.check_condition_3(content_str)
        if has_issue_3:
            hit_conditions.append('3')
            all_issues.extend(issues_3)
        
        # 检查条件4
        has_issue_4, issues_4 = self.check_condition_4(content_str)
        if has_issue_4:
            hit_conditions.append('4')
            all_issues.extend(issues_4)
        
        # 检查条件5
        has_issue_5, issues_5 = self.check_condition_5(content_str)
        if has_issue_5:
            hit_conditions.append('5')
            all_issues.extend(issues_5)
        
        # 检查条件6
        has_issue_6, issues_6 = self.check_condition_6(content_str)
        if has_issue_6:
            hit_conditions.append('6')
            all_issues.extend(issues_6)
            
        # 检查条件7
        has_issue_7, issues_7 = self.check_condition_7(content_str)
        if has_issue_7:
            hit_conditions.append('7')
            all_issues.extend(issues_7)
            
        # 检查条件8
        has_issue_8, issues_8 = self.check_condition_8(content_str)
        if has_issue_8:
            hit_conditions.append('8')
            all_issues.extend(issues_8)
            
        # 检查条件9
        has_issue_9, issues_9 = self.check_condition_9(content_str)
        if has_issue_9:
            hit_conditions.append('9')
            all_issues.extend(issues_9)
        # 检查条件10
        has_issue_10, issues_10 = self.check_condition_10(content_str)
        if has_issue_10:
            hit_conditions.append('10')
            all_issues.extend(issues_10)
        
        # 如果有任何问题，创建验证结果
        if hit_conditions:
            # 生成修复建议
            suggestion = ''
            # 如果没有命中条件1，尝试自动修复
            if '1' not in hit_conditions:
                suggestion = content_str
                # 按照新的条件顺序应用修复
                if '2' in hit_conditions:
                    suggestion = self.auto_fix_condition_2(suggestion)  # 原条件5
                if '3' in hit_conditions:
                    suggestion = self.auto_fix_condition_3(suggestion)  # 新条件3
                if '4' in hit_conditions:
                    suggestion = self.auto_fix_condition_4(suggestion)  # 新条件4
                if '5' in hit_conditions:
                    suggestion = self.auto_fix_condition_5(suggestion)  # 原条件7
                if '6' in hit_conditions:
                    suggestion = self.auto_fix_condition_6(suggestion)  # 原条件4
                if '7' in hit_conditions:
                    suggestion = self.auto_fix_condition_7(suggestion)  # 新条件7
                if '8' in hit_conditions:
                    suggestion = self.auto_fix_condition_8(suggestion)  # 原条件6
                if '9' in hit_conditions:
                    suggestion = self.auto_fix_condition_9(suggestion)  # 条件9
                if '10' in hit_conditions:
                    suggestion = self.auto_fix_condition_10(suggestion)  # 条件10
            
            validation_results.append({
                'selected': True,  # 默认选中
                'title': column_name,
                'row': row_index + 2,  # 表头算第1行，数据从第2行开始
                'type': ' '.join(hit_conditions),
                'original_text': content_str,
                'suggestion': suggestion,  # 使用生成的修复建议
                'param1': param1_str  # 添加param1字段
            })
        
        return validation_results
    
    def init_ui(self):
        """初始化用户界面"""
        main_layout = QVBoxLayout()  # 改为垂直布局，使页签容器占据100%宽度
        self.setLayout(main_layout)
        
        # 创建页签容器
        self.tab_widget = QTabWidget()
        
        # 第一个页签：内容质检
        quality_tab = QWidget()
        quality_layout = QVBoxLayout()
        quality_tab.setLayout(quality_layout)
        
        # 内容质检页签内部使用水平分割器
        quality_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # 左侧面板（移入内容质检页签内）
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        left_widget.setLayout(left_layout)
        
        # 打开文件按钮和一键检验按钮的布局
        file_btn_layout = QHBoxLayout()
        self.open_files_btn = QPushButton("打开文件")
        self.open_files_btn.setFixedWidth(100) 
        self.open_files_btn.clicked.connect(self.open_files)
        file_btn_layout.addWidget(self.open_files_btn)
        
        # 添加"添加文件"按钮
        self.add_files_btn = QPushButton("添加文件")
        self.add_files_btn.setFixedWidth(100)
        self.add_files_btn.clicked.connect(self.add_files)
        file_btn_layout.addWidget(self.add_files_btn)
        
        file_btn_layout.addStretch()  # 添加弹性空间，使一键检验按钮右对齐
        
        self.batch_validate_btn = QPushButton("一键检验")
        self.batch_validate_btn.setFixedWidth(100)
        self.batch_validate_btn.clicked.connect(self.batch_validate_files)
        self.batch_validate_btn.setEnabled(False)  # 初始状态禁用，有文件时启用
        file_btn_layout.addWidget(self.batch_validate_btn)
        self.batch_fix_btn = QPushButton("☢️批量修复")
        self.batch_fix_btn.setFixedWidth(120)
        self.batch_fix_btn.clicked.connect(self.batch_auto_fix_files)
        self.batch_fix_btn.setEnabled(False)
        file_btn_layout.addWidget(self.batch_fix_btn)
        
        left_layout.addLayout(file_btn_layout)
        
        # 文件列表
        self.file_list = QListWidget()
        self.file_list.itemClicked.connect(self.on_file_item_clicked)
        left_layout.addWidget(self.file_list)
        
        # 配置面板
        config_panel = QGroupBox()
        config_panel.setFixedHeight(80)  # 增加高度以容纳边框和标题
        config_layout = QGridLayout()
        config_layout.setContentsMargins(5, 5, 5, 5)  # 设置内边距
        config_panel.setLayout(config_layout)
        
        # 第一行：配置目录按钮（权重3）、显示目录文本框（权重7）
        self.config_dir_btn = QPushButton("配置目录")
        self.config_dir_btn.clicked.connect(self.select_config_directory)
        config_layout.addWidget(self.config_dir_btn, 0, 0, 1, 3)  # 行0，列0，跨3列
        
        self.config_dir_display = QLineEdit()
        self.config_dir_display.setReadOnly(True)
        self.config_dir_display.setPlaceholderText("请选择配置输出目录...")
        # 从配置文件读取初始值并转换为绝对路径
        speech_config_output = self.config_manager.get('PATH', 'data_output_directory', './output/data')
        absolute_path = os.path.abspath(speech_config_output)
        self.config_dir_display.setText(absolute_path)
        config_layout.addWidget(self.config_dir_display, 0, 3, 1, 6)  # 行0，列3，跨6列
        
        # 第二行：打开目录按钮（权重4）、生成配置文件按钮（权重6）
        self.open_dir_btn = QPushButton("打开目录")
        self.open_dir_btn.clicked.connect(self.open_config_directory)
        config_layout.addWidget(self.open_dir_btn, 0, 9, 1, 1)  # 行0，列9，跨1列
        
        self.speech_checkbox = QCheckBox("Speech")
        self.speech_checkbox.setChecked(True)
        config_layout.addWidget(self.speech_checkbox, 1, 0, 1, 2)

        self.template_checkbox = QCheckBox("TemplateConfig")
        self.template_checkbox.setChecked(True)
        config_layout.addWidget(self.template_checkbox, 1, 2, 1, 2)

        self.generate_config_btn = QPushButton("生成配置文件")
        self.generate_config_btn.clicked.connect(self.generate_config_file)
        config_layout.addWidget(self.generate_config_btn, 1, 4, 1, 6)  # 行1，列4，跨6列
        
        left_layout.addWidget(config_panel)
        
        # 将左侧面板添加到分割器
        quality_splitter.addWidget(left_widget)
        
        # 右侧面板（内容质检表格区域）
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_widget.setLayout(right_layout)
        
        # 添加按钮区域
        buttons_layout = QHBoxLayout()
        
        # 左侧添加"保存"按钮
        self.save_btn = QPushButton("保存")
        self.save_btn.setFixedWidth(100)
        self.save_btn.clicked.connect(self.save_to_xlsx)
        self.save_btn.setEnabled(False)  # 初始状态禁用
        buttons_layout.addWidget(self.save_btn)
        
        # 添加问题数据行数标签
        self.content_issue_count_label = QLabel("问题数据：0行")
        buttons_layout.addWidget(self.content_issue_count_label)
        
        buttons_layout.addStretch()  # 添加弹性空间，使"一键修复"按钮右对齐
        
        # 右侧添加"一键修复"按钮
        self.quick_fix_btn = QPushButton("一键修复")
        self.quick_fix_btn.setFixedWidth(100)
        self.quick_fix_btn.clicked.connect(self.quick_fix_selected)
        self.quick_fix_btn.setEnabled(False)  # 初始状态禁用
        buttons_layout.addWidget(self.quick_fix_btn)
        
        right_layout.addLayout(buttons_layout)
        
        # 校验结果表格
        self.validation_table = QTableView()
        self.validation_model = ValidationTableModel()
        self.validation_table.setModel(self.validation_model)
        
        # 设置表格样式
        header = self.validation_table.horizontalHeader()
        
        # 设置各列宽度
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # 选择列 - 窄
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)  # title列 - 窄
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)  # row列 - 窄
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # type列 - 稍微增加宽度
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # 原文列 - 宽
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # 修复建议列 - 宽
        
        # 设置固定列的具体宽度
        self.validation_table.setColumnWidth(0, 30)   # 选择列
        self.validation_table.setColumnWidth(1, 150)  # title列 - 增加宽度
        self.validation_table.setColumnWidth(2, 60)   # row列
        self.validation_table.setColumnWidth(3, 80)   # type列 - 从原来的自适应改为固定80px
        
        self.validation_table.setAlternatingRowColors(True)
        self.validation_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        
        # 设置右键菜单策略
        self.validation_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.validation_table.customContextMenuRequested.connect(self.show_context_menu)
        
        # 添加点击事件处理
        self.validation_table.clicked.connect(self.on_table_clicked)
        
        # 设置tooltip样式，控制tooltip控件的最大宽度
        self.validation_table.setStyleSheet("""
            QToolTip {
                color: black;
                white-space: normal;
                background-color: orange;
            }
        """)
        
        right_layout.addWidget(self.validation_table)
        
        # 将右侧面板添加到分割器
        quality_splitter.addWidget(right_widget)
        
        # 设置分割器比例 (1:4)
        quality_splitter.setStretchFactor(0, 1)
        quality_splitter.setStretchFactor(1, 4)
        
        # 将分割器添加到内容质检页签布局
        quality_layout.addWidget(quality_splitter)
        
        # 第二个页签：格式校验
        format_tab = QWidget()
        format_layout = QVBoxLayout()
        format_tab.setLayout(format_layout)
        
        # 格式校验页签内部使用水平分割器
        format_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # 左侧面板（格式校验）
        format_left_widget = QWidget()
        format_left_layout = QVBoxLayout()
        format_left_widget.setLayout(format_left_layout)
        
        # 打开文件按钮
        self.format_open_file_btn = QPushButton("打开文件")
        self.format_open_file_btn.setFixedWidth(100)
        self.format_open_file_btn.clicked.connect(self.open_format_file)
        format_left_layout.addWidget(self.format_open_file_btn)
        
        # sheets列表
        self.sheets_list = QListWidget()
        self.sheets_list.itemClicked.connect(self.on_sheet_item_clicked)
        format_left_layout.addWidget(self.sheets_list)
        
        # 将左侧面板添加到分割器
        format_splitter.addWidget(format_left_widget)
        
        # 右侧面板（格式校验）
        format_right_widget = QWidget()
        format_right_layout = QVBoxLayout()
        format_right_widget.setLayout(format_right_layout)
        
        # 按钮区域
        format_buttons_layout = QHBoxLayout()
        
        # 左侧添加"保存"按钮
        self.format_save_btn = QPushButton("保存")
        self.format_save_btn.setFixedWidth(100)
        self.format_save_btn.clicked.connect(self.save_format_file)
        self.format_save_btn.setEnabled(False)  # 初始状态禁用
        format_buttons_layout.addWidget(self.format_save_btn)
        
        # 添加问题数据行数标签
        self.format_issue_count_label = QLabel("问题数据：0行")
        format_buttons_layout.addWidget(self.format_issue_count_label)
        
        format_buttons_layout.addStretch()  # 添加弹性空间，使右侧按钮右对齐
        
        # 右侧添加"一键修复"按钮
        self.format_fix_btn = QPushButton("一键修复")
        self.format_fix_btn.setFixedWidth(100)
        self.format_fix_btn.clicked.connect(self.format_quick_fix)
        self.format_fix_btn.setEnabled(False)  # 初始状态禁用
        format_buttons_layout.addWidget(self.format_fix_btn)
        
        format_right_layout.addLayout(format_buttons_layout)
        
        # 格式校验表格
        self.format_table = QTableView()
        self.format_model = ValidationTableModel()  # 暂时使用相同的模型
        self.format_table.setModel(self.format_model)
        
        # 设置表格样式
        format_header = self.format_table.horizontalHeader()
        
        # 设置各列宽度
        format_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # 选择列 - 窄
        format_header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)  # id列 - 窄
        format_header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)  # stage_id列 - 窄
        format_header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # type列 - 窄
        format_header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # 原文列 - 宽
        format_header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # 修复建议列 - 宽
        
        # 设置固定列的具体宽度
        self.format_table.setColumnWidth(0, 30)   # 选择列
        self.format_table.setColumnWidth(1, 60)  # id列
        self.format_table.setColumnWidth(2, 200)   # stage_id列
        self.format_table.setColumnWidth(3, 60)   # type列
        
        self.format_table.setAlternatingRowColors(True)
        self.format_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        
        # 启用右键菜单
        self.format_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.format_table.customContextMenuRequested.connect(self.show_format_context_menu)

        self.format_table.setStyleSheet("""
            QToolTip {
                color: black;
                white-space: normal;
                background-color: orange;
            }
        """)
        
        format_right_layout.addWidget(self.format_table)
        
        # 将右侧面板添加到分割器
        format_splitter.addWidget(format_right_widget)
        
        # 设置分割器比例 (1:4)
        format_splitter.setSizes([200, 800])
        
        # 将分割器添加到格式校验页签布局
        format_layout.addWidget(format_splitter)
        
        # 第三个页签：关联检测（原数据校验）
        validation_tab = QWidget()
        validation_layout = QVBoxLayout()
        validation_tab.setLayout(validation_layout)
        
        # 数据校验按钮区域
        validation_button_layout = QHBoxLayout()
        
        # 数据校验按钮
        self.data_validation_button = QPushButton("数据校验")
        self.data_validation_button.clicked.connect(self.perform_data_validation)
        validation_button_layout.addWidget(self.data_validation_button)
        
        # 导出结果按钮
        self.export_results_button = QPushButton("导出结果")
        self.export_results_button.clicked.connect(self.export_validation_results)
        validation_button_layout.addWidget(self.export_results_button)
        
        # 添加弹性空间，使左侧按钮左对齐，右侧按钮右对齐
        validation_button_layout.addStretch()
        
        # 尝试修复按钮（右对齐）
        self.try_fix_button = QPushButton("尝试修复")
        self.try_fix_button.clicked.connect(self.try_fix_validation_errors)
        validation_button_layout.addWidget(self.try_fix_button)
        
        validation_layout.addLayout(validation_button_layout)
        
        # 数据校验文本框（大的多行文本框，背景色C0C0C0，可选中不可编辑）
        self.validation_text = QTextEdit()
        self.validation_text.setReadOnly(True)  # 不可编辑
        self.validation_text.setStyleSheet("background-color: #303030;")  # 背景色C0C0C0
        validation_layout.addWidget(self.validation_text)
        
        # 第四个页签：数据拆解
        data_split_tab = QWidget()
        data_split_layout = QVBoxLayout()
        data_split_tab.setLayout(data_split_layout)
        
        # 数据拆解按钮区域
        data_split_button_layout = QHBoxLayout()
        
        # 打开文件按钮
        self.data_split_open_file_button = QPushButton("打开文件")
        self.data_split_open_file_button.clicked.connect(self.open_data_split_file)
        data_split_button_layout.addWidget(self.data_split_open_file_button)
        
        # 文件路径显示标签
        self.data_split_file_path_label = QLabel("未选择文件")
        self.data_split_file_path_label.setStyleSheet("color: #888888; margin-left: 10px;")
        data_split_button_layout.addWidget(self.data_split_file_path_label)
        
        # 添加弹性空间，将后面的控件推到右边
        data_split_button_layout.addStretch()
        
        # 处理模式选择区域
        mode_group_box = QGroupBox("")
        mode_group_box.setStyleSheet("QGroupBox { border: none; }")  # 去掉边框
        mode_layout = QHBoxLayout()
        mode_group_box.setLayout(mode_layout)
        
        # 创建单选按钮组
        self.data_split_mode_group = QButtonGroup()
        
        # 全量模式（默认选中）
        self.full_mode_radio = QRadioButton("全量")
        self.full_mode_radio.setChecked(True)
        self.full_mode_radio.setToolTip("所有dialog_name都参与输出，输出路径：\\split\\[sheet]")
        self.data_split_mode_group.addButton(self.full_mode_radio, 0)
        mode_layout.addWidget(self.full_mode_radio)
        
        # 仅中文模式
        self.chinese_mode_radio = QRadioButton("仅中文")
        self.chinese_mode_radio.setToolTip("仅处理stage_id第6个元素为e0的数据，输出路径：\\chinese\\[sheet]")
        self.data_split_mode_group.addButton(self.chinese_mode_radio, 1)
        mode_layout.addWidget(self.chinese_mode_radio)
        
        # 仅结构模式
        self.frame_mode_radio = QRadioButton("仅结构")
        self.frame_mode_radio.setToolTip("每个stage_name仅生成第一组dialog_name，输出路径：\\frame\\[sheet]")
        self.data_split_mode_group.addButton(self.frame_mode_radio, 2)
        mode_layout.addWidget(self.frame_mode_radio)
        
        data_split_button_layout.addWidget(mode_group_box)
        
        # 开始拆解按钮（放在最右边）
        self.data_split_process_button = QPushButton("开始拆解")
        self.data_split_process_button.clicked.connect(self.process_data_split)
        self.data_split_process_button.setEnabled(False)  # 初始状态禁用
        data_split_button_layout.addWidget(self.data_split_process_button)
        
        data_split_layout.addLayout(data_split_button_layout)
        
        # 数据拆解处理文本框（显示处理过程和结果）
        self.data_split_text = QTextEdit()
        self.data_split_text.setReadOnly(True)
        self.data_split_text.setStyleSheet("background-color: #303030;")
        data_split_layout.addWidget(self.data_split_text)
        
        # 添加页签到容器
        self.tab_widget.addTab(quality_tab, "内容质检")
        self.tab_widget.addTab(format_tab, "格式校验")
        self.tab_widget.addTab(validation_tab, "关联检测")
        self.tab_widget.addTab(data_split_tab, "数据拆解")

        image_tab = QWidget()
        image_layout = QVBoxLayout()
        image_tab.setLayout(image_layout)
        image_splitter = QSplitter(Qt.Orientation.Horizontal)

        image_left_widget = QWidget()
        image_left_layout = QVBoxLayout()
        image_left_widget.setLayout(image_left_layout)

        image_left_btns = QHBoxLayout()
        self.image_open_btn = QPushButton("打开文件")
        self.image_open_btn.setFixedWidth(100)
        self.image_open_btn.clicked.connect(self.open_image_config_file)
        image_left_btns.addWidget(self.image_open_btn)
        self.image_left_sheet_combo = QComboBox()
        self.image_left_sheet_combo.setFixedWidth(120)
        self.image_left_sheet_combo.currentTextChanged.connect(self._on_image_left_sheet_changed)
        image_left_btns.addStretch()
        image_left_btns.addWidget(self.image_left_sheet_combo)
        image_left_layout.addLayout(image_left_btns)

        self.image_left_table = QTableWidget()
        self.image_left_table.setColumnCount(4)
        self.image_left_table.setHorizontalHeaderLabels(["id", "npc1_character", "npc2_character", "npcSceneId"])
        self.image_left_table.horizontalHeader().resizeSection(0, 120)
        self.image_left_table.horizontalHeader().resizeSection(1, 180)
        self.image_left_table.horizontalHeader().resizeSection(2, 180)
        self.image_left_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        image_left_layout.addWidget(self.image_left_table)
        self.image_select_dir_btn = QPushButton("选择图片目录")
        self.image_select_dir_btn.setFixedWidth(150)
        self.image_select_dir_btn.setEnabled(False)
        self.image_select_dir_btn.clicked.connect(self._on_select_image_dir)
        self.image_dir_path_edit = QLineEdit("")
        self.image_dir_path_edit.setReadOnly(True)
        self.image_start_btn = QPushButton("开始配置")
        self.image_start_btn.setFixedWidth(150)
        self.image_start_btn.setEnabled(False)
        self.image_start_btn.clicked.connect(self._on_start_image_config)
        image_left_bottom = QHBoxLayout()
        image_left_bottom.addWidget(self.image_select_dir_btn)
        image_left_bottom.addWidget(self.image_dir_path_edit, 1)
        image_left_bottom.addWidget(self.image_start_btn, 0, Qt.AlignmentFlag.AlignRight)
        image_left_layout.addLayout(image_left_bottom)

        image_splitter.addWidget(image_left_widget)

        image_right_widget = QWidget()
        image_right_layout = QVBoxLayout()
        image_right_widget.setLayout(image_right_layout)

        image_right_top = QHBoxLayout()
        self.image_right_open_btn = QPushButton("打开文件")
        self.image_right_open_btn.setFixedWidth(100)
        self.image_right_open_btn.clicked.connect(self._on_open_right_excel_clicked)
        image_right_top.addWidget(self.image_right_open_btn)
        self.image_save_as_btn = QPushButton("另存为")
        self.image_save_as_btn.setFixedWidth(100)
        self.image_save_as_btn.clicked.connect(self._on_save_as_clicked)
        image_right_top.addWidget(self.image_save_as_btn)
        image_right_top.addStretch()
        self.image_sheet_combo = QComboBox()
        self.image_sheet_combo.setFixedWidth(120)
        image_right_top.addWidget(self.image_sheet_combo)
        self.image_path_transform_btn = QPushButton("路径转换")
        self.image_path_transform_btn.setFixedWidth(100)
        self.image_path_transform_btn.clicked.connect(self._on_path_transform_clicked)
        image_right_top.addWidget(self.image_path_transform_btn)
        self.image_upload_btn = QPushButton("上传文件")
        self.image_upload_btn.setFixedWidth(100)
        self.image_upload_btn.clicked.connect(self._on_upload_images_clicked)
        image_right_top.addWidget(self.image_upload_btn)
        image_right_layout.addLayout(image_right_top)

        self.image_right_table = QTableWidget()
        self.image_right_table.setColumnCount(7)
        self.image_right_table.setHorizontalHeaderLabels(["id", "templateType", "npcSceneId", "stage_type", "npc1_base_id", "npc2_base_id", "url_image"])
        self.image_right_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.image_right_table.horizontalHeader().resizeSection(0, 60)
        self.image_right_table.horizontalHeader().resizeSection(1, 120)
        self.image_right_table.horizontalHeader().resizeSection(2, 120)
        self.image_right_table.horizontalHeader().resizeSection(3, 120)
        self.image_right_table.horizontalHeader().resizeSection(4, 120)
        self.image_right_table.horizontalHeader().resizeSection(5, 120)
        self.image_right_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self.image_right_table.cellClicked.connect(self._on_right_table_cell_clicked)
        image_right_layout.addWidget(self.image_right_table)

        image_splitter.addWidget(image_right_widget)
        image_splitter.setSizes([400, 600])
        image_layout.addWidget(image_splitter)

        self.image_config_tab_index = self.tab_widget.addTab(image_tab, "图片配置")
        self.tab_widget.currentChanged.connect(self._on_tab_changed)
        self.image_sheet_combo.currentTextChanged.connect(self._on_image_sheet_changed)

        online_tab = QWidget()
        online_layout = QVBoxLayout()
        online_tab.setLayout(online_layout)
        online_top = QHBoxLayout()
        self.online_check_start_button = QPushButton("开始检测")
        self.online_check_start_button.clicked.connect(self._on_start_online_check)
        online_top.addWidget(self.online_check_start_button)
        online_top.addStretch()
        self.online_check_save_button = QPushButton("保存日志")
        self.online_check_save_button.clicked.connect(self._on_save_online_check_log)
        online_top.addWidget(self.online_check_save_button)
        online_layout.addLayout(online_top)
        self.online_check_text = QTextEdit()
        self.online_check_text.setReadOnly(True)
        self.online_check_text.setStyleSheet("background-color: #303030;")
        online_layout.addWidget(self.online_check_text)
        self.tab_widget.addTab(online_tab, "线上检测")

        # 将页签容器添加到主布局
        main_layout.addWidget(self.tab_widget)

    def open_image_config_file(self):
        cfg = ConfigManager()
        last_path = cfg.get('PATH', 'content_validator_last_open_path', './')
        file_path, _ = QFileDialog.getOpenFileName(self, "打开文件", last_path, "Excel 文件 (*.xlsx)")
        if not file_path:
            return
        try:
            from duoki_editor.utils.excel_handler import ExcelHandler
            excel_handler = ExcelHandler()
            excel_data = excel_handler.load_excel(file_path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开文件失败：{str(e)}")
            return
        try:
            d = os.path.dirname(file_path)
            cfg.config['PATH']['content_validator_last_open_path'] = d
            cfg.save_config()
        except Exception:
            pass
        self._image_left_data = {}
        for sheet, df in excel_data.items():
            cols = [c for c in df.columns]
            col_map = {str(c).lower(): c for c in cols}
            need = ['id', 'npc1_character', 'npc2_character']
            ok = all(n in col_map for n in need)
            npc_col = None
            if 'npcsceneid' in col_map:
                npc_col = col_map['npcsceneid']
            elif 'npcsceneid' in [str(c) for c in cols] or 'npcSceneId' in cols:
                npc_col = 'npcSceneId' if 'npcSceneId' in cols else col_map.get('npcsceneid')
            else:
                npc_col = None
            ok = ok and (npc_col is not None or ('npcSceneId' in cols))
            if ok:
                df2 = df.copy()
                if npc_col and npc_col != 'npcSceneId':
                    df2.rename(columns={npc_col: 'npcSceneId'}, inplace=True)
                sel_cols = ['id', 'npc1_character', 'npc2_character', 'npcSceneId']
                sel = [c for c in sel_cols if c in df2.columns]
                if len(sel) == 4:
                    self._image_left_data[str(sheet)] = df2[sel].iloc[1:].reset_index(drop=True)
        if not self._image_left_data:
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast("格式不符，请选择模板配置表")
            else:
                QMessageBox.warning(self, "提示", "格式不符，请选择模板配置表")
            return
        # 填充左侧sheet下拉并显示默认第一个sheet
        names = list(self._image_left_data.keys())
        self.image_left_sheet_combo.clear()
        self.image_left_sheet_combo.addItems(names)
        if names:
            self._set_left_table_sheet(names[0])
        # 左侧数据加载后允许选择图片目录，重置路径与开始配置状态
        try:
            self.image_select_dir_btn.setEnabled(bool(names))
            self.image_dir_path_edit.setText("")
            self.image_start_btn.setEnabled(False)
        except Exception:
            pass
        try:
            self.image_sheet_combo.clear()
            self._image_right_data = {}
            self._load_show_image_npc_save()
        except Exception:
            pass

    def _on_image_left_sheet_changed(self, name):
        self._set_left_table_sheet(name)

    def _set_left_table_sheet(self, name):
        df = self._image_left_data.get(name)
        if df is None:
            self.image_left_table.setRowCount(0)
            return
        self._populate_left_table(df)

    def _on_select_image_dir(self):
        cfg = ConfigManager()
        base = cfg.get_story_image_last_open_path()
        selected_dir = QFileDialog.getExistingDirectory(self, "选择图片目录", base)
        if not selected_dir:
            return
        names = list(self._image_left_data.keys())
        if not names:
            return
        ok = True
        for name in names:
            sub = os.path.join(selected_dir, str(name))
            if not os.path.isdir(sub):
                ok = False
                break
        if ok:
            abs_path = os.path.abspath(selected_dir)
            norm_path = abs_path.replace('\\', '/')
            self.image_dir_path_edit.setText(norm_path)
            self.image_start_btn.setEnabled(True)
            cfg.set_story_image_last_open_path(norm_path)
        else:
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast("所选目录下没有表单对应的子目录")
            else:
                QMessageBox.warning(self, "提示", "所选目录下没有表单对应的子目录")

    def _populate_left_table(self, df):
        self.image_left_table.setRowCount(len(df))
        for r in range(len(df)):
            for c, col in enumerate(['id', 'npc1_character', 'npc2_character', 'npcSceneId']):
                val = ''
                try:
                    raw = df.iloc[r][col]
                    if pd.isna(raw) or str(raw).strip().lower() in ['nan', 'none']:
                        val = ''
                    else:
                        val = str(raw)
                except Exception:
                    val = ''
                item = QTableWidgetItem(val)
                self.image_left_table.setItem(r, c, item)

    def _on_tab_changed(self, idx):
        if idx == self.image_config_tab_index:
            self._load_show_image_npc_save()

    def _load_show_image_npc_save(self):
        base = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'data', 'save', 'ShowImageNpc_save.xlsx')
        self.image_sheet_combo.clear()
        self._image_right_data = {}
        if not os.path.exists(base):
            return
        try:
            xls = pd.ExcelFile(base)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, header=None)
                if len(df) > 1:
                    cols = df.iloc[0].tolist()
                    data = df.iloc[2:].reset_index(drop=True)
                    data.columns = cols
                    self._image_right_data[sheet] = data
            names = list(self._image_right_data.keys())
            self.image_sheet_combo.addItems(names)
            if names:
                self._set_right_table_sheet(names[0])
        except Exception:
            pass

    def _on_image_sheet_changed(self, name):
        self._set_right_table_sheet(name)

    def _set_right_table_sheet(self, name):
        df = self._image_right_data.get(name)
        if df is None:
            self.image_right_table.setRowCount(0)
            return
        cols = ['id', 'templateType', 'npcSceneId', 'stage_type', 'npc1_base_id', 'npc2_base_id', 'url_image']
        for col in cols:
            if col not in df.columns:
                df[col] = ''
        view = df[cols]
        self.image_right_table.setRowCount(len(view))
        for r in range(len(view)):
            for c, col in enumerate(cols):
                val = ''
                try:
                    raw = view.iloc[r][col]
                    if pd.isna(raw) or str(raw).strip().lower() in ['nan', 'none']:
                        val = ''
                    else:
                        val = str(raw)
                except Exception:
                    val = ''
                item = QTableWidgetItem(val)
                raw_url = str(df.iloc[r].get('url_image') or '').strip()
                is_local_abs = os.path.isabs(raw_url) and not raw_url.lower().startswith(('http://', 'https://'))
                status = str(df.iloc[r].get('_upload_status') or '').strip().lower()
                miss = False
                if '_missing_image' in df.columns:
                    miss = bool(df.iloc[r].get('_missing_image')) and is_local_abs
                if status == 'missing':
                    item.setForeground(QBrush(QColor(255, 165, 0)))
                elif status == 'failed':
                    item.setForeground(QBrush(QColor(255, 80, 80)))
                elif status == 'success':
                    item.setForeground(QBrush(QColor(34, 139, 34)))
                elif miss:
                    item.setForeground(QBrush(QColor(255, 80, 80)))
                else:
                    if col == 'url_image' and val:
                        font = QFont()
                        font.setUnderline(True)
                        item.setFont(font)
                        item.setForeground(QBrush(QColor(70, 130, 180)))
                self.image_right_table.setItem(r, c, item)

    def _on_right_table_cell_clicked(self, row, col):
        if col != 6:
            return
        item = self.image_right_table.item(row, col)
        if not item:
            return
        text = str(item.text() or '').strip()
        if not text:
            return
        is_local_abs = os.path.isabs(text) and not text.lower().startswith(('http://', 'https://'))
        if is_local_abs:
            if os.path.exists(text):
                dlg = QDialog(self)
                dlg.setWindowTitle("图片预览")
                v = QVBoxLayout()
                lbl = QLabel()
                pix = QPixmap(text)
                lbl.setPixmap(pix)
                lbl.setScaledContents(True)
                lbl.setMinimumSize(400, 300)
                v.addWidget(lbl)
                btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
                btns.accepted.connect(dlg.accept)
                v.addWidget(btns)
                dlg.setLayout(v)
                dlg.exec()
            else:
                QMessageBox.warning(self, "图片", f"文件不存在: {text}")
            return
        base_url = "https://portal-test.qidianlingzhi.com:10199/client_resources/getFile?path=client/restaurant/image/"
        full_url = base_url + text
        try:
            from duoki_editor.modules.online_data.online_data_viewer import ImagePreviewDialog
            dlg = ImagePreviewDialog(text, auth_manager=self.auth_manager, parent=self)
            dlg.load_image(full_url)
            dlg.exec()
        except Exception:
            QMessageBox.information(self, "图片", full_url)

    def _to_pinyin(self, text):
        from re import sub
        t = str(text or '').strip().replace('_', '')
        from pypinyin import pinyin, Style
        pys = pinyin(t, style=Style.NORMAL, strict=False)
        flat = [s[0] for s in pys if s]
        res = '-'.join(flat)
        res = sub(r"[^a-zA-Z0-9\-]", "", res)
        return res.lower()

    def _on_path_transform_clicked(self):
        try:
            cfg = ConfigManager()
            base_out = cfg.get('PATH', 'scene_output_directory', './output/scene')
            if not base_out:
                base_out = './output/scene'
            print(f"[路径转换] 开始，输出目录: {base_out}")
            for sheet_name, df in list(getattr(self, '_image_right_data', {}).items()):
                if df is None or df.empty:
                    continue
                out_dir = os.path.abspath(os.path.join(base_out, str(sheet_name), 'story'))
                os.makedirs(out_dir, exist_ok=True)
                changed = False
                converted_count = 0
                if '_missing_image' not in df.columns:
                    df['_missing_image'] = False
                for i in range(len(df)):
                    raw_url = str(df.iloc[i].get('url_image') or '').strip()
                    if not raw_url:
                        continue
                    if not (os.path.isabs(raw_url) and not raw_url.lower().startswith(('http://','https://'))):
                        continue
                    if not os.path.exists(raw_url):
                        df.at[i, '_missing_image'] = True
                        continue
                    npc1_base = str(df.iloc[i].get('npc1_base_id') or '').strip()
                    npc2_base = str(df.iloc[i].get('npc2_base_id') or '').strip()
                    base_name = os.path.basename(raw_url)
                    name_no_ext, ext = os.path.splitext(base_name)
                    if not ext:
                        ext = '.jpg'
                    name_no_ext_str = str(name_no_ext)
                    parts = name_no_ext_str.split('_')
                    if len(parts) >= 2:
                        front_part = '_'.join(parts[:-1])
                        dash_parts = front_part.split('-')
                        if dash_parts and len(dash_parts[-1]) == 1 and dash_parts[-1].isalpha():
                            dash_parts = dash_parts[:-1]
                        s1 = str(sheet_name)
                        s2 = ''
                        s3 = ''
                        if len(dash_parts) >= 2:
                            s2 = self._to_pinyin(dash_parts[1])
                        if len(dash_parts) >= 3:
                            s3 = dash_parts[2]
                        front_segments = [s1]
                        if s2:
                            front_segments.append(s2)
                        if s3:
                            front_segments.append(str(s3))
                        new_front = '-'.join(front_segments)
                    else:
                        f_segs = name_no_ext_str.split('-')
                        s1 = str(sheet_name)
                        s2 = self._to_pinyin(f_segs[1]) if len(f_segs) >= 2 else ''
                        new_front = f"{s1}-{s2}".strip('-')
                    new_back = f"{npc1_base}-{npc2_base}"
                    new_name = f"story_{new_front}_{new_back}{ext}"
                    src = raw_url
                    dst = os.path.join(out_dir, new_name)
                    import shutil
                    shutil.copy2(src, dst)
                    rel_path = os.path.join(str(sheet_name), 'story', new_name).replace('\\','/')
                    df.at[i, 'url_image'] = rel_path
                    df.at[i, '_missing_image'] = False
                    changed = True
                    converted_count += 1
                if changed:
                    self._image_right_data[sheet_name] = df
                    print(f"[路径转换] sheet={sheet_name} 转换 {converted_count} 个")
            curr = self.image_sheet_combo.currentText()
            if curr:
                self._set_right_table_sheet(curr)
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast('路径转换成功')
            out_dir2 = cfg.get('PATH', 'data_output_directory', './output/data')
            if not out_dir2:
                out_dir2 = './output/data'
            os.makedirs(out_dir2, exist_ok=True)
            target_xlsx = os.path.join(out_dir2, 'ShowImageNpc.xlsx')
            print(f"[导出] 自动导出到 {target_xlsx}")
            self._export_right_data_to_excel(target_xlsx)
        except Exception as e:
            print(f"[路径转换] 失败: {e}")
            QMessageBox.critical(self, '路径转换', str(e))

    def _on_open_right_excel_clicked(self):
        cfg = ConfigManager()
        last_path = cfg.get('PATH', 'content_validator_last_open_path', './')
        file_path, _ = QFileDialog.getOpenFileName(self, "打开文件", last_path, "Excel 文件 (*.xlsx)")
        if not file_path:
            return
        import pandas as pd
        loaded = {}
        xls = pd.ExcelFile(file_path)
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            if len(df) > 1:
                cols = df.iloc[0].tolist()
                data = df.iloc[2:].reset_index(drop=True)
                data.columns = cols
                loaded[sheet] = data
        if not loaded:
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast('格式不符，请选择剧情图片配置表')
            return
        for _, d in loaded.items():
            cols = [str(c) for c in list(d.columns)]
            if 'npcSceneId' not in cols or 'url_image' not in cols:
                if hasattr(self, 'toast_manager') and self.toast_manager:
                    self.toast_manager.show_toast('格式不符，请选择剧情图片配置表')
                return
        self._image_right_data = loaded
        names = list(loaded.keys())
        self.image_sheet_combo.clear()
        self.image_sheet_combo.addItems(names)
        if names:
            self._set_right_table_sheet(names[0])

    def _export_right_data_to_excel(self, file_path):
        cols = ['id', 'templateType', 'npcSceneId', 'stage_type', 'npc1_base_id', 'npc2_base_id', 'url_image']
        import xlsxwriter
        print(f"[导出] 开始，目标文件: {file_path}")
        wb = xlsxwriter.Workbook(file_path)
        try:
            for sheet_name, df in list(getattr(self, '_image_right_data', {}).items()):
                if df is None:
                    continue
                sheet = str(sheet_name)
                invalid = ['\\', '/', '?', '*', '[', ']']
                for ch in invalid:
                    sheet = sheet.replace(ch, '_')
                ws = wb.add_worksheet(sheet[:31])
                for c, h in enumerate(cols):
                    ws.write(0, c, h)
                fmt_row = ['int','string','string','string','int','int','string']
                for c, t in enumerate(fmt_row):
                    ws.write(1, c, t)
                for r in range(len(df)):
                    rowdata = []
                    for h in cols:
                        v = df.iloc[r].get(h)
                        if pd.isna(v):
                            v = ''
                        rowdata.append(v)
                    try:
                        v_id = rowdata[0]
                        rowdata[0] = int(v_id) if str(v_id).strip().isdigit() else v_id
                    except Exception:
                        pass
                    try:
                        v1 = rowdata[4]
                        rowdata[4] = int(v1) if str(v1).strip().isdigit() else v1
                    except Exception:
                        pass
                    try:
                        v2 = rowdata[5]
                        rowdata[5] = int(v2) if str(v2).strip().isdigit() else v2
                    except Exception:
                        pass
                    for c, v in enumerate(rowdata):
                        ws.write(2 + r, c, v)
                print(f"[导出] sheet={sheet} 行数={len(df)}")
        finally:
            wb.close()
            print(f"[导出] 完成: {file_path}")

    def _on_save_as_clicked(self):
        cfg = ConfigManager()
        base_dir = cfg.get('PATH', 'data_output_directory', './output/data')
        if not base_dir:
            base_dir = './output/data'
        os.makedirs(base_dir, exist_ok=True)
        default_path = os.path.join(base_dir, 'ShowImageNpc.xlsx')
        file_path, _ = QFileDialog.getSaveFileName(self, '另存为', default_path, 'Excel 文件 (*.xlsx)')
        if not file_path:
            return
        print(f"[另存为] 保存到 {file_path}")
        self._export_right_data_to_excel(file_path)

    def _on_upload_images_clicked(self):
        cfg = ConfigManager()
        base_out = cfg.get('PATH', 'scene_output_directory', './output/scene')
        if not base_out:
            base_out = './output/scene'
        print(f"[上传] 开始，基目录: {base_out}")
        cookie = cfg.get('KEY', 'cookie', '')
        headers = {}
        if cookie:
            headers['Cookie'] = cookie
        url_api = 'https://portal-test.qidianlingzhi.com:10199/client_resources/upload'
        tasks = []
        missing_stats = {}
        for sheet_name, df in list(getattr(self, '_image_right_data', {}).items()):
            if df is None or df.empty:
                continue
            missing_stats[sheet_name] = {'missing': 0, 'success': 0, 'fail': 0}
            for i in range(len(df)):
                npc_scene_id_raw = df.iloc[i].get('npcSceneId')
                npc_scene_id = '' if pd.isna(npc_scene_id_raw) else str(npc_scene_id_raw).strip()
                if not npc_scene_id or npc_scene_id.lower() in ['nan', 'none']:
                    continue
                url_rel = str(df.iloc[i].get('url_image') or '').strip()
                if not url_rel:
                    df.at[i, '_upload_status'] = 'missing'
                    missing_stats[sheet_name]['missing'] += 1
                    continue
                local_path = os.path.abspath(os.path.join(base_out, url_rel))
                if not os.path.exists(local_path):
                    df.at[i, '_upload_status'] = 'missing'
                    missing_stats[sheet_name]['missing'] += 1
                    continue
                remote_name = 'client/restaurant/image/' + url_rel.replace('\\', '/')
                tasks.append({
                    'sheet': sheet_name,
                    'row': i,
                    'url_rel': url_rel,
                    'local_path': local_path,
                    'remote_name': remote_name,
                    'headers': headers,
                    'api': url_api
                })
            self._image_right_data[sheet_name] = df
        if not tasks:
            curr = self.image_sheet_combo.currentText()
            if curr:
                self._set_right_table_sheet(curr)
            print("[上传] 无可上传项")
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast('上传图片完成，请查看结果')
            return
        def worker_fn(t):
            with open(t['local_path'], 'rb') as f:
                files = {'files': (t['remote_name'], f, 'image/jpeg')}
                resp = requests.post(t['api'], headers=t['headers'], files=files, verify=False)
            return {'sheet': t['sheet'], 'row': t['row'], 'url_rel': t['url_rel'], 'code': resp.status_code}
        runner = ThreadPoolRunner(tasks, worker_fn, max_workers=10)
        self._set_upload_busy(True)
        self.image_upload_btn.setText(f"上传中 0/{len(tasks)}")
        def on_done(res):
            sh = res.get('sheet')
            r = res.get('row')
            code = int(res.get('code') or 0)
            if 200 <= code < 300:
                self._image_right_data[sh].at[r, '_upload_status'] = 'success'
                missing_stats[sh]['success'] += 1
                print(f"[上传] 成功 {sh}/{res.get('url_rel')} -> {code}")
            else:
                self._image_right_data[sh].at[r, '_upload_status'] = 'failed'
                missing_stats[sh]['fail'] += 1
                print(f"[上传] 失败 {sh}/{res.get('url_rel')} -> {code}")
        def on_failed(err):
            t = err.get('task')
            sh = t.get('sheet')
            r = t.get('row')
            self._image_right_data[sh].at[r, '_upload_status'] = 'failed'
            missing_stats[sh]['fail'] += 1
            print(f"[上传] 异常 {sh}/{t.get('url_rel')}: {err.get('error')}")
        def on_progress(cur, tot, msg):
            self.image_upload_btn.setText(f"上传中 {cur}/{tot}")
        def on_finished():
            curr = self.image_sheet_combo.currentText()
            if curr:
                self._set_right_table_sheet(curr)
            for sh, s in missing_stats.items():
                print(f"[上传] sheet={sh} 缺失={s['missing']} 成功={s['success']} 失败={s['fail']}")
            print("[上传] 完成")
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast('上传图片完成，请查看结果')
            self._set_upload_busy(False)
            self.image_upload_btn.setText("上传文件")
        runner.item_done.connect(on_done)
        runner.item_failed.connect(on_failed)
        runner.progress_updated.connect(on_progress)
        runner.finished.connect(on_finished)
        self._upload_runner = runner
        runner.start()

    def _on_start_online_check(self):
        url = 'https://portal-test.qidianlingzhi.com:10199/server/sendGMCmd'
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        cookie_header = ''
        if hasattr(self, 'auth_manager') and self.auth_manager:
            try:
                cookie_header = self.auth_manager.get_cookie_header()
            except Exception:
                cookie_header = ''
        if not cookie_header:
            try:
                cfg = ConfigManager()
                cookie_header = cfg.get('KEY', 'cookie', '')
            except Exception:
                cookie_header = ''
        if cookie_header:
            headers['Cookie'] = cookie_header
        data = {
            'serviceName': 'restaurant',
            'command': 'level_game.check+'
        }
        print('[线上检测] 开始调用接口')
        try:
            resp = requests.post(url, headers=headers, data=data, timeout=15, verify=False)
            status = resp.status_code
            text = resp.text or ''
            print(f"[线上检测] 状态码: {status}")
            content = str(text).strip()
            if content:
                self.online_check_text.setPlainText(content)
            else:
                self.online_check_text.setPlainText('检验通过，未命中已知问题')
            if hasattr(self, 'toast_manager') and self.toast_manager:
                if 200 <= status < 300:
                    self.toast_manager.show_success('检测成功，请查看检测结果')
                else:
                    msg = f"检测失败，错误原因：HTTP {status}"
                    self.toast_manager.show_error(msg)
            print('[线上检测] 接口调用完成')
        except requests.exceptions.RequestException as e:
            err = str(e)
            self.online_check_text.setPlainText(err)
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_error(f"检测失败，错误原因：{err}")
            print(f"[线上检测] 调用失败: {err}")

    def _on_save_online_check_log(self):
        text = ''
        try:
            text = self.online_check_text.toPlainText()
        except Exception:
            text = ''
        if not text:
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_warning('没有可保存的日志')
            return
        try:
            from datetime import datetime
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_dir = './output/logs'
            os.makedirs(base_dir, exist_ok=True)
            default_path = os.path.join(base_dir, f'online_check_{ts}.txt')
            file_path, _ = QFileDialog.getSaveFileName(self, '保存日志', default_path, '文本文件 (*.txt)')
            if not file_path:
                return
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(text)
            print(f"[线上检测] 日志已保存: {file_path}")
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_success('日志已保存')
        except Exception as e:
            msg = str(e)
            print(f"[线上检测] 日志保存失败: {msg}")
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_error(f"日志保存失败: {msg}")

    def _set_upload_busy(self, busy):
        b = bool(busy)
        self.image_open_btn.setEnabled(not b)
        self.image_save_as_btn.setEnabled(not b)
        self.image_path_transform_btn.setEnabled(not b)
        self.image_upload_btn.setEnabled(not b)
        self.image_start_btn.setEnabled(not b)
        self.image_select_dir_btn.setEnabled(not b)

    def _on_start_image_config(self):
        try:
            if not hasattr(self, '_image_left_data') or not self._image_left_data:
                raise RuntimeError('未加载左侧数据')
            img_dir = str(self.image_dir_path_edit.text() or '').strip()
            if not img_dir:
                raise RuntimeError('未选择图片目录')

            if hasattr(self, '_image_right_data') and isinstance(self._image_right_data, dict):
                for sh, df in list(self._image_right_data.items()):
                    if df is None:
                        continue
                    if '_upload_status' in df.columns:
                        df['_upload_status'] = ''
                    if '_missing_image' in df.columns:
                        for i in range(len(df)):
                            raw_url = str(df.iloc[i].get('url_image') or '').strip()
                            is_local_abs = os.path.isabs(raw_url) and not raw_url.lower().startswith(('http://','https://'))
                            if is_local_abs:
                                df.at[i, '_missing_image'] = (not os.path.exists(raw_url))
                    self._image_right_data[sh] = df
                curr = self.image_sheet_combo.currentText()
                if curr:
                    self._set_right_table_sheet(curr)

            left_names = [self.image_left_sheet_combo.itemText(i) for i in range(self.image_left_sheet_combo.count())]
            if not hasattr(self, '_image_right_data'):
                self._image_right_data = {}
            right_names = set(self._image_right_data.keys())
            map3 = get_npc_id_map(3)
            inv_map3 = {}
            for k, v in (map3 or {}).items():
                try:
                    inv_map3[str(v).strip()] = str(k)
                except Exception:
                    pass
            for sheet_name in left_names:
                left_df = self._image_left_data.get(sheet_name)
                if left_df is None:
                    continue
                if sheet_name not in right_names:
                    self._image_right_data[sheet_name] = pd.DataFrame(columns=['id','templateType','npcSceneId','stage_type','npc1_base_id','npc2_base_id','url_image','_missing_image'])
                right_df = self._image_right_data.get(sheet_name)
                base = left_names.index(sheet_name) * 10000
                existing_ids = pd.to_numeric(right_df['id'], errors='coerce') if 'id' in right_df.columns else pd.Series(dtype='int64')
                max_id = int(existing_ids.max()) if len(existing_ids) > 0 and not pd.isna(existing_ids.max()) else 0
                next_id = max(max_id, base) + 1
                for _, row in left_df.iterrows():
                    npc_scene_id_raw = row.get('npcSceneId')
                    npc_scene_id = '' if pd.isna(npc_scene_id_raw) else str(npc_scene_id_raw).strip()
                    if not npc_scene_id or npc_scene_id.lower() in ['nan','none']:
                        continue
                    id_raw = row.get('id')
                    id_str = '' if pd.isna(id_raw) else str(id_raw).strip().lower()
                    if not id_str.endswith('_a7_e0'):
                        continue
                    parts = id_str.split('_')
                    template_type = parts[1] if len(parts) >= 2 else ''
                    npc1_char = row.get('npc1_character')
                    npc2_char = row.get('npc2_character')
                    npc1_base = self.character_table_manager.get_base_id_by_id(npc1_char)
                    npc2_base = self.character_table_manager.get_base_id_by_id(npc2_char)
                    npc1_name = inv_map3.get(str(npc1_base).strip(), '') if npc1_base is not None else ''
                    npc2_name = inv_map3.get(str(npc2_base).strip(), '') if npc2_base is not None else ''
                    file_name = f"{npc_scene_id}_{npc1_name}-{npc2_name}.jpg"
                    abs_dir = os.path.join(img_dir, str(sheet_name))
                    abs_path = os.path.abspath(os.path.join(abs_dir, file_name))
                    missing = not os.path.exists(abs_path)
                    key_npc1 = str(int(npc1_base)) if isinstance(npc1_base, (int,)) else str(npc1_base or '').strip()
                    key_npc2 = str(int(npc2_base)) if isinstance(npc2_base, (int,)) else str(npc2_base or '').strip()
                    exists_mask = (
                        right_df['npcSceneId'].astype(str).str.strip() == npc_scene_id
                    ) & (
                        right_df['npc1_base_id'].astype(str).str.strip() == key_npc1
                    ) & (
                        right_df['npc2_base_id'].astype(str).str.strip() == key_npc2
                    ) & (
                        right_df['stage_type'].astype(str).str.strip() == 'show_image_npc'
                    ) 
                    if exists_mask.any():
                        continue
                    new_row = {
                        'id': next_id,
                        'templateType': template_type,
                        'npcSceneId': npc_scene_id,
                        'stage_type': 'show_image_npc',
                        'npc1_base_id': int(npc1_base) if isinstance(npc1_base, (int,)) else (npc1_base if npc1_base is not None else ''),
                        'npc2_base_id': int(npc2_base) if isinstance(npc2_base, (int,)) else (npc2_base if npc2_base is not None else ''),
                        'url_image': abs_path,
                        '_missing_image': missing
                    }
                    self._image_right_data[sheet_name] = pd.concat([right_df, pd.DataFrame([new_row])], ignore_index=True)
                    right_df = self._image_right_data.get(sheet_name)
                    next_id += 1
            self.image_sheet_combo.clear()
            self.image_sheet_combo.addItems(list(self._image_right_data.keys()))
            curr = self.image_sheet_combo.currentText()
            if curr:
                self._set_right_table_sheet(curr)
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast('图片文件配置完成')
        except Exception as e:
            QMessageBox.critical(self, '错误', str(e))
    
    def toggle_select_all(self):
        """切换全选状态"""
        self.validation_model.toggle_select_all()
        # 更新按钮文本
        if self.validation_model.select_all_state:
            self.select_all_btn.setText("全选")
        else:
            self.select_all_btn.setText("取消全选")
            
    def open_format_file(self):
        """打开格式校验文件"""
        # 从配置文件获取上次打开路径
        config = ConfigParser()
        config.read(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), encoding='utf-8')
        
        last_path = ""
        if config.has_section('PATH') and config.has_option('PATH', 'format_validator_last_open_path'):
            last_path = config.get('PATH', 'format_validator_last_open_path')
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "打开Excel文件", last_path, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # 保存路径到配置文件
                config = ConfigParser()
                config.read(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), encoding='utf-8')
                
                if not config.has_section('PATH'):
                    config.add_section('PATH')
                
                config.set('PATH', 'format_validator_last_open_path', os.path.dirname(file_path))
                with open(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), 'w', encoding='utf-8') as f:
                    config.write(f)
                
                # 读取Excel文件
                self.format_df = pd.read_excel(file_path, sheet_name=None)
                # 保存当前文件路径，用于保存功能
                self.current_format_file = file_path
                
                # 检查是否存在必要的列
                required_columns = ['stage_type', 'speaker', 'param1']
                missing_columns = []
                
                # 清空sheets列表
                self.sheets_list.clear()
                
                # 检查每个sheet是否包含必要的列
                valid_sheets = []
                for sheet_name, df in self.format_df.items():
                    sheet_columns = df.columns.tolist()
                    current_missing = [col for col in required_columns if col not in sheet_columns]
                    
                    if not current_missing:
                        valid_sheets.append(sheet_name)
                        # 为每个sheet添加5个带后缀的项目
                        param_suffixes = ['[param1]', '[param2]', '[param3]', '[param4]', '[param5]']
                        for suffix in param_suffixes:
                            self.sheets_list.addItem(f"{sheet_name} {suffix}")
                    else:
                        missing_columns.extend(current_missing)
                
                # 如果没有有效的sheet，显示警告
                if not valid_sheets:
                    missing_columns = list(set(missing_columns))  # 去重
                    QMessageBox.warning(
                        self, 
                        "表格数据不符合要求", 
                        f"缺少必要的列({', '.join(required_columns)})"
                    )
                else:
                    # 如果有有效的sheet，选中第一个
                    self.sheets_list.setCurrentRow(0)
                    self.on_sheet_item_clicked(self.sheets_list.item(0))
                    
            except Exception as e:
                QMessageBox.critical(self, "错误", f"打开文件时发生错误: {str(e)}")
    
    def on_sheet_item_clicked(self, item):
        """处理sheet列表项点击事件"""
        if not item:
            return
            
        full_name = item.text()
        # 提取实际sheet名称（去除后缀）
        sheet_name = full_name.split(" [param")[0]
        # 获取参数后缀
        param_suffix = ""
        if "[param" in full_name:
            param_suffix = full_name.split(" ")[1]
            
        if sheet_name in self.format_df:
            df = self.format_df[sheet_name]
            
            # 保存当前选中的工作表名称和参数列，用于保存功能
            self.current_format_sheet = sheet_name
            
            # 准备数据模型
            validation_data = []
            
            # 确定要显示的参数列
            param_column = 'param1'  # 默认显示param1
            if param_suffix:
                # 从[param1]中提取数字
                param_num = param_suffix.replace('[', '').replace(']', '')
                param_column = param_num  # 使用对应的参数列
            
            # 保存当前参数列
            self.current_format_param_column = param_column
            
            # 遍历DataFrame的每一行，从第3行开始（跳过前两行）
            for idx, row in df.iterrows():
                # 跳过前两行（Excel行号从1开始，标题行是1，所以数据从2开始，跳过前两行就是从4开始）
                if idx < 2:  # 跳过前两行
                    continue
                    
                text = str(row.get(param_column, '')) if row.get(param_column) is not None else ''
                
                # 对文本进行校验
                has_issues = False
                issues_list = []
                
                # 创建一个包含必要信息的字典
                item_data = {
                    'selected': False,
                    'stage_id': row.get('stage_id', ''),
                    'row': idx + 2,  # Excel行号从1开始，标题行是1，所以数据从2开始
                    'speaker': row.get('speaker', ''),
                    'original_text': text,
                    'suggestion': '',  # 初始没有修复建议
                    'type': ''  # 初始没有问题类型
                }
                
                # 使用check_condition_2至check_condition_9进行校验
                check_methods = [
                    (2, self.check_condition_2, self.auto_fix_condition_2),
                    (3, self.check_condition_3, self.auto_fix_condition_3),
                    (4, self.check_condition_4, self.auto_fix_condition_4),
                    (5, self.check_condition_5, self.auto_fix_condition_5),
                    (6, self.check_condition_6, self.auto_fix_condition_6),
                    (7, self.check_condition_7, self.auto_fix_condition_7),
                    (8, self.check_condition_8, self.auto_fix_condition_8),
                    (9, self.check_condition_9, self.auto_fix_condition_9)
                ]
                
                # 创建修改建议文本的副本
                fixed_text = text
                
                for condition_id, check_method, fix_method in check_methods:
                    hit, issues = check_method(text)
                    if hit:
                        has_issues = True
                        # 记录问题类型
                        if item_data['type']:
                            item_data['type'] += f", {condition_id}"
                        else:
                            item_data['type'] = str(condition_id)
                        
                        # 应用自动修复方法
                        fixed_text = fix_method(fixed_text)
                
                # 设置修改建议为自动修复后的文本
                if has_issues:
                    item_data['suggestion'] = fixed_text
                    validation_data.append(item_data)
            
            # 创建新的格式校验表格模型类
            class FormatValidationTableModel(QAbstractTableModel):
                def __init__(self, data):
                    super().__init__()
                    self._data = data
                    self._headers = ["选择", "id", "stage_id", "type", "原文", "修改建议"]
                    # 为type列定义tooltip描述
                self.condition_descriptions = {
                    '1': '条件1：英文相关变量内容与模板相比种类不匹配',
                    '2': '条件2：不应出现中括号中只有数字的情况',
                    '3': '条件3：内容中不应该出现全角或半角冒号及波浪线',
                    '4': '条件4：知识点前后不应该出现双引号，知识点后不能有符号',
                    '5': '条件5：内容的首尾不应出现配对的单引号或反引号',
                    '6': '条件6：内容中不应该出现连续标点符号',
                    '7': '条件7：无方括号的{word_en}或{phrase_en}必须在纯英文语句中，{word_cn}或{phrase_cn}必须不在纯英文语句中',
                    '8': '条件8：半角句号前后两句不应出现中英文混合',
                    '9': '条件9：{words}两侧不能有中括号',
                    '10': '条件10：中括号或大括号两侧的引号必须成对，单侧有引号视为错误'
                }
                
                def rowCount(self, parent=None):
                    return len(self._data)
                
                def columnCount(self, parent=None):
                    return len(self._headers)
                
                def _generate_tooltip_for_type(self, type_value):
                    """为type列的数字生成tooltip"""
                    # 解析type值中的数字（可能是"1 2"这样的组合）
                    numbers = type_value.split()
                    tooltips = []
                    
                    for num in numbers:
                        if num in self.condition_descriptions:
                            tooltips.append(self.condition_descriptions[num])
                    
                    return "\n".join(tooltips) if tooltips else ""
                
                def data(self, index, role=Qt.ItemDataRole.DisplayRole):
                    if not index.isValid():
                        return None
                    
                    row = index.row()
                    col = index.column()
                    
                    if role == Qt.ItemDataRole.DisplayRole:
                        if col == 0:
                            return ""  # 选择列不显示文本
                        elif col == 1:
                            return self._data[row]['row']  # 使用row作为id
                        elif col == 2:
                            return self._data[row]['stage_id']
                        elif col == 3:
                            # 格式化问题ID，用空格分隔
                            if self._data[row]['type']:
                                return " ".join(self._data[row]['type'].split(", "))
                            return ""
                        elif col == 4:
                            return self._data[row]['original_text']
                        elif col == 5:
                            return self._data[row]['suggestion']
                    
                    elif role == Qt.ItemDataRole.CheckStateRole and col == 0:
                        return Qt.CheckState.Checked if self._data[row]['selected'] else Qt.CheckState.Unchecked
                    
                    elif role == Qt.ItemDataRole.ForegroundRole:
                        if col == 3:  # type列使用蓝色
                            return QColor(0, 100, 200)
                        elif col == 4:  # 原文列，如果已修复则使用橙色
                            if self._data[row].get('fixed', False):
                                return QColor(255, 140, 0)  # 橙色
                        elif col == 5:  # 修改建议列使用橙色
                            return QColor(255, 140, 0)  # 橙色
                    
                    elif role == Qt.ItemDataRole.FontRole:
                        if col == 3:  # type列添加下划线
                            font = QFont()
                            font.setUnderline(True)
                            return font
                    
                    elif role == Qt.ItemDataRole.ToolTipRole:
                        if col == 3:  # 为type列添加tooltip
                            type_value = " ".join(self._data[row]['type'].split(", ")) if self._data[row]['type'] else ""
                            return self._generate_tooltip_for_type(type_value)
                        elif col == 4 or col == 5:  # 原文和修改建议列添加tooltip
                            return self._data[row]['original_text'] if col == 4 else self._data[row]['suggestion']
                    
                    return None
                
                def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
                    if not index.isValid():
                        return False
                    
                    row = index.row()
                    col = index.column()
                    
                    if role == Qt.ItemDataRole.CheckStateRole and col == 0:
                        self._data[row]['selected'] = (value == Qt.CheckState.Checked)
                        self.dataChanged.emit(index, index)
                        return True
                    
                    return False
                
                def contextMenuEvent(self, event, view, index):
                    """处理右键菜单事件"""
                    if not index.isValid():
                        return
                        
                    col = index.column()
                    row = index.row()
                    
                    # 只为原文和修改建议列添加右键菜单
                    if col == 4 or col == 5:
                        menu = QMenu()
                        copy_action = menu.addAction("复制")
                        
                        action = menu.exec(event.globalPos())
                        
                        if action == copy_action:
                            # 获取要复制的文本
                            text = self._data[row]['original_text'] if col == 4 else self._data[row]['suggestion']
                            
                            # 复制到剪贴板
                            clipboard = QApplication.clipboard()
                            clipboard.setText(text)
                            
                            # 显示提示
                            from duoki_editor.ui.toast import Toast
                            Toast.show_message("已复制到剪贴板", parent=view)
                
                def flags(self, index):
                    if not index.isValid():
                        return Qt.ItemFlag.NoItemFlags
                    
                    if index.column() == 0:
                        return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable
                    
                    return Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
                
                def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
                    if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
                        return self._headers[section]
                    
                    return None
            
            # 更新格式校验表格的数据模型
            self.format_model = FormatValidationTableModel(validation_data)
            self.format_table.setModel(self.format_model)
            
            # 启用相关按钮
            self.format_save_btn.setEnabled(len(validation_data) > 0)
            self.format_fix_btn.setEnabled(len(validation_data) > 0)
            
            # 更新问题数据行数标签
            self.format_issue_count_label.setText(f"问题数据：{len(validation_data)}行")
            
            # 自动选中修复建议列非空的行
            for row in range(len(validation_data)):
                suggestion = str(validation_data[row].get('suggestion', ''))
                if suggestion.strip():
                    # 设置该行为选中状态
                    self.format_model.setData(
                        self.format_model.index(row, 0),
                        Qt.CheckState.Checked,
                        Qt.ItemDataRole.CheckStateRole
                    )
            
            # 如果没有发现问题数据，显示提示
            if len(validation_data) == 0:
                from duoki_editor.ui.toast import ToastManager
                toast_manager = ToastManager(self)
                toast_manager.show_toast("当前表格未发现问题数据")
    
    def on_table_clicked(self, index):
        """处理表格点击事件"""
        if index.column() == 3:  # type列
            if hasattr(self.validation_model._data, 'iloc') and not self.validation_model._data.empty:
                type_value = str(self.validation_model._data.iloc[index.row()]['type'])
                error_text = self.validation_model._generate_tooltip_for_type(type_value)
                
                # 创建弹窗显示错误条件文本
                dialog = QDialog(self)
                dialog.setWindowTitle("错误条件详情")
                dialog.setFixedSize(400, 200)
                
                layout = QVBoxLayout()
                
                # 添加文本显示
                text_edit = QTextEdit()
                text_edit.setPlainText(error_text)
                text_edit.setReadOnly(True)
                layout.addWidget(text_edit)
                
                # 添加确定按钮
                button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
                button_box.accepted.connect(dialog.accept)
                layout.addWidget(button_box)
                
                dialog.setLayout(layout)
                dialog.exec()

    def show_context_menu(self, position):
        """显示右键菜单"""
        index = self.validation_table.indexAt(position)
        if not index.isValid():
            return
            
        # 在原文列（第4列）和修复建议列（第5列）显示菜单
        if index.column() not in [4, 5]:
            return
            
        menu = QMenu(self)
        
        # 添加复制选项
        copy_action = QAction("复制", self)
        if index.column() == 4:
            # 原文列的复制
            copy_action.triggered.connect(lambda: self.copy_original_cell(index))
        else:
            # 修复建议列的复制
            copy_action.triggered.connect(lambda: self.copy_suggestion_cell(index))
        menu.addAction(copy_action)
        
        # 只在修复建议列（第5列）添加编辑选项
        if index.column() == 5:
            edit_action = QAction("编辑", self)
            edit_action.triggered.connect(lambda: self.edit_suggestion_cell(index))
            menu.addAction(edit_action)
        
        # 显示菜单
        menu.exec(self.validation_table.mapToGlobal(position))
    
    def copy_original_cell(self, index):
        """复制原文单元格内容到剪贴板"""
        if not index.isValid() or index.column() != 4:
            return
            
        # 获取单元格内容
        cell_data = self.validation_model.data(index, Qt.ItemDataRole.DisplayRole)
        if not cell_data:
            cell_data = ""
        
        # 复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(str(cell_data))
        
        # 显示Toast提示
        if hasattr(self, 'toast_manager') and self.toast_manager:
            self.toast_manager.show_toast("已复制到剪贴板")
        else:
            # 如果没有toast管理器，使用MessageBox作为备选
            QMessageBox.information(self, "提示", "已复制到剪贴板")
    
    def copy_suggestion_cell(self, index):
        """复制修复建议单元格内容到剪贴板"""
        if not index.isValid() or index.column() != 5:
            return
            
        # 获取单元格内容
        cell_data = self.validation_model.data(index, Qt.ItemDataRole.DisplayRole)
        if not cell_data:
            cell_data = ""
        
        # 复制到剪贴板
        clipboard = QApplication.clipboard()
        clipboard.setText(str(cell_data))
        
        # 显示Toast提示
        if hasattr(self, 'toast_manager') and self.toast_manager:
            self.toast_manager.show_toast("已复制到剪贴板")
        else:
            # 如果没有toast管理器，使用MessageBox作为备选
            QMessageBox.information(self, "提示", "已复制到剪贴板")
    
    def show_format_context_menu(self, pos):
        """显示格式校验表格的右键菜单"""
        index = self.format_table.indexAt(pos)
        if not index.isValid():
            return
            
        col = index.column()
        
        # 只为原文和修改建议列添加右键菜单
        if col == 4 or col == 5:
            menu = QMenu(self)
            copy_action = menu.addAction("复制")
            
            # 只为修改建议列添加编辑功能
            edit_action = None
            if col == 5:
                edit_action = menu.addAction("编辑")
            
            action = menu.exec(self.format_table.viewport().mapToGlobal(pos))
            
            if action == copy_action:
                # 获取要复制的文本
                text = self.format_model.data(index, Qt.ItemDataRole.DisplayRole)
                
                # 复制到剪贴板
                clipboard = QApplication.clipboard()
                clipboard.setText(str(text))
                
                # 显示提示
                from duoki_editor.ui.toast import ToastManager
                toast_manager = ToastManager(self)
                toast_manager.show_toast("已复制到剪贴板")
            elif action == edit_action:
                self.edit_format_suggestion_cell(index)
    
    def format_quick_fix(self):
        """一键修复格式校验表格中选中行的内容"""
        if not hasattr(self.format_model, '_data') or not self.format_model._data:
            return
            
        # 获取选中的行
        selected_rows = []
        for row in range(len(self.format_model._data)):
            if self.format_model.data(self.format_model.index(row, 0), Qt.ItemDataRole.CheckStateRole) == Qt.CheckState.Checked:
                selected_rows.append(row)
                
        # 应用修复
        fixed_count = 0
        for row in selected_rows:
            # 获取修复建议
            suggestion = self.format_model.data(self.format_model.index(row, 5), Qt.ItemDataRole.DisplayRole)
            if suggestion and str(suggestion).strip():
                # 将修复建议应用到原文
                self.format_model.setData(
                    self.format_model.index(row, 4),  # 原文列
                    suggestion,
                    Qt.ItemDataRole.EditRole
                )
                # 更新数据源中的原文
                if row < len(self.format_model._data):
                    self.format_model._data[row]['original_text'] = suggestion
                    # 设置fixed标记
                    self.format_model._data[row]['fixed'] = True
                fixed_count += 1
                
        # 显示提示
        if fixed_count > 0:
            self.format_model.layoutChanged.emit()  # 强制更新视图
            QMessageBox.information(self, "成功", "已将选中行的非空修复建议应用到原文")
        else:
            QMessageBox.information(self, "提示", "没有选中包含非空修复建议的行")
            
    def save_format_file(self):
        """保存格式校验表格中的原文数据到Excel文件"""
        # 确认对话框
        reply = QMessageBox.question(
            self,
            "确认保存",
            "保存修复内容不可撤掉，请谨慎操作",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
            
        if not hasattr(self, 'current_format_file') or not self.current_format_file:
            QMessageBox.warning(self, "警告", "没有打开的格式校验文件")
            return
            
        if not hasattr(self, 'current_format_sheet') or not self.current_format_sheet:
            QMessageBox.warning(self, "警告", "没有选中的工作表")
            return
            
        if not hasattr(self, 'current_format_param_column') or not self.current_format_param_column:
            QMessageBox.warning(self, "警告", "没有选中的参数列")
            return
            
        try:
            # 读取原始Excel文件
            workbook = openpyxl.load_workbook(self.current_format_file)
            sheet = workbook[self.current_format_sheet]
            
            # 获取表格数据
            if hasattr(self.format_model, '_data') and self.format_model._data:
                for row_data in self.format_model._data:
                    # 获取行号和原文内容
                    excel_row = int(row_data.get('row', 0))
                    original_text = str(row_data.get('original_text', ''))
                    
                    # 获取当前参数列的列号
                    param_col = None
                    for col in range(1, sheet.max_column + 1):
                        if sheet.cell(row=1, column=col).value == self.current_format_param_column:
                            param_col = col
                            break
                    
                    if param_col:
                        # 更新Excel中的数据
                        sheet.cell(row=excel_row, column=param_col).value = original_text
            
            # 保存Excel文件
            workbook.save(self.current_format_file)
            
            Toast.success("已保存修改", self)
            
            # 保存成功后重新校验当前选中的内容
            self.revalidate_current_format_content()
            
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"保存文件时出错: {str(e)}")
    
    def revalidate_current_format_content(self):
        """重新校验当前选中的格式内容"""
        try:
            # 重新读取Excel文件
            if hasattr(self, 'current_format_file') and self.current_format_file:
                self.format_df = pd.read_excel(self.current_format_file, sheet_name=None)
                
                # 获取当前选中的项目
                current_item = self.sheets_list.currentItem()
                if current_item:
                    # 重新触发校验
                    self.on_sheet_item_clicked(current_item)             

        except Exception as e:
            QMessageBox.warning(self, "重新校验失败", f"重新校验时出错: {str(e)}")
            
    def edit_format_suggestion_cell(self, index):
        """编辑格式校验修复建议单元格内容"""
        if not index.isValid() or index.column() != 5:
            return
            
        # 获取当前内容
        current_text = self.format_model.data(index, Qt.ItemDataRole.DisplayRole)
        current_text = str(current_text) if current_text else ""
        
        # 创建编辑对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("编辑修复建议")
        dialog.setModal(True)
        dialog.resize(500, 300)
        
        # 创建布局
        layout = QVBoxLayout(dialog)
        
        # 创建文本编辑框
        text_edit = QTextEdit(dialog)
        text_edit.setPlainText(current_text)
        layout.addWidget(text_edit)
        
        # 创建按钮布局
        button_layout = QHBoxLayout()
        
        # 创建确定和取消按钮
        ok_button = QPushButton("确定", dialog)
        cancel_button = QPushButton("取消", dialog)
        
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        # 连接按钮信号
        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        
        # 显示对话框
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # 获取编辑后的文本
            new_text = text_edit.toPlainText()
            
            # 更新数据模型
            row = index.row()
            self.format_model._data[row]['suggestion'] = new_text
            
            # 通知视图更新
            self.format_model.dataChanged.emit(index, index)
        
    def edit_suggestion_cell(self, index):
        """编辑修复建议单元格内容"""
        if not index.isValid() or index.column() != 5:
            return
            
        # 获取当前内容
        current_text = self.validation_model.data(index, Qt.ItemDataRole.DisplayRole)
        
        # 如果修复建议列为空，则显示原文列的内容
        if not current_text or str(current_text).strip() == "":
            # 获取同一行的原文列内容（第4列）
            original_index = self.validation_model.index(index.row(), 4)
            original_text = self.validation_model.data(original_index, Qt.ItemDataRole.DisplayRole)
            current_text = str(original_text) if original_text else ""
        else:
            current_text = str(current_text)
        
        # 创建编辑对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("编辑修复建议")
        dialog.setModal(True)
        dialog.resize(500, 300)
        
        layout = QVBoxLayout(dialog)
        
        # 添加说明标签
        info_label = QLabel("编辑修复建议内容：")
        layout.addWidget(info_label)
        
        # 添加文本编辑框
        text_edit = QTextEdit()
        text_edit.setPlainText(current_text)
        layout.addWidget(text_edit)
        
        # 添加按钮
        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        # 显示对话框
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_text = text_edit.toPlainText()
            
            # 更新数据模型
            self.validation_model.setData(index, new_text, Qt.ItemDataRole.EditRole)
            
            # 根据文本内容自动勾选或取消勾选
            select_index = self.validation_model.index(index.row(), 0)
            if new_text.strip():  # 如果有文本内容（去除空白字符后长度大于0）
                self.validation_model.setData(select_index, Qt.CheckState.Checked, Qt.ItemDataRole.CheckStateRole)
            else:  # 如果没有文本内容（空文本或只有空白字符）
                self.validation_model.setData(select_index, Qt.CheckState.Unchecked, Qt.ItemDataRole.CheckStateRole)
            
            # 使用toast提示替代弹窗
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast("修复建议已更新")
            else:
                # 如果没有toast管理器，使用MessageBox作为备选
                QMessageBox.information(self, "提示", "修复建议已更新")
    
    def open_files(self):
        """打开Excel文件"""
        # 获取上次打开的路径
        last_path = self.config_manager.get('PATH', 'content_validator_last_open_path', './')
        
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择Excel文件",
            last_path,
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_paths:
            # 检查文件名首字母是否为~
            invalid_files = []
            valid_files = []
            
            for file_path in file_paths:
                file_name = os.path.basename(file_path)
                if file_name.startswith('~'):
                    invalid_files.append(file_name)
                else:
                    valid_files.append(file_path)
            
            # 如果有无效文件，显示警告
            if invalid_files:
                invalid_list = '\n'.join(invalid_files)
                QMessageBox.warning(
                    self,
                    "文件名警告",
                    f"以下文件名首字母不能是'~'，已跳过：\n\n{invalid_list}"
                )
            
            # 处理有效文件
            if valid_files:
                # 清空当前文件列表
                self.clear_files()
                
                # 保存最后打开的路径
                last_dir = os.path.dirname(valid_files[0])
                self.config_manager.set('PATH', 'content_validator_last_open_path', last_dir)
                self.config_manager.save()
                
                for file_path in valid_files:
                    if file_path not in self.file_paths:
                        self.file_paths.append(file_path)
                        # 只显示文件名，但记录完整路径
                        file_name = os.path.basename(file_path)
                        item = QListWidgetItem(file_name)
                        item.setData(Qt.ItemDataRole.UserRole, file_path)  # 存储完整路径
                        self.file_list.addItem(item)
                
                # 启用一键检验与批量修复按钮
                self.batch_validate_btn.setEnabled(True)
                if hasattr(self, 'batch_fix_btn'):
                    self.batch_fix_btn.setEnabled(True)
                
                # 自动执行校验
                # self.validate_files()  # 暂时注释掉自动校验
    
    def add_files(self):
        """添加Excel文件（不清空现有列表）"""
        # 获取上次打开的路径
        last_path = self.config_manager.get('PATH', 'content_validator_last_open_path', './')
        
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择Excel文件",
            last_path,
            "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_paths:
            # 检查文件名首字母是否为~
            invalid_files = []
            valid_files = []
            duplicate_files = []
            
            for file_path in file_paths:
                file_name = os.path.basename(file_path)
                if file_name.startswith('~'):
                    invalid_files.append(file_name)
                elif file_path in self.file_paths:
                    duplicate_files.append(file_name)
                else:
                    valid_files.append(file_path)
            
            # 如果有无效文件，显示警告
            if invalid_files:
                invalid_list = '\n'.join(invalid_files)
                QMessageBox.warning(
                    self,
                    "文件名警告",
                    f"以下文件名首字母不能是'~'，已跳过：\n\n{invalid_list}"
                )
            
            # 如果有重复文件，显示提示
            if duplicate_files:
                duplicate_list = '\n'.join(duplicate_files)
                QMessageBox.information(
                    self,
                    "重复文件提示",
                    f"以下文件已存在于列表中，已跳过：\n\n{duplicate_list}"
                )
            
            # 处理有效文件
            if valid_files:
                # 保存最后打开的路径
                last_dir = os.path.dirname(valid_files[0])
                self.config_manager.set('PATH', 'content_validator_last_open_path', last_dir)
                self.config_manager.save()
                
                # 添加新文件到现有列表（不清空）
                for file_path in valid_files:
                    self.file_paths.append(file_path)
                    # 只显示文件名，但记录完整路径
                    file_name = os.path.basename(file_path)
                    item = QListWidgetItem(file_name)
                    item.setData(Qt.ItemDataRole.UserRole, file_path)  # 存储完整路径
                    self.file_list.addItem(item)
                
                # 启用一键检验与批量修复按钮
                self.batch_validate_btn.setEnabled(True)
                if hasattr(self, 'batch_fix_btn'):
                    self.batch_fix_btn.setEnabled(True)
    
    def on_file_item_clicked(self, item):
        """点击文件列表项时的处理"""
        file_path = item.data(Qt.ItemDataRole.UserRole)
        if not file_path:
            return
            
        # 设置当前文件索引
        self.current_file_index = self.file_list.row(item)
            
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 检查是否包含必需字段
            required_fields = ['param1']
            missing_fields = [field for field in required_fields if field not in df.columns]
            
            if missing_fields:
                # 文件格式不对，显示警告并标记为红色
                QMessageBox.warning(
                    self,
                    "文件格式不对",
                    f"文件 '{os.path.basename(file_path)}' 缺少以下字段：\n{', '.join(missing_fields)}"
                )
                # 将文件名标记为红色
                item.setForeground(QColor(255, 0, 0))  # 红色
                # 禁用按钮
                self.quick_fix_btn.setEnabled(False)
                self.save_btn.setEnabled(False)
                return
            
            # 启用按钮
            self.quick_fix_btn.setEnabled(True)
            self.save_btn.setEnabled(True)
            
            # 开始内容校验
            all_validation_results = []
            
            # 定义要忽略的列
            ignored_columns = ['stage_id', 'stage_name', 'step_name', 'speaker']
            
            # 遍历每一行数据
            for row_index, row in df.iterrows():
                param1_content = row.get('param1', '')
                
                # 遍历所有列（除了param1列本身和忽略的列）
                for column_name in df.columns:
                    if column_name == 'param1' or column_name in ignored_columns:
                        continue
                    
                    cell_content = row[column_name]
                    
                    # 校验单元格内容
                    validation_results = self.validate_cell_content(
                        cell_content, param1_content, column_name, row_index
                    )
                    
                    all_validation_results.extend(validation_results)
            
            # 将校验结果转换为DataFrame并显示
            if all_validation_results:
                validation_df = pd.DataFrame(all_validation_results)
                self.validation_model.load_data(validation_df)
                self.quick_fix_btn.setEnabled(True)
                
                # 检查是否命中条件1
                has_condition_1 = any(re.search(r'\b1\b', str(result.get('type', ''))) for result in all_validation_results)
                
                if has_condition_1:
                    # 命中条件1，设置为红色
                    item.setForeground(QColor('red'))
                else:
                    # 命中其他条件，设置为橙色
                    item.setForeground(QColor('orange'))

            else:
                # 没有发现问题
                empty_df = pd.DataFrame(columns=['selected', 'title', 'row', 'type', 'original_text', 'suggestion'])
                self.validation_model.load_data(empty_df)
                self.quick_fix_btn.setEnabled(False)
                # 文件没有问题，标记为绿色
                item.setForeground(QColor('green'))  # 绿色
                            
                # 如果没有发现问题数据，显示提示
                from duoki_editor.ui.toast import ToastManager
                toast_manager = ToastManager(self)
                toast_manager.show_toast("当前表格未发现问题数据")
                
            # 更新内容质检问题数据行数
            row_count = len(all_validation_results) if hasattr(all_validation_results, '__len__') else 0
            self.content_issue_count_label.setText(f"问题数据：{row_count}行")
                
        except Exception as e:
            # 文件读取错误
            QMessageBox.warning(
                self,
                "文件读取错误",
                f"无法读取文件 '{os.path.basename(file_path)}'：\n{str(e)}"
            )
            # 将文件名标记为红色
            item.setForeground(QColor(255, 0, 0))  # 红色
    
    def validate_files(self):
        """校验文件内容"""
        if not self.file_paths:
            return
        
        # 清空表格数据
        self.validation_model.load_data(pd.DataFrame())
        self.quick_fix_btn.setEnabled(False)
    
    def perform_validation(self, df: pd.DataFrame, file_name: str) -> List[Dict[str, Any]]:
        """执行具体的校验逻辑"""
        validation_results = []
        
        # 定义要忽略的列
        ignored_columns = ['stage_id', 'stage_name', 'step_name', 'speaker']
        
        # 遍历DataFrame的每一行进行验证
        for index, row in df.iterrows():
            # 获取需要验证的列
            for column_name in df.columns:
                if column_name == 'param1' or column_name in ignored_columns:  # 跳过param1列和忽略的列
                    continue
                    
                content = row[column_name]
                param1_content = row.get('param1', '')
                
                # 跳过空内容
                if pd.isna(content) or str(content).strip() == '':
                    continue
                
                # 执行验证
                issues = self.validate_cell_content(str(content), str(param1_content), column_name, index)
                
                # 如果有问题，添加到结果中
                if issues:
                    for issue in issues:
                        validation_results.append({
                            'file_name': file_name,
                            'row_index': index,
                            'column_name': column_name,
                            'content': str(content),
                            'param1': str(param1_content),
                            'type': issue['type'],
                            'suggestion': issue['suggestion']  # 直接使用validate_cell_content生成的建议
                        })
        
        return validation_results
    
    def batch_validate_files(self):
        """一键检验所有文件"""
        if not self.file_paths:
            QMessageBox.information(self, "提示", "请先添加文件")
            return
        
        # 先重置所有文件的颜色为白色
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            item.setForeground(QColor('black'))  # 设置为默认黑色文字
        
        # 遍历所有文件进行检验
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            file_path = item.data(Qt.ItemDataRole.UserRole)
            
            try:
                # 读取Excel文件
                df = pd.read_excel(file_path)
                
                # 检查是否包含必需字段
                required_fields = ['param1']
                missing_fields = [field for field in required_fields if field not in df.columns]
                
                if missing_fields:
                    # 文件格式不正确，设置为红色
                    item.setForeground(QColor('red'))
                    continue
                
                # 执行验证
                validation_results = self.perform_validation(df, os.path.basename(file_path))
                
                # 根据验证结果和条件类型设置文件名颜色
                if validation_results:
                    # 检查是否命中条件1
                    has_condition_1 = any(re.search(r'\b1\b', str(result.get('type', ''))) for result in validation_results)
                    
                    if has_condition_1:
                        # 命中条件1，设置为红色
                        item.setForeground(QColor('red'))
                    else:
                        # 命中其他条件，设置为橙色
                        item.setForeground(QColor('orange'))
                else:
                    # 无问题，设置为绿色
                    item.setForeground(QColor('green'))
                    
            except Exception as e:
                # 读取文件出错，设置为红色
                item.setForeground(QColor('red'))
        
        # 显示完成提示
        QMessageBox.information(self, "提示", "批量检验完成！文件名颜色已更新：\n绿色 - 无问题\n橙色 - 命中其他条件\n红色 - 命中条件1或格式错误")
    
    def clear_files(self):
        """清空文件列表和验证结果"""
        self.file_paths.clear()
        self.file_list.clear()
        self.validation_results = pd.DataFrame()
        self.validation_model.load_data(self.validation_results)
        self.quick_fix_btn.setEnabled(False)  # 使用quick_fix_btn替代fix_btn
        self.save_btn.setEnabled(False)  # 同时禁用保存按钮
        self.batch_validate_btn.setEnabled(False)
        if hasattr(self, 'batch_fix_btn'):
            self.batch_fix_btn.setEnabled(False)
    
    def select_config_directory(self):
        """选择配置输出目录"""
        current_dir = self.config_dir_display.text() or './output'
        directory = QFileDialog.getExistingDirectory(
            self,
            "选择配置输出目录",
            current_dir
        )
        
        if directory:
            # 确保显示绝对路径
            absolute_path = os.path.abspath(directory)
            self.config_dir_display.setText(absolute_path)
            # 保存到配置文件
            self.config_manager.set('PATH', 'speech_config_output', directory)
            self.config_manager.save()
    
    def open_config_directory(self):
        """打开配置目录"""
        directory = self.config_dir_display.text()
        if not directory:
            QMessageBox.warning(self, "警告", "请先选择配置目录")
            return
        
        if not os.path.exists(directory):
            # 如果目录不存在，询问是否创建
            reply = QMessageBox.question(
                self,
                "目录不存在",
                f"目录 '{directory}' 不存在，是否创建？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    os.makedirs(directory, exist_ok=True)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"创建目录失败：{str(e)}")
                    return
            else:
                return
        
        # 打开目录
        try:
            if os.name == 'nt':  # Windows
                os.startfile(directory)
            elif os.name == 'posix':  # macOS and Linux
                os.system(f'open "{directory}"')
        except Exception as e:
            QMessageBox.critical(self, "错误", f"打开目录失败：{str(e)}")
    
    def generate_config_file(self):
        """生成配置文件"""
        # 自动切换到数据校验页签
        self.tab_widget.setCurrentIndex(2)  # 切换到第二个页签（数据校验）
        
        # 清空数据校验文本框并显示开始信息
        self.validation_text.clear()
        self.validation_text.append("开始生成配置文件...")
        
        directory = self.config_dir_display.text()
        if not directory:
            self.validation_text.append("错误：请先选择配置目录")
            QMessageBox.warning(self, "警告", "请先选择配置目录")
            return
        
        # 确保目录存在
        if not os.path.exists(directory):
            try:
                os.makedirs(directory, exist_ok=True)
                self.validation_text.append(f"创建目录：{directory}")
            except Exception as e:
                error_msg = f"创建目录失败：{str(e)}"
                self.validation_text.append(f"错误：{error_msg}")
                QMessageBox.critical(self, "错误", error_msg)
                return
        
        # 获取文件列表中的所有xlsx文件
        excel_files = []
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            # 获取存储在UserRole中的完整文件路径
            file_path = item.data(Qt.ItemDataRole.UserRole)
            if file_path and file_path.endswith('.xlsx') and not os.path.basename(file_path).startswith('~'):
                excel_files.append(file_path)
        
        if not excel_files:
            error_msg = "文件列表中没有找到xlsx文件"
            self.validation_text.append(f"错误：{error_msg}")
            QMessageBox.warning(self, "警告", error_msg)
            return
        
        self.validation_text.append(f"找到 {len(excel_files)} 个Excel文件")

        do_speech = self.speech_checkbox.isChecked() if hasattr(self, 'speech_checkbox') else True
        do_template = self.template_checkbox.isChecked() if hasattr(self, 'template_checkbox') else True
        if not do_speech and not do_template:
            if hasattr(self, 'toast_manager') and self.toast_manager:
                self.toast_manager.show_toast("请至少选择生成一个文件")
            else:
                QMessageBox.warning(self, "提示", "请至少选择生成一个文件")
            return

        try:
            # 调用生成配置文件的方法
            self.generate_speech_config(excel_files, directory, do_speech, do_template)
            files_list = []
            if do_speech:
                files_list.append("Speech_ai.xlsx")
            if do_template:
                files_list.append("TemplateConfig_ai.xlsx")
            success_msg = "配置文件已生成到：" + directory + "\n\n生成的文件：\n- " + "\n- ".join(files_list)
            self.validation_text.append("生成完成！")
            self.validation_text.append(success_msg)
            QMessageBox.information(self, "成功", success_msg)
        except Exception as e:
            error_msg = f"生成配置文件失败：{str(e)}"
            self.validation_text.append(f"错误：{error_msg}")
            QMessageBox.critical(self, "错误", error_msg)

    def generate_speech_config(self, excel_files, output_dir, generate_speech=True, generate_template=True):
        """生成语音配置文件，调用post_process.py的统一接口"""
        
        # 设置输出重定向
        text_redirector = TextRedirector(self.validation_text)
        logging_handler = LoggingHandler(self.validation_text)
        
        # 保存原始的stdout和stderr
        original_stdout = sys.stdout
        original_stderr = sys.stderr
        
        # 保存原始的logging配置
        original_handlers = logging.root.handlers[:]
        
        try:
            # 重定向stdout和stderr
            sys.stdout = text_redirector
            sys.stderr = text_redirector
            
            # 添加自定义的logging处理器
            logging.root.addHandler(logging_handler)
            
            # 在文本框中显示开始信息
            self.validation_text.append('<span style="color: orange; font-weight: bold;">开始生成配置文件...</span>')
            self.validation_text.append("=" * 50)
            QApplication.processEvents()  # 刷新界面
            
            # 定义进度回调函数
            def progress_callback(message):
                self.validation_text.append(f'<span style="color: orange;">{message}</span>')
                QApplication.processEvents()  # 刷新界面
            
            # 调用post_process.py中的统一接口
            from duoki_editor.utils.post_process import generate_config_files
            
            # 设置保存数据目录路径
            save_data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'resources', 'data', 'save')
            
            success = generate_config_files(input_files=excel_files, output_directory=output_dir, save_data_dir=save_data_dir, progress_callback=progress_callback, generate_speech=generate_speech, generate_template=generate_template)
            
            if not success:
                raise Exception("配置文件生成失败")
            
            # 在文本框中显示完成信息
            self.validation_text.append("=" * 50)
            self.validation_text.append('<span style="color: green; font-weight: bold;">配置文件生成完成！</span>')
            self.validation_text.append(f'<span style="color: green;">输出目录: {output_dir}</span>')
            QApplication.processEvents()  # 刷新界面
            
        except Exception as e:
            # 在文本框中显示错误信息
            self.validation_text.append("=" * 50)
            self.validation_text.append(f'<span style="color: red; font-weight: bold;">生成配置文件时出错: {str(e)}</span>')
            QApplication.processEvents()  # 刷新界面
            raise e
            
        finally:
            # 恢复原始的stdout和stderr
            sys.stdout = original_stdout
            sys.stderr = original_stderr
            
            # 恢复原始的logging配置
            logging.root.handlers = original_handlers

    def export_validation_results(self):
        """导出校验结果到txt文件"""
        try:
            # 获取文本框内容
            content = self.validation_text.toPlainText()
            
            if not content.strip():
                QMessageBox.information(self, "提示", "文本框内容为空，无法导出")
                return
            
            # 打开文件保存对话框
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出校验结果",
                "validation_results.txt",
                "文本文件 (*.txt);;所有文件 (*)"
            )
            
            if file_path:
                # 保存文件
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                
                QMessageBox.information(self, "成功", f"校验结果已导出到：\n{file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败：{str(e)}")

    def perform_data_validation(self):
        """执行数据校验"""
        try:
            # 获取data_output_directory
            config_manager = ConfigManager()
            data_output_directory = config_manager.get('PATH', 'data_output_directory')
            
            if not data_output_directory:
                QMessageBox.warning(self, "警告", "请先设置数据输出目录")
                return
            
            # 检查文件是否存在
            template_config_path = os.path.join(data_output_directory, "TemplateConfig_ai.xlsx")
            speech_ai_path = os.path.join(data_output_directory, "Speech_ai.xlsx")
            
            missing_files = []
            if not os.path.exists(template_config_path):
                missing_files.append("TemplateConfig_ai.xlsx")
            if not os.path.exists(speech_ai_path):
                missing_files.append("Speech_ai.xlsx")
            
            if missing_files:
                for file_name in missing_files:
                    QMessageBox.warning(self, "文件不存在", f"文件 {file_name} 不存在于目录：\n{data_output_directory}")
                return
            
            # 禁用UI
            self._disable_ui_during_validation()
            
            # 清空文本框
            self.validation_text.clear()
            
            # 开始校验
            self.validation_text.append("开始数据校验...")
            self.validation_text.append("=" * 50)
            
            # 校验TemplateConfig_ai.xlsx
            if os.path.exists(template_config_path):
                self._validate_template_config(template_config_path)
            
            # 校验Speech_ai.xlsx
            if os.path.exists(speech_ai_path):
                self._validate_speech_ai(speech_ai_path, template_config_path)
            
            self.validation_text.append("=" * 50)
            self.validation_text.append("数据校验完成！")
            
        except Exception as e:
            self.validation_text.append(f"校验过程中出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"校验失败：{str(e)}")
        finally:
            # 恢复UI
            self._enable_ui_after_validation()

    def _validate_template_config(self, file_path):
        """校验TemplateConfig_ai.xlsx"""
        try:
            self.validation_text.append("TemplateConfig_ai.xlsx开始判定")
            self.validation_text.append("-" * 30)
            QApplication.processEvents()  # 刷新界面
            
            # 加载Excel文件
            from duoki_editor.utils.excel_handler import ExcelHandler
            excel_handler = ExcelHandler()
            template_data = excel_handler.load_excel(file_path)
            
            if not template_data:
                self.validation_text.append("错误：无法读取TemplateConfig_ai.xlsx文件")
                return
            
            # 获取SceneGraphManager实例
            from duoki_editor.core.scene_graph_manager import SceneGraphManager
            scene_manager = SceneGraphManager()
            scene_graph_data = scene_manager.get_scene_graph_origin_data()
            
            if scene_graph_data.empty:
                self.validation_text.append("错误：SceneGraph数据为空")
                return
            
            # 背景图有效性校验
            self.validation_text.append("背景图有效性判定：")
            QApplication.processEvents()  # 刷新界面
            
            error_count = 0
            for sheet_name, sheet_data in template_data.items():
                if sheet_data.empty:
                    continue
                
                # 检查是否有sceneId列
                if 'sceneId' not in sheet_data.columns:
                    continue
                
                # 实时输出当前处理的sheet
                self.validation_text.append(f"正在校验sheet: {sheet_name}")
                QApplication.processEvents()  # 刷新界面
                
                # 按sheet名称过滤SceneGraph数据
                # 使用SceneGraphManager的方法获取特定sheet的数据
                sheet_scene_data = scene_manager.get_origin_data_by_sheet_name(sheet_name)
                
                if sheet_scene_data.empty:
                    self.validation_text.append(f"警告：SceneGraph中没有找到名为'{sheet_name}'的sheet数据")
                    QApplication.processEvents()  # 刷新界面
                
                # 遍历每一行数据
                for index, row in sheet_data.iterrows():
                    scene_id = row.get('sceneId')
                    stage_id = row.get('stage_id', '')
                    
                    if pd.isna(scene_id):
                        continue
                    
                    scene_id_str = str(scene_id)
                    
                    # 在同名sheet的SceneGraph数据中查找匹配的scene_id
                    matching_rows = sheet_scene_data[sheet_scene_data['scene_id'] == scene_id_str]
                    
                    if matching_rows.empty:
                        # 未找到匹配，输出错误信息
                        error_msg = f"[{sheet_name}] 第{index + 2}行：[{stage_id}] 的 [{scene_id_str}] 在 SceneGraph 中无法找到对应模板的场景数据"
                        self._append_colored_text(error_msg)
                        error_count += 1
                        QApplication.processEvents()  # 刷新界面
            
            if error_count == 0:
                self.validation_text.append("✅ 背景图有效性校验通过")
            else:
                self.validation_text.append(f"❌ 背景图有效性校验发现 {error_count} 个错误")
            QApplication.processEvents()  # 刷新界面
                
        except Exception as e:
            self.validation_text.append(f"TemplateConfig_ai.xlsx校验出错: {str(e)}")
            QApplication.processEvents()  # 刷新界面

    def _validate_speech_ai(self, speech_file_path, template_file_path):
        """校验Speech_ai.xlsx"""
        try:
            self.validation_text.append("\nSpeech_ai.xlsx开始判定")
            self.validation_text.append("-" * 30)
            QApplication.processEvents()  # 刷新界面
            
            # 加载Excel文件
            from duoki_editor.utils.excel_handler import ExcelHandler
            excel_handler = ExcelHandler()
            speech_data = excel_handler.load_excel(speech_file_path)
            template_data = excel_handler.load_excel(template_file_path)
            
            if not speech_data:
                self.validation_text.append("错误：无法读取Speech_ai.xlsx文件")
                return
            
            if not template_data:
                self.validation_text.append("错误：无法读取TemplateConfig_ai.xlsx文件")
                return
            
            # 获取SceneGraphManager实例
            from duoki_editor.core.scene_graph_manager import SceneGraphManager
            scene_manager = SceneGraphManager()
            
            # 环节有效性校验
            self.validation_text.append("环节有效性判定：")
            QApplication.processEvents()  # 刷新界面
            
            error_count = 0
            for sheet_name, sheet_data in speech_data.items():
                if sheet_data.empty:
                    continue
                
                # 检查必要的列
                if 'stage_id' not in sheet_data.columns or 'stage_name' not in sheet_data.columns:
                    continue
                
                # 实时输出当前处理的sheet
                self.validation_text.append(f"正在校验sheet: {sheet_name}")
                QApplication.processEvents()  # 刷新界面
                
                # 获取同名sheet的TemplateConfig数据
                template_sheet_data = template_data.get(sheet_name)
                if template_sheet_data is None or template_sheet_data.empty:
                    self.validation_text.append(f"警告：TemplateConfig_ai中没有找到名为'{sheet_name}'的sheet数据")
                    QApplication.processEvents()  # 刷新界面
                    continue
                
                # 构建当前sheet的stage_id到sceneId的映射
                sheet_stage_to_scene_map = {}
                if 'stage_id' in template_sheet_data.columns and 'sceneId' in template_sheet_data.columns:
                    for _, row in template_sheet_data.iterrows():
                        stage_id = row.get('stage_id')
                        scene_id = row.get('sceneId')
                        if not pd.isna(stage_id) and not pd.isna(scene_id):
                            sheet_stage_to_scene_map[str(stage_id)] = str(scene_id)
                
                # 获取当前sheet的SceneGraph数据
                sheet_scene_data = scene_manager.get_data_by_sheet_name(sheet_name)
                if sheet_scene_data.empty:
                    self.validation_text.append(f"警告：SceneGraph中没有找到名为'{sheet_name}'的sheet数据")
                    QApplication.processEvents()  # 刷新界面
                
                # 遍历每一行数据
                for index, row in sheet_data.iterrows():
                    stage_id = row.get('stage_id')
                    stage_name = row.get('stage_name')
                    
                    if pd.isna(stage_id) or pd.isna(stage_name):
                        continue
                    
                    stage_id_str = str(stage_id)
                    stage_name_str = str(stage_name)
                    
                    # 在同名sheet的TemplateConfig中查找对应的stage_id
                    scene_id = sheet_stage_to_scene_map.get(stage_id_str)
                    
                    if not scene_id:
                        # 在同名sheet的TemplateConfig中找不到对应的stage_id
                        error_msg = f"[{sheet_name}] 第{index + 2}行：[{stage_id_str}] 在 TemplateConfig_ai 中无法找到对应的场景数据"
                        self._append_colored_text(error_msg)
                        error_count += 1
                        QApplication.processEvents()  # 刷新界面
                        continue
                    
                    # 在同名sheet的SceneGraph中查找匹配的scene_id
                    scene_rows = sheet_scene_data[sheet_scene_data['scene_id'] == scene_id]
                    
                    if scene_rows.empty:
                        error_msg = f"[{sheet_name}] 第{index + 2}行：[{stage_id_str}] 的 [{stage_name_str}] 环节 在 SceneGraph 中的 [{scene_id}] 下无法找到对应环节的场景数据"
                        self._append_colored_text(error_msg)
                        error_count += 1
                        QApplication.processEvents()  # 刷新界面
                        continue
                    
                    # 检查stage_name是否在SceneGraph的相关行中存在
                    # 这里需要根据具体的SceneGraph结构来判断
                    # 假设SceneGraph中有stage_name或类似的列
                    stage_name_found = False
                    for scene_col in scene_rows.columns:
                        if scene_rows[scene_col].astype(str).str.contains(stage_name_str, na=False).any():
                            stage_name_found = True
                            break
                    
                    if not stage_name_found:
                        error_msg = f"[{sheet_name}] 第{index + 2}行：[{stage_id_str}] 的 [{stage_name_str}] 环节 在 SceneGraph 中的 [{scene_id}] 下无法找到对应环节的场景数据"
                        self._append_colored_text(error_msg)
                        error_count += 1
                        QApplication.processEvents()  # 刷新界面
            
            if error_count == 0:
                self.validation_text.append("✅ 环节有效性校验通过")
            else:
                self.validation_text.append(f"❌ 环节有效性校验发现 {error_count} 个错误")
            QApplication.processEvents()  # 刷新界面
                
        except Exception as e:
            self.validation_text.append(f"Speech_ai.xlsx校验出错: {str(e)}")
            QApplication.processEvents()  # 刷新界面

    def _append_colored_text(self, text):
        """添加带颜色的文本到validation_text"""
        # 使用HTML格式来显示橙色的方括号内容
        import re
        
        # 将方括号内的内容标记为橙色
        colored_text = re.sub(r'\[([^\]]+)\]', r'<span style="color: orange;">[\1]</span>', text)
        
        # 添加到文本框
        self.validation_text.append(colored_text)

    def _disable_ui_during_validation(self):
        """校验过程中禁用相关UI控件"""
        self.export_results_button.setEnabled(False)
        self.generate_config_btn.setEnabled(False)
        self.batch_validate_btn.setEnabled(False)
        self.tab_widget.setTabEnabled(0, False)  # 禁用内容质检页签
        
    def _enable_ui_after_validation(self):
        """校验结束后恢复UI控件"""
        self.export_results_button.setEnabled(True)
        self.generate_config_btn.setEnabled(True)
        self.batch_validate_btn.setEnabled(True)
        self.tab_widget.setTabEnabled(0, True)  # 启用内容质检页签
        
    def quick_fix_selected(self):
        """一键修复选中行的功能"""
        # 获取选中的行
        selected_rows = self.validation_model.get_selected_rows()
        if not selected_rows:
            QMessageBox.information(self, "提示", "请先选择需要修复的行")
            return
            
        # 获取当前文件索引
        if not hasattr(self, 'current_file_index') or self.current_file_index is None:
            QMessageBox.warning(self, "警告", "请先选择一个文件")
            return
            
        # 修改数据
        modified = False
        for row_idx in selected_rows:
            if row_idx < len(self.validation_model._data):
                # 获取修复建议
                suggestion = self.validation_model._data.iloc[row_idx]['suggestion']
                
                # 只有当修复建议非空时才替换
                if suggestion and not pd.isna(suggestion) and str(suggestion).strip():
                    # 更新原文
                    self.validation_model._data.at[self.validation_model._data.index[row_idx], 'original_text'] = suggestion
                    # 设置fixed标记
                    self.validation_model._data.at[self.validation_model._data.index[row_idx], 'fixed'] = True
                    modified = True
        
        if modified:
            # 刷新表格显示
            self.validation_model.layoutChanged.emit()
            Toast.success("已将选中行的非空修复建议应用到原文", self)
        else:
            Toast.info("选中的行没有可用的修复建议", self)
            
    def save_to_xlsx(self):
        """保存修改到原xlsx文件"""
        # 确认对话框
        reply = QMessageBox.question(
            self,
            "确认保存",
            "保存修复内容不可撤掉，请谨慎操作",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
            
        # 获取当前文件索引
        if not hasattr(self, 'current_file_index') or self.current_file_index is None:
            QMessageBox.warning(self, "警告", "请先选择一个文件")
            return
            
        # 获取当前文件路径
        if self.current_file_index >= len(self.file_paths):
            QMessageBox.warning(self, "警告", "文件索引无效")
            return
            
        file_path = self.file_paths[self.current_file_index]
        
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 获取修改后的数据
            modified_data = self.validation_model._data
            
            # 应用修改
            for _, row in modified_data.iterrows():
                excel_row = row['row'] - 2  # 表头算第1行，数据从第2行开始
                column_name = row['title']
                
                if 0 <= excel_row < len(df) and column_name in df.columns:
                    # 更新Excel中的单元格内容
                    df.at[excel_row, column_name] = row['original_text']
            
            # 保存回Excel文件
            df.to_excel(file_path, index=False)
            
            # 重新加载文件并刷新表格 - 直接调用on_file_item_clicked方法
            self.on_file_item_clicked(self.file_list.item(self.current_file_index))
            
            Toast.success("已保存修改", self)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：{str(e)}")
            print(f"保存失败：{str(e)}")

    def try_fix_validation_errors(self):
        """尝试修复校验错误"""
        try:
            # 获取data_output_directory
            config_manager = ConfigManager()
            data_output_directory = config_manager.get('PATH', 'data_output_directory')
            
            if not data_output_directory:
                QMessageBox.warning(self, "警告", "请先设置数据输出目录")
                return
            
            # 检查TemplateConfig_ai.xlsx文件是否存在
            template_config_path = os.path.join(data_output_directory, "TemplateConfig_ai.xlsx")
            
            if not os.path.exists(template_config_path):
                QMessageBox.warning(self, "文件不存在", f"文件 TemplateConfig_ai.xlsx 不存在于目录：\n{data_output_directory}")
                return
            
            # 立即创建备份文件
            import shutil
            from datetime import datetime
            
            # 生成备份文件名（格式：_yymmddhhmm）
            timestamp = datetime.now().strftime("%y%m%d%H%M")
            backup_path = template_config_path.replace('.xlsx', f'_{timestamp}.xlsx')
            
            # 复制原文件作为备份（覆盖同名文件）
            shutil.copy2(template_config_path, backup_path)
            
            # 禁用UI
            self._disable_ui_during_validation()
            
            # 清空文本框
            self.validation_text.clear()
            
            # 开始修复
            self.validation_text.append("开始尝试修复...")
            self.validation_text.append(f"✅ 已创建备份文件: {os.path.basename(backup_path)}")
            self.validation_text.append("=" * 50)
            QApplication.processEvents()
            
            # 执行修复
            self._perform_template_config_fix(template_config_path)
            
            self.validation_text.append("=" * 50)
            self.validation_text.append("尝试修复完成！")
            
        except Exception as e:
            self.validation_text.append(f"修复过程中出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"修复失败：{str(e)}")
        finally:
            # 恢复UI
            self._enable_ui_after_validation()

    def batch_auto_fix_files(self):
        if not getattr(self, 'file_paths', None):
            from duoki_editor.ui.toast import Toast
            Toast.info("请先添加文件", self)
            return
        reply = QMessageBox.question(
            self,
            "确认批量修复",
            "将对所有文件进行批量修复并覆盖原文件，是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            file_path = item.data(Qt.ItemDataRole.UserRole)
            file_name = os.path.basename(file_path)
            total_fixed = 0
            attempts = 0
            while attempts < 10:
                attempts += 1
                df = pd.read_excel(file_path)
                if 'param1' not in df.columns:
                    item.setForeground(QColor('red'))
                    print(f"{file_name}：缺少必要字段param1，跳过")
                    break
                results = self.perform_validation(df, os.path.basename(file_path))
                if not results:
                    item.setForeground(QColor('green'))
                    print(f"{file_name}：无问题跳过")
                    break
                types = [str(r.get('type', '')).strip() for r in results]
                only_cond1 = bool(types) and all(
                    (len(re.findall(r'\d+', t)) > 0 and all(num == '1' for num in re.findall(r'\d+', t)))
                    for t in types
                )
                if only_cond1:
                    item.setForeground(QColor('red'))
                    print(f"{file_name}：只有条件1问题跳过")
                    break
                item.setForeground(QColor('orange'))
                fixed_this_round = 0
                for r in results:
                    if str(r.get('type', '')).strip() == '1':
                        continue
                    suggestion = r.get('suggestion')
                    if suggestion and str(suggestion).strip():
                        row_idx = int(r.get('row_index'))
                        col = str(r.get('column_name'))
                        if col in df.columns and 0 <= row_idx < len(df):
                            df.at[row_idx, col] = suggestion
                            fixed_this_round += 1
                if fixed_this_round == 0:
                    print(f"{file_name}：无法应用修复建议，跳过")
                    break
                try:
                    df.to_excel(file_path, index=False)
                except Exception as e:
                    from duoki_editor.ui.toast import Toast
                    Toast.error(f"保存失败：{str(e)}", self)
                    print(f"保存失败：{file_name}：{str(e)}")
                    return
                total_fixed += fixed_this_round
            if total_fixed > 0:
                print(f"{file_name}：修复{total_fixed}条错误")
        from duoki_editor.ui.toast import Toast
        Toast.success("修复完成", self)

    def _perform_template_config_fix(self, file_path):
        """执行TemplateConfig_ai.xlsx的修复"""
        try:
            self.validation_text.append("TemplateConfig_ai.xlsx开始修复")
            self.validation_text.append("-" * 30)
            QApplication.processEvents()
            
            # 加载Excel文件
            from duoki_editor.utils.excel_handler import ExcelHandler
            excel_handler = ExcelHandler()
            template_data = excel_handler.load_excel(file_path)
            
            if not template_data:
                self.validation_text.append("错误：无法读取TemplateConfig_ai.xlsx文件")
                return
            
            # 获取SceneGraphManager实例
            from duoki_editor.core.scene_graph_manager import SceneGraphManager
            scene_manager = SceneGraphManager()
            scene_graph_data = scene_manager.get_scene_graph_data()
            
            if scene_graph_data.empty:
                self.validation_text.append("错误：SceneGraph数据为空")
                return
            
            # 背景图有效性修复
            self.validation_text.append("背景图有效性修复：")
            QApplication.processEvents()
            
            fix_count = 0
            modified_sheets = {}  # 记录修改的sheet数据
            
            for sheet_name, sheet_data in template_data.items():
                if sheet_data.empty:
                    continue
                
                # 检查是否有sceneId列
                if 'sceneId' not in sheet_data.columns:
                    continue
                
                # 实时输出当前处理的sheet
                self.validation_text.append(f"正在修复sheet: {sheet_name}")
                QApplication.processEvents()
                
                # 获取当前sheet的SceneGraph数据
                sheet_scene_data = scene_manager.get_data_by_sheet_name(sheet_name)
                
                if sheet_scene_data.empty:
                    self.validation_text.append(f"警告：SceneGraph中没有找到名为'{sheet_name}'的sheet数据")
                    QApplication.processEvents()
                    continue
                
                # 获取有效的scene_id列表
                valid_scene_ids = set(sheet_scene_data['scene_id'].astype(str).tolist())
                
                # 复制sheet数据用于修改
                modified_sheet_data = sheet_data.copy()
                sheet_modified = False
                
                # 遍历每一行进行修复
                for index, row in sheet_data.iterrows():
                    scene_id = str(row['sceneId']) if pd.notna(row['sceneId']) else ""
                    
                    if scene_id and scene_id not in valid_scene_ids:
                        # 尝试修复
                        fixed_scene_id = self._try_fix_scene_id(scene_id, sheet_name, valid_scene_ids, scene_manager)
                        
                        if fixed_scene_id and fixed_scene_id != scene_id:
                            # 修复成功，更新数据
                            modified_sheet_data.loc[index, 'sceneId'] = fixed_scene_id
                            sheet_modified = True
                            fix_count += 1
                            
                            # 获取stage_id用于输出
                            stage_id = str(row.get('stage_id', '未知')) if 'stage_id' in row else '未知'
                            
                            # 输出修复信息
                            self._append_colored_text(f"[{sheet_name}] 第{index + 2}行：[{stage_id}] 的 [{scene_id}] 在 SceneGraph 中未找到完全匹配项 但可以找到 [{fixed_scene_id}] 并成功进行替换")
                            QApplication.processEvents()
                
                # 如果sheet有修改，记录下来
                if sheet_modified:
                    modified_sheets[sheet_name] = modified_sheet_data
            
            # 如果有修改，保存文件
            if modified_sheets:
                self._save_modified_excel(file_path, template_data, modified_sheets)
                self.validation_text.append(f"修复完成！共修复 {fix_count} 个错误")
            else:
                self.validation_text.append("没有发现可以修复的错误")
                
        except Exception as e:
            self.validation_text.append(f"TemplateConfig_ai.xlsx修复出错: {str(e)}")
            QApplication.processEvents()

    def _try_fix_scene_id(self, scene_id, sheet_name, valid_scene_ids, scene_manager):
        """
        尝试修复scene_id
        
        Args:
            scene_id (str): 原始scene_id
            sheet_name (str): sheet名称
            valid_scene_ids (set): 有效的scene_id集合
            scene_manager: SceneGraphManager实例
            
        Returns:
            str or None: 修复后的scene_id，如果无法修复则返回None
        """
        # 预处理：将所有"_"替换成"-"
        processed_scene_id = scene_id.replace('_', '-')
        
        # 第一次尝试：添加前缀
        try:
            # 使用SceneGraphManager的_add_scene_id_prefix方法
            import pandas as pd
            temp_df = pd.DataFrame({'scene_id': [processed_scene_id]})
            prefixed_df = scene_manager._add_scene_id_prefix(temp_df, sheet_name)
            
            if not prefixed_df.empty and 'scene_id' in prefixed_df.columns:
                prefixed_scene_id = str(prefixed_df.iloc[0]['scene_id'])
                # 检查添加前缀后是否在有效列表中
                if prefixed_scene_id in valid_scene_ids and prefixed_scene_id != scene_id:
                    return prefixed_scene_id
        except Exception as e:
            print(f"第一次修复尝试失败: {e}")
        
        # 第二次尝试：在第一次的基础上，将最后一个元素替换为"通用"
        try:
            # 先添加前缀
            import pandas as pd
            temp_df = pd.DataFrame({'scene_id': [processed_scene_id]})
            prefixed_df = scene_manager._add_scene_id_prefix(temp_df, sheet_name)
            
            if not prefixed_df.empty and 'scene_id' in prefixed_df.columns:
                prefixed_scene_id = str(prefixed_df.iloc[0]['scene_id'])
                
                # 用"-"分割，将最后一个元素替换为"通用"
                parts = prefixed_scene_id.split('-')
                if len(parts) > 1:
                    parts[-1] = '通用'
                    generic_scene_id = '-'.join(parts)
                    
                    # 检查替换为"通用"后是否在有效列表中
                    if generic_scene_id in valid_scene_ids and generic_scene_id != scene_id:
                        return generic_scene_id
        except Exception as e:
            print(f"第二次修复尝试失败: {e}")
        
        return None

    def _save_modified_excel(self, file_path, original_data, modified_sheets):
        """
        保存修改后的Excel文件
        
        Args:
            file_path (str): 原始文件路径
            original_data (dict): 原始数据字典 {sheet_name: DataFrame}
            modified_sheets (dict): 修改后的sheet数据 {sheet_name: DataFrame}
        """
        try:
            # 直接保存修改后的数据到原文件
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, sheet_data in original_data.items():
                    if sheet_name in modified_sheets:
                        # 使用修改后的数据
                        modified_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
                        self.validation_text.append(f"✅ 已更新sheet: {sheet_name}")
                    else:
                        # 使用原始数据
                        sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
                    QApplication.processEvents()
            
            self.validation_text.append(f"✅ 文件修复完成并已保存: {os.path.basename(file_path)}")
            QApplication.processEvents()
            
        except Exception as e:
            error_msg = f"❌ 保存文件失败: {str(e)}"
            self.validation_text.append(error_msg)
            QApplication.processEvents()
            # 如果保存失败，也要通知用户
            QMessageBox.critical(self, "保存失败", f"无法保存修复后的文件：\n{str(e)}")
            raise

    def open_data_split_file(self):
        """打开数据拆解文件"""
        # 从配置文件获取上次打开路径
        config = ConfigParser()
        config.read(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), encoding='utf-8')
        
        last_path = ""
        if config.has_section('PATH') and config.has_option('PATH', 'data_split_last_open_path'):
            last_path = config.get('PATH', 'data_split_last_open_path')
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "打开Excel文件", last_path, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # 保存路径到配置文件
                config = ConfigParser()
                config.read(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), encoding='utf-8')
                
                if not config.has_section('PATH'):
                    config.add_section('PATH')
                
                config.set('PATH', 'data_split_last_open_path', os.path.dirname(file_path))
                with open(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'config.ini'), 'w', encoding='utf-8') as f:
                    config.write(f)
                
                # 读取Excel文件
                self.data_split_df = pd.read_excel(file_path, sheet_name=None)
                self.current_data_split_file = file_path
                
                # 更新文件路径显示
                file_name = os.path.basename(file_path)
                self.data_split_file_path_label.setText(f"已选择: {file_name}")
                self.data_split_file_path_label.setStyleSheet("color: #4CAF50; margin-left: 10px;")
                self.data_split_file_path_label.setToolTip(file_path)
                
                # 检查是否存在必要的列
                required_columns = ['stage_id', 'param1', 'param2', 'param3', 'param4', 'param5']
                missing_columns = []
                
                # 检查每个sheet是否包含必要的列
                valid_sheets = []
                for sheet_name, df in self.data_split_df.items():
                    sheet_columns = df.columns.tolist()
                    current_missing = [col for col in required_columns if col not in sheet_columns]
                    
                    if not current_missing:
                        valid_sheets.append(sheet_name)
                    else:
                        missing_columns.extend(current_missing)
                
                # 如果没有有效的sheet，显示警告
                if not valid_sheets:
                    missing_columns = list(set(missing_columns))  # 去重
                    QMessageBox.warning(
                        self, 
                        "表格数据不符合要求", 
                        f"缺少必要的列: {', '.join(required_columns)}\n当前缺少: {', '.join(missing_columns)}"
                    )
                    self.data_split_process_button.setEnabled(False)
                else:
                    # 如果有有效的sheet，启用处理按钮
                    self.data_split_process_button.setEnabled(True)
                    self.data_split_text.clear()
                    self.data_split_text.append(f"✅ 文件加载成功: {os.path.basename(file_path)}")
                    self.data_split_text.append(f"📊 有效工作表: {', '.join(valid_sheets)}")
                    
            except Exception as e:
                QMessageBox.critical(self, "错误", f"打开文件时发生错误: {str(e)}")
                self.data_split_process_button.setEnabled(False)

    def process_data_split(self):
        """处理数据拆解"""
        if not hasattr(self, 'data_split_df') or not self.data_split_df:
            QMessageBox.warning(self, "警告", "请先打开一个有效的Excel文件")
            return
        
        try:
            # 获取选择的处理模式
            if self.full_mode_radio.isChecked():
                mode = "full"
                mode_name = "全量"
            elif self.chinese_mode_radio.isChecked():
                mode = "chinese"
                mode_name = "仅中文"
            elif self.frame_mode_radio.isChecked():
                mode = "structure"
                mode_name = "仅结构"
            else:
                mode = "full"  # 默认全量模式
                mode_name = "全量"
            
            self.data_split_text.clear()
            self.data_split_text.append(f"🚀 开始数据拆解处理（{mode_name}模式）...")
            QApplication.processEvents()
            
            # 加载step_name映射表
            self.data_split_text.append("📋 加载step_name映射表...")
            QApplication.processEvents()
            
            step_name_mapping_path = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
                'resources', 'data', 'mapping', 'step_name_mapping.xlsx'
            )
            
            if not os.path.exists(step_name_mapping_path):
                error_msg = f"❌ step_name映射表不存在: {step_name_mapping_path}"
                self.data_split_text.append(error_msg)
                QMessageBox.critical(self, "错误", error_msg)
                return
            
            try:
                self.step_name_mapping_df = pd.read_excel(step_name_mapping_path)
                self.data_split_text.append(f"✅ step_name映射表加载成功，共 {len(self.step_name_mapping_df)} 条记录")
                QApplication.processEvents()
            except Exception as e:
                error_msg = f"❌ 加载step_name映射表失败: {str(e)}"
                self.data_split_text.append(error_msg)
                QMessageBox.critical(self, "错误", error_msg)
                return
            
            # 获取输出目录
            output_base_dir = self.config_manager.get('PATH', 'data_output_directory', './output/data')
            output_base_dir = os.path.abspath(output_base_dir)
            
            # 确保输出目录存在
            os.makedirs(output_base_dir, exist_ok=True)
            
            # 处理每个sheet
            for sheet_name, df in self.data_split_df.items():
                self.data_split_text.append(f"\n📋 处理工作表: {sheet_name}")
                QApplication.processEvents()
                
                # 检查必要的列
                required_columns = ['stage_id', 'param1', 'param2', 'param3', 'param4', 'param5']
                if not all(col in df.columns for col in required_columns):
                    self.data_split_text.append(f"⚠️ 跳过工作表 {sheet_name}：缺少必要的列")
                    continue
                
                # 根据模式过滤数据
                if mode == "chinese":
                    # 仅中文模式：只处理stage_id第6个元素为e0的数据
                    filtered_df = df[df['stage_id'].apply(lambda x: self._is_chinese_stage(str(x)))]
                    self.data_split_text.append(f"🇨🇳 仅中文模式：从 {len(df)} 条记录中筛选出 {len(filtered_df)} 条中文记录")
                else:
                    filtered_df = df
                
                # 第一步：按game_name分组
                game_groups = self._group_by_game_name(filtered_df)
                self.data_split_text.append(f"🎮 发现 {len(game_groups)} 个游戏组")
                
                # 第二步：对每个game_name组进行dialog_name分组
                for game_name, game_data in game_groups.items():
                    self.data_split_text.append(f"  📝 处理游戏: {game_name}")
                    QApplication.processEvents()
                    
                    dialog_groups = self._group_by_dialog_name(game_data)
                    self.data_split_text.append(f"    💬 发现 {len(dialog_groups)} 个对话组")
                    
                    # 第三步：生成Excel文件
                    self._generate_excel_file(game_name, dialog_groups, sheet_name, output_base_dir, mode)
            
            self.data_split_text.append(f"\n✅ 数据拆解处理完成！")
            self.data_split_text.append(f"📁 输出目录: {output_base_dir}")
            
        except Exception as e:
            error_msg = f"❌ 数据拆解处理失败: {str(e)}"
            self.data_split_text.append(error_msg)
            QMessageBox.critical(self, "处理失败", error_msg)

    def _is_chinese_stage(self, stage_id):
        """判断stage_id是否为中文stage（第6个元素为e0）"""
        parts = stage_id.split('_')
        if len(parts) >= 7:
            return parts[6] == 'e0'
        return False

    def _group_by_game_name(self, df):
        """按game_name分组数据"""
        game_groups = {}
        
        for idx, row in df.iterrows():
            stage_id = str(row['stage_id'])
            
            # 处理stage_id：用"_"split，取第0、1、3三个元素，用"-"join起来
            parts = stage_id.split('_')
            if len(parts) >= 4:
                game_name = '-'.join([parts[0], parts[1], parts[3]])
                
                if game_name not in game_groups:
                    game_groups[game_name] = []
                
                game_groups[game_name].append(row)
        
        # 转换为DataFrame
        for game_name in game_groups:
            game_groups[game_name] = pd.DataFrame(game_groups[game_name])
        
        return game_groups

    def _group_by_dialog_name(self, game_df):
        """按dialog_name分组数据"""
        dialog_groups = {}
        
        for idx, row in game_df.iterrows():
            stage_id = str(row['stage_id'])
            
            # 处理stage_id：用"_"split，取第0、1、3三个元素，用"-"join起来，
            # 取第4、5、6三个元素，把5和6直接连在一起，并跟4一起用"-"join起来
            parts = stage_id.split('_')
            if len(parts) >= 7:
                # 取第4、5、6三个元素（索引4、5、6）
                if len(parts) > 6:
                    dialog_name = '-'.join([parts[4], parts[5] + parts[6]])
                elif len(parts) > 5:
                    dialog_name = '-'.join([parts[4], parts[5]])
                else:
                    dialog_name = parts[4]
                
                if dialog_name not in dialog_groups:
                    dialog_groups[dialog_name] = []
                
                dialog_groups[dialog_name].append(row)
        
        # 转换为DataFrame
        for dialog_name in dialog_groups:
            dialog_groups[dialog_name] = pd.DataFrame(dialog_groups[dialog_name])
        
        return dialog_groups

    def _generate_excel_file(self, game_name, dialog_groups, sheet_name, output_base_dir, mode="full"):
        """生成Excel文件"""
        try:
            # 根据模式确定输出目录
            if mode == "chinese":
                output_dir = os.path.join(output_base_dir, "chinese", sheet_name)
            elif mode == "structure":
                output_dir = os.path.join(output_base_dir, "frame", sheet_name)
            else:  # full模式
                output_dir = os.path.join(output_base_dir, "split", sheet_name)
            
            os.makedirs(output_dir, exist_ok=True)
            
            # 创建Excel文件
            output_file = os.path.join(output_dir, f"{game_name}.xlsx")
            
            # 准备输出数据
            output_data = []
            
            # 获取所有dialog_name的列表
            dialog_names = list(dialog_groups.keys())
            
            if not dialog_names:
                self.data_split_text.append(f"    ⚠️ 游戏 {game_name} 没有有效的对话组")
                return
            
            # 仅结构模式：只处理第一个dialog_name
            if mode == "structure":
                dialog_names = dialog_names[:1]
                self.data_split_text.append(f"    🏗️ 仅结构模式：只处理第一个对话组 {dialog_names[0]}")
            
            # 初始化每个dialog_name下stage_name+speaker组合的出现次数统计
            dialog_occurrence_count = {}
            for dialog_name in dialog_names:
                dialog_occurrence_count[dialog_name] = {}
            
            # 处理第一个dialog_name
            first_dialog_name = dialog_names[0]
            first_dialog_data = dialog_groups[first_dialog_name]
            
            # 获取第一个dialog_name下的第一条数据，用于生成stage_id模板
            if not first_dialog_data.empty:
                first_row = first_dialog_data.iloc[0]
                template_stage_id = self._generate_template_stage_id(str(first_row['stage_id']))
                
                # 遍历第一个dialog_name下的每条数据
                for idx, row in first_dialog_data.iterrows():
                    # 检查param1-param5，如果有内容则创建行
                    for param_num in range(1, 6):
                        param_col = f'param{param_num}'
                        param_value = row[param_col]
                        
                        # 检查是否有内容（非空字符串或na）
                        if pd.notna(param_value) and str(param_value).strip() != '' and str(param_value).lower() != 'na':
                            # 获取step_name
                            step_name, warning_msg = self._get_step_name(
                                row.get('stage_type', ''), 
                                row.get('speaker', ''), 
                                first_dialog_name, 
                                dialog_occurrence_count
                            )
                            
                            # 如果有警告信息，添加到输出
                            if warning_msg:
                                self.data_split_text.append(f"    🔴 {warning_msg}")
                            
                            new_row = {
                                'stage_id': template_stage_id,
                                'stage_name': row.get('stage_type', ''),
                                'step_name': step_name,
                                'speaker': row.get('speaker', ''),
                                'param1': ''
                            }
                            
                            # 仅结构模式不添加dialog_name字段
                            if mode != "structure":
                                new_row[first_dialog_name] = str(param_value)
                            
                            output_data.append(new_row)
            
            # 处理剩余的dialog_name（只添加列，不添加行）
            # 仅结构模式跳过这部分处理
            if mode != "structure":
                for dialog_name in dialog_names[1:]:
                    dialog_data = dialog_groups[dialog_name]
                    
                    # 为现有的每一行添加这个dialog_name的列数据
                    row_index = 0
                    for idx, row in dialog_data.iterrows():
                        # 检查param1-param5，如果有内容则填入对应行
                        for param_num in range(1, 6):
                            param_col = f'param{param_num}'
                            param_value = row[param_col]
                            
                            # 检查是否有内容（非空字符串或na）
                            if pd.notna(param_value) and str(param_value).strip() != '' and str(param_value).lower() != 'na':
                                if row_index < len(output_data):
                                    output_data[row_index][dialog_name] = str(param_value)
                                else:
                                    # 如果行数不够，创建新行
                                    # 获取step_name
                                    step_name, warning_msg = self._get_step_name(
                                        row.get('stage_type', ''), 
                                        row.get('speaker', ''), 
                                        dialog_name, 
                                        dialog_occurrence_count
                                    )
                                    
                                    # 如果有警告信息，添加到输出
                                    if warning_msg:
                                        self.data_split_text.append(f"    🔴 {warning_msg}")
                                    
                                    new_row = {
                                        'stage_id': template_stage_id,
                                        'stage_name': row.get('stage_type', ''),
                                        'step_name': step_name,
                                        'speaker': row.get('speaker', ''),
                                        'param1': '',
                                        dialog_name: str(param_value)
                                    }
                                    # 为之前的dialog_name添加空值
                                    for prev_dialog in dialog_names[:dialog_names.index(dialog_name)]:
                                        if prev_dialog not in new_row:
                                            new_row[prev_dialog] = ''
                                    output_data.append(new_row)
                                row_index += 1
            
            # 创建DataFrame并保存
            if output_data:
                output_df = pd.DataFrame(output_data)
                
                # 确保所有dialog_name列都存在（仅结构模式跳过）
                if mode != "structure":
                    for dialog_name in dialog_names:
                        if dialog_name not in output_df.columns:
                            output_df[dialog_name] = ''
                
                # 保存到Excel文件
                output_df.to_excel(output_file, index=False)
                self.data_split_text.append(f"    ✅ 生成文件: {os.path.basename(output_file)}")
            else:
                self.data_split_text.append(f"    ⚠️ 游戏 {game_name} 没有有效数据，跳过文件生成")
                
        except Exception as e:
            self.data_split_text.append(f"    ❌ 生成文件失败: {str(e)}")

    def _generate_template_stage_id(self, original_stage_id):
        """生成模板stage_id"""
        parts = original_stage_id.split('_')
        if len(parts) >= 7:
            # 将第4个元素用"通用"替换，第5个元素用"a5"替换，第6个元素用"e0"替换
            parts[4] = '通用'
            parts[5] = 'a5'
            parts[6] = 'e0'
            return '_'.join(parts)
        return original_stage_id

    def _get_step_name(self, stage_type, speaker, dialog_name, dialog_occurrence_count):
        """
        根据stage_type和speaker双主键和在dialog_name下的出现次数获取step_name
        stage_type匹配step_name_mapping表的stage_name字段
        返回: (step_name, warning_message)
        """
        # 始终根据映射表计算step_name（不再直接返回stage_type）
        
        # 构建双主键标识符（仅用于计数）
        key = f"{stage_type}+{speaker}"
        
        # 初始化该dialog_name下的计数器（如果不存在）
        if key not in dialog_occurrence_count[dialog_name]:
            dialog_occurrence_count[dialog_name][key] = 0
        
        # 增加出现次数
        dialog_occurrence_count[dialog_name][key] += 1
        occurrence_count = dialog_occurrence_count[dialog_name][key]
        
        # 检查出现次数是否超过5
        if occurrence_count > 5:
            warning_msg = f"双主键组合 (stage_type='{stage_type}', speaker='{speaker}') 在dialog_name '{dialog_name}' 下出现第 {occurrence_count} 次，超过最大限制5次"
            return stage_type, warning_msg
        
        # 在映射表中查找匹配的行（stage_type匹配step_name_mapping表的stage_name字段）
        matching_rows = self.step_name_mapping_df[
            (self.step_name_mapping_df['stage_name'] == stage_type) & 
            (self.step_name_mapping_df['speaker'] == speaker)
        ]
        
        if matching_rows.empty:
            warning_msg = f"双主键组合 (stage_type='{stage_type}', speaker='{speaker}') 在step_name映射表中找不到匹配的行"
            return stage_type, warning_msg
        
        # 获取第一个匹配行
        matching_row = matching_rows.iloc[0]
        
        # 根据出现次数选择对应的step_name_x字段
        step_name_col = f'step_name_{occurrence_count}'
        step_name_value = matching_row[step_name_col]
        
        # 检查step_name_x字段的值
        if pd.isna(step_name_value) or str(step_name_value).strip() == '' or str(step_name_value).lower() == 'na':
            warning_msg = f"双主键组合 (stage_type='{stage_type}', speaker='{speaker}') 的 {step_name_col} 字段为空或na"
            return stage_type, warning_msg
        
        # 返回有效的step_name
        return str(step_name_value), None
